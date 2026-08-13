# Per-neighborhood ablation (Reviewer comment R1-23).
"""
Per-neighborhood contribution (Table 8)
=======================================
Place this file in the same folder as RMFS_main.py and the instance file.

Implementation note on the neighborhood definitions
---------------------------------------------------
variable_neighborhood_search in RMFS_main.py indexes its moves k = 1..4:

    k=1  exchange the contents of two groups, leaving yy unchanged
    k=2  move one order between two groups
    k=3  exchange the station and sequence position of two groups
    k=4  exchange one item between two waves

k=1 and k=3 are two mechanisms for the same move: both end in a different
assignment of groups to (station, sequence) pairs and leave everything else
untouched, so a configuration containing both counts one neighborhood twice.
The proposed VNS is therefore defined as ACTIVE_NEIGHBORHOODS in RMFS_main.py,
and every script reporting a headline result imports that constant, so there
is a single definition of the algorithm across all tables.

k=4 implements the item/wave exchange described in Section 4.4. It is
evaluated here as a candidate third neighborhood but is not part of the
proposed configuration.

Label mapping: N1/N2/N3 in the manuscript correspond to k=2/3/4 here, in that
order.

Usage:
  set RMFS_DATA_FILE=Data1.xlsx
  set NBHD_N_REPLICATIONS=3
  set NBHD_OUTPUT_FILE=neighborhood_ablation_Data1.xlsx
  python run_neighborhood_ablation.py
"""

import sys
_CPLEX_API = r'C:\Program Files\IBM\ILOG\CPLEX_Studio221\cplex\python\3.7\x64_win64'
# On another machine, set the CPLEX_PYTHON_API environment variable to the
# folder holding the CPLEX Python API, e.g.
#   Linux : /opt/ibm/ILOG/CPLEX_Studio221/cplex/python/3.7/x86-64_linux
#   macOS : /Applications/CPLEX_Studio221/cplex/python/3.7/x86-64_osx
import os
_CPLEX_API = os.environ.get('CPLEX_PYTHON_API', _CPLEX_API)
if os.path.isdir(_CPLEX_API):
    sys.path = [_CPLEX_API] + sys.path
import random
import numpy as np
import time
import os
import openpyxl

SEED = 42
random.seed(SEED)
np.random.seed(SEED)

N_REPLICATIONS = int(os.environ.get('NBHD_N_REPLICATIONS', '3'))
OUTPUT_FILE = os.environ.get('NBHD_OUTPUT_FILE', 'neighborhood_ablation_results.xlsx')
INSTANCE_LABEL = os.environ.get('NBHD_INSTANCE_LABEL', os.environ.get('RMFS_DATA_FILE', 'Data11.xlsx'))
FAIL_SENTINEL = 10000000

DATA_FILE = os.environ.get('RMFS_DATA_FILE', 'Data11.xlsx')
workbook = openpyxl.load_workbook(DATA_FILE)
sheet = workbook.active

def read_matrix(sheet, name):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == name:
                sr, sc = row + 1, col
                rows, cols = 0, 0
                while sheet.cell(row=sr + rows, column=sc).value is not None:
                    rows += 1
                while sheet.cell(row=sr, column=sc + cols).value is not None:
                    cols += 1
                data = []
                for r in range(sr, sr + rows):
                    row_data = [sheet.cell(row=r, column=c).value for c in range(sc, sc + cols)]
                    data.append(row_data)
                return data
    return None

def read_scalar(sheet, name):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == name:
                return sheet.cell(row=row + 1, column=col).value
    return None

Demand               = read_matrix(sheet, 'Demand')
Sum_demand           = np.sum(Demand, axis=1)
orders1              = read_scalar(sheet, 'orders1')
orders               = np.arange(orders1)
num_picking_stations = read_scalar(sheet, 'num_picking_stations')
group_capacity       = read_scalar(sheet, 'group_capacity')
max_iterations       = read_scalar(sheet, 'max_iterations')
sequences            = read_scalar(sheet, 'sequences')
items                = read_scalar(sheet, 'items')
waves                = read_scalar(sheet, 'waves')

from RMFS_main import (
    create_order_groups, assigning_order_groups,
    evaluate_order_groups, variable_neighborhood_search,
    ACTIVE_NEIGHBORHOODS,
)


# Guard against a silent instance mismatch. evaluate_order_groups and the other
# RMFS_main functions read the demand matrix, capacities and wave data from
# RMFS_main's own module-level worksheet, NOT from the arguments passed in. If
# this script and RMFS_main were pointed at different workbooks, every result
# would be computed from a mixture of two instances and no error would be
# raised. Both now read RMFS_DATA_FILE, so this check should never fire.
import RMFS_main as _rmfs_main
if os.path.abspath(_rmfs_main.DATA_FILE) != os.path.abspath(DATA_FILE):
    raise RuntimeError(
        "Instance mismatch: this script loaded %r but RMFS_main loaded %r. "
        "Set the RMFS_DATA_FILE environment variable so that both use the "
        "same workbook." % (DATA_FILE, _rmfs_main.DATA_FILE))

print(f"Instance label: {INSTANCE_LABEL}")
print(f"Instance: {orders1} orders | max_iterations={max_iterations}")
print("=" * 65)

CONFIGS = [
    ("Construction only", None),
    ("VNS: Order exchange only (N1)", (2,)),
    ("VNS: Station/sequence exchange only (N2)", (3,)),
    ("VNS: Item/wave exchange only (N3)", (4,)),
    ("Full VNS (proposed)", tuple(ACTIVE_NEIGHBORHOODS)),
]


def run_config(active_nbhd):
    t0 = time.time()
    if active_nbhd is None:
        groups = create_order_groups(Sum_demand, orders, group_capacity)
        yy, _, _ = assigning_order_groups(groups, num_picking_stations, Demand)
        cost = evaluate_order_groups(groups, yy)
    else:
        _, cost, _ = variable_neighborhood_search(
            Sum_demand, orders, group_capacity, max_iterations,
            num_picking_stations, Demand, sequences,
            active_neighborhoods=active_nbhd,
            items=items, waves=waves)
    return cost, time.time() - t0


results = {name: {'costs': [], 'times': []} for name, _ in CONFIGS}

for rep in range(N_REPLICATIONS):
    seed = SEED + rep
    print(f"\nReplication {rep+1}/{N_REPLICATIONS}  (seed={seed})")
    print("-" * 65)
    for name, active_nbhd in CONFIGS:
        random.seed(seed)
        np.random.seed(seed)
        print(f"  {name}...", end=' ', flush=True)
        cost, t = run_config(active_nbhd)
        results[name]['costs'].append(cost)
        results[name]['times'].append(t)
        status = "" if cost < FAIL_SENTINEL else "  [FAILED]"
        print(f"cost={cost:.4f}  time={t:.1f}s{status}")

print(f"\n\n{'='*80}")
print(f"PER-NEIGHBORHOOD ABLATION -- {INSTANCE_LABEL}")
print(f"{'='*80}")
print(f"{'Configuration':<55} {'Mean(ok)':>9} {'Std(ok)':>9} {'Success':>8}")
print("-" * 80)

stats = {}
baseline = None
for name, _ in CONFIGS:
    data = results[name]
    ok = [c for c in data['costs'] if c < FAIL_SENTINEL]
    mean_c = np.mean(ok) if ok else None
    std_c = np.std(ok) if ok else None
    stats[name] = {'mean_c': mean_c, 'std_c': std_c, 'n_ok': len(ok)}
    if name == "Construction only":
        baseline = mean_c

for name, _ in CONFIGS:
    s = stats[name]
    mean_str = f"{s['mean_c']:.4f}" if s['mean_c'] is not None else "N/A"
    std_str = f"{s['std_c']:.4f}" if s['std_c'] is not None else "N/A"
    succ_str = f"{s['n_ok']}/{N_REPLICATIONS}"
    print(f"{name:<55} {mean_str:>9} {std_str:>9} {succ_str:>8}")

print("-" * 80)
print("Improvement over Construction-only baseline:")
for name, _ in CONFIGS:
    if name == "Construction only":
        continue
    s = stats[name]
    if baseline and s['mean_c'] is not None:
        improv = (baseline - s['mean_c']) / baseline * 100
        print(f"  {name}: {improv:.2f}%")
print("=" * 80)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Neighborhood Ablation"
ws.append(["Instance", INSTANCE_LABEL])
ws.append(["Configuration", "Replication", "Seed", "Cost", "Time (s)", "Failed?"])
for name, _ in CONFIGS:
    for rep, (c, t) in enumerate(zip(results[name]['costs'], results[name]['times'])):
        ws.append([name, rep + 1, SEED + rep,
                   round(c, 4) if c < FAIL_SENTINEL else "FAILED",
                   round(t, 2), c >= FAIL_SENTINEL])

ws.append([])
ws.append(["Summary (means over successful replications only)"])
ws.append(["Configuration", "Mean Cost (ok)", "Std Cost (ok)", "Successes", "Total Reps",
           "Improvement vs Construction-only (%)"])
for name, _ in CONFIGS:
    s = stats[name]
    improv = None
    if name != "Construction only" and baseline and s['mean_c'] is not None:
        improv = (baseline - s['mean_c']) / baseline * 100
    ws.append([name,
               round(s['mean_c'], 4) if s['mean_c'] is not None else "N/A",
               round(s['std_c'], 4) if s['std_c'] is not None else "N/A",
               s['n_ok'], N_REPLICATIONS,
               round(improv, 2) if improv is not None else ("—" if name == "Construction only" else "N/A")])

wb.save(OUTPUT_FILE)
print(f"\n✅ Results saved to {OUTPUT_FILE}")
