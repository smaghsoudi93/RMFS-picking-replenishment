from docplex.mp.model import Model
import sys
try:
    sys.stdout.reconfigure(line_buffering=True)
except Exception:
    pass  # older Python without reconfigure(); harmless to skip
sys.path = [r'C:\Program Files\IBM\ILOG\CPLEX_Studio221\cplex\python\3.7\x64_win64'] + sys.path
import cplex
import random
import copy
import numpy as np
import time
import threading
import pandas as pd
import concurrent.futures
#import params
import openpyxl

# Per-solve CPLEX time limit (seconds) and MIP-gap tolerance.
# evaluate_order_groups() is called many times inside a single VNS run
# (observed: ~20 solves for max_iterations=5, since the "reset k to 1 on
# improvement" logic can trigger far more evaluations than max_iterations
# alone suggests), so unbounded/near-exact solves multiply badly -- a
# single run took ~886s (many solves hit the 60s cap without even
# confirming optimality). VNS only needs a solution GOOD ENOUGH to rank
# candidates against each other, not a certified global optimum at every
# step, so a modest MIP gap tolerance (accept a solution within X% of the
# best proven bound) cuts solve time drastically with little effect on
# search quality. Tune both down further for quick testing, or raise
# SOLVE_MIPGAP toward 0 (and raise the time limit) for a final, more
# precise run of the best solution found.
SOLVE_TIME_LIMIT_SECONDS = 10
SOLVE_MIPGAP = 0.05  # accept within 5% of the proven optimal bound
# Moderate thread count: forcing threads=1 earlier was meant to rule out a
# multi-thread deadlock, but if CPU usage was actually high during the long
# stalls, threads=1 was just making every solve much slower (single-thread
# CPLEX vs. multi-thread). 4 threads is a reasonable middle ground -- lower
# it back to 1 only if you see genuine 0%-CPU hangs again.
SOLVE_THREADS = 4
# The conflict refiner was essential while diagnosing the structural
# infeasibility bugs (now fixed) but is expensive on larger models and is
# no longer needed for normal runs. Set to True only if you need to debug
# a NEW kind of failure.
ENABLE_CONFLICT_REFINER = False

import os
DATA_FILE = os.environ.get('RMFS_DATA_FILE', 'Data11.xlsx')
workbook = openpyxl.load_workbook(DATA_FILE)
sheet = workbook.active


for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=row, column=col).value == 'Demand':
            demand_start_row = row + 1
            demand_start_col = col
            break

demand_rows = 0
while sheet.cell(row=demand_start_row + demand_rows, column=demand_start_col).value is not None:
    demand_rows += 1

demand_cols = 0
while sheet.cell(row=demand_start_row, column=demand_start_col + demand_cols).value is not None:
    demand_cols += 1

# Read the Demand data
Demand = []
for row in range(demand_start_row, demand_start_row + demand_rows):
    row_data = []
    for col in range(demand_start_col, demand_start_col + demand_cols):
        row_data.append(sheet.cell(row=row, column=col).value)
    Demand.append(row_data)

#for row in Demand:
#    print(row)

#xl_file = pd.read_excel("C:\Users\saman\OneDrive\Desktop\Data1.xlsx")
#Demand = xl_file.parse('Demand')

def create_order_groups(Sum_demand, orders, group_capacity):
    

    start1 = time.time()
    order_groups = []
    remaining_orders = Sum_demand.copy()
    remaining_numbers = orders.copy()
    group_count = 0
    order_groups.append([])

    while len(remaining_orders) > 0:
        sorted_pairs = sorted(zip(remaining_orders, remaining_numbers), reverse=True)
        remaining_orders = [demand for demand, number in sorted_pairs]
        remaining_numbers = [number for demand, number in sorted_pairs]
        order_groups[group_count].append(remaining_numbers[0])
        remaining_orders.remove(remaining_orders[0])
        remaining_numbers.remove(remaining_numbers[0])


        if len(remaining_orders) > 0:
            if len(order_groups[group_count]) < group_capacity:
                sorted_pairs = sorted(zip(remaining_orders, remaining_numbers))
                remaining_orders = [demand for demand, number in sorted_pairs]
                remaining_numbers = [number for demand, number in sorted_pairs]
                order_groups[group_count].append(remaining_numbers[0])
                remaining_orders.remove(remaining_orders[0])
                remaining_numbers.remove(remaining_numbers[0])

            else:
                group_count = group_count + 1
                order_groups.append([])
                sorted_pairs = sorted(zip(remaining_orders, remaining_numbers))
                remaining_orders = [demand for demand, number in sorted_pairs]
                remaining_numbers = [number for demand, number in sorted_pairs]
                order_groups[group_count].append(remaining_numbers[0])
                remaining_orders.remove(remaining_orders[0])
                remaining_numbers.remove(remaining_numbers[0])
        if len(order_groups[group_count]) >= group_capacity and len(remaining_orders) > 0:
            group_count = group_count + 1
            order_groups.append([])

    end1 = time.time()
    print("order_grouping_time:",end1 - start1)
    return order_groups

def swap_neighborhood(yy, sequences, num_picking_stations):
    new_y = yy.copy()

    # Select two random groups
    group1, group2 = random.sample(range(len(new_y)), 2)
    

    # Swap the values of ss and jj for the selected groups
    ss1, jj1 = find_ss_jj(new_y, group1, sequences, num_picking_stations)
    ss2, jj2 = find_ss_jj(new_y, group2, sequences, num_picking_stations)

    new_y[group1][ss1][jj1] = 0
    new_y[group2][ss2][jj2] = 0
    new_y[group2][ss1][jj1] = 1
    new_y[group1][ss2][jj2] = 1

    #print(new_y)
    return new_y

def find_ss_jj(y, group, sequences, num_picking_stations):
    for ss in range(num_picking_stations):
        for jj in range(sequences):
            if y[group][ss][jj] == 1:
                return ss, jj

#-----------------------------------------------------------------------------
# Initialize the problem data
#-----------------------------------------------------------------------------

#def calculate_similarity_matrix(Demand):
# Calculate similarity matrix (presence of same items between orders)
    #similarity_matrix = np.zeros((len(Demand), len(Demand)))
    #for i in range(len(Demand)):
        #for j in range(i + 1, len(Demand)):
            #similarity = sum(1 for x, y in zip(Demand[i], Demand[j]) if x > 0 and y > 0)
            #similarity_matrix[i][j] = similarity
            #similarity_matrix[j][i] = similarity
    #print("si:,",similarity_matrix)
    #return similarity_matrix

#def create_initial_order_groups(group_capacity):

    #Demand = [[10, 3, 0, 4, 5, 2, 3, 0, 0, 0, 0],
          #[4, 4, 4, 3, 0, 2, 2, 1, 3, 0, 1],
          #[3, 7, 6, 0, 1, 5, 0, 2, 1, 2, 1],
          #[0, 5, 2, 2, 3, 0, 1, 0, 2, 1, 7],
          #[1, 2, 1, 1, 1, 0, 2, 3, 5, 6, 0],
          #[5, 12, 3, 0, 4, 1, 2, 1, 0, 3, 2],
          #[6, 3, 1, 4, 0, 1, 4, 0, 5, 0, 6],
          #[2, 2, 3, 6, 7, 0, 0, 1, 2, 3, 4]]
    #order_groups = []
    #assigned_orders = set()
    #group_count = 0

    #while len(assigned_orders) < len(Demand):
        #similarity_matrix = calculate_similarity_matrix(Demand)

        # Find the indices of the maximum value
        #max_index = np.unravel_index(np.argmax(similarity_matrix), similarity_matrix.shape)
        # Get the row and column indices
        #row_index = max_index[0]
        #column_index = max_index[1]

        #order_groups.append([])
        #order_groups[group_count].append(row_index)
        #order_groups[group_count].append(column_index)

        #assigned_orders.add(row_index)
        #assigned_orders.add(column_index)

        #while len(order_groups[group_count]) < group_capacity and len(assigned_orders) < len(Demand):
            #for i in range(np.array(Demand).shape[1]):
                #Demand[row_index][i] = Demand[row_index][i] + Demand[column_index][i]
                #Demand[column_index][i] = 0

        #similarity_matrix = np.zeros((len(Demand), len(Demand)))

        #for i in range(len(Demand)):
            #for j in range(i + 1, len(Demand)):
                #similarity = sum(1 for x, y in zip(Demand[i], Demand[j]) if x > 0 and y > 0)
                #similarity_matrix[i][j] = similarity
                #similarity_matrix[j][i] = similarity
        #column_index = np.argmax(similarity_matrix[row_index])
        #order_groups[group_count].append(column_index)
        #assigned_orders.add(column_index)
        #if len(assigned_orders) < len(Demand):
            #group_count = group_count + 1
            #for i in range(np.array(Demand).shape[1]):
                #Demand[row_index][i] = 0
                #Demand[column_index][i] = 0

    #print("og:,",order_groups)
    #return order_groups


    #---------------------------------------------------- random assignment of groups to stations and sequencing them ---------------------------------

def assigning_order_groups(order_groups, num_picking_stations, Demand):
    Groups = list(range(len(order_groups)))
    Sequences = list(range(len(Demand)))
    Picking_stations = list(range(num_picking_stations))
    ss = 0
    jj = 0

    yy = np.zeros((len(order_groups), num_picking_stations, len(Demand)))
    active = np.zeros((len(order_groups), num_picking_stations, len(Demand)))

    for g in Groups:
        yy[g, ss, jj] = 1
        active[g, ss, jj] = 1
        ss = ss + 1
        if ss >= num_picking_stations:
            jj = jj + 1
            ss = 0
    #print('yy:',yy)            
    return yy, active, jj


#---------------------------------------------------- Solving sub problem of order groups and stations ---------------------------------

def _evaluate_order_groups_once(order_groups, yy, time_limit=None, mipgap=None, run_conflict_refiner=True, wave_item_swaps=None):
    _tl = SOLVE_TIME_LIMIT_SECONDS if time_limit is None else time_limit
    _gap = SOLVE_MIPGAP if mipgap is None else mipgap

    ##########################################################Read the data###########################################################################

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'matrix1':
                matrix1_start_row = row + 1
                matrix1_start_col = col
                break

    matrix1_rows = 0
    while sheet.cell(row=matrix1_start_row + matrix1_rows, column=matrix1_start_col).value is not None:
        matrix1_rows += 1

    matrix1_cols = 0
    while sheet.cell(row=matrix1_start_row, column=matrix1_start_col + matrix1_cols).value is not None:
        matrix1_cols += 1

    # Read the matrix1 data
    matrix1 = []
    for row in range(matrix1_start_row, matrix1_start_row + matrix1_rows):
        row_data1 = []
        for col in range(matrix1_start_col, matrix1_start_col + matrix1_cols):
            row_data1.append(sheet.cell(row=row, column=col).value)
        matrix1.append(row_data1)
        
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'Demand1':
                demand1_start_row = row + 1
                demand1_start_col = col
                break

    demand1_rows = 0
    while sheet.cell(row=demand1_start_row + demand1_rows, column=demand1_start_col).value is not None:
        demand1_rows += 1

    demand1_cols = 0
    while sheet.cell(row=demand1_start_row, column=demand1_start_col + demand1_cols).value is not None:
        demand1_cols += 1

    # Read the Demand data
    Demand1 = []
    for row in range(demand1_start_row, demand1_start_row + demand1_rows):
        row_data2 = []
        for col in range(demand1_start_col, demand1_start_col + demand1_cols):
            row_data2.append(sheet.cell(row=row, column=col).value)
        Demand1.append(row_data2)

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'items':
                items_row = row
                items_col = col
                break

    items = sheet.cell(row=items_row + 1, column=items_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'orders1':
                orders1_row = row
                orders1_col = col
                break

    orders1 = sheet.cell(row=orders1_row + 1, column=orders1_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'shelves':
                shelves_row = row
                shelves_col = col
                break

    shelves = sheet.cell(row=shelves_row + 1, column=shelves_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'num_picking_stations':
                num_picking_stations_row = row
                num_picking_stations_col = col
                break

    num_picking_stations = sheet.cell(row=num_picking_stations_row + 1, column=num_picking_stations_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'num_replenishment_stations':
                num_replenishment_stations_row = row
                num_replenishment_stations_col = col
                break

    num_replenishment_stations = sheet.cell(row=num_replenishment_stations_row + 1, column=num_replenishment_stations_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'waves':
                waves_row = row
                waves_col = col
                break

    waves = sheet.cell(row=waves_row + 1, column=waves_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'Pods_capacity':
                Pods_capacity_row = row
                Pods_capacity_col = col
                break

    Pods_capacity = sheet.cell(row=Pods_capacity_row + 1, column=Pods_capacity_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'group_capacity':
                group_capacity_row = row
                group_capacity_col = col
                break

    group_capacity = sheet.cell(row=group_capacity_row + 1, column=group_capacity_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'Replenishment_station_capacity':
                Replenishment_station_capacity_row = row
                Replenishment_station_capacity_col = col
                break

    Replenishment_station_capacity = sheet.cell(row=Replenishment_station_capacity_row + 1, column=Replenishment_station_capacity_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'Arriving_times':
                arriving_start_row = row + 1
                arriving_start_col = col
                break

    arriving_rows = 0
    while sheet.cell(row=arriving_start_row + arriving_rows, column=arriving_start_col).value is not None:
        arriving_rows += 1

    arriving_cols = 0
    while sheet.cell(row=arriving_start_row, column=arriving_start_col + arriving_cols).value is not None:
        arriving_cols += 1

    # Read the data
    Arriving_times = []
    for row in range(arriving_start_row, arriving_start_row + arriving_rows):
        row_data = []
        for col in range(arriving_start_col, arriving_start_col + arriving_cols):
            row_data.append(sheet.cell(row=row, column=col).value)
        Arriving_times = row_data

#--------------------------------

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'Duration':
                Duration_start_row = row + 1
                Duration_start_col = col
                break

    Duration_rows = 0
    while sheet.cell(row=Duration_start_row + Duration_rows, column=Duration_start_col).value is not None:
        Duration_rows += 1

    Duration_cols = 0
    while sheet.cell(row=Duration_start_row, column=Duration_start_col + Duration_cols).value is not None:
        Duration_cols += 1

    # Read the data
    Duration = []
    for row in range(Duration_start_row, Duration_start_row + Duration_rows):
        row_data = []
        for col in range(Duration_start_col, Duration_start_col + Duration_cols):
            row_data.append(sheet.cell(row=row, column=col).value)
        Duration = row_data


    ##################################################################################################################################################


    Groups = list(range(len(order_groups)))
    Items = list(range(items))
    Orders = list(range(orders1))
    Sequences = list(range(orders1))
    Shelves = list(range(shelves))
    Picking_stations = list(range(num_picking_stations))
    Replenishment_stations = list(range(num_replenishment_stations))
    Waves1 = list(range(waves))
    # A tight Big-M instead of an arbitrary 1,000,000. Used both for
    # time-scale bounds (ct/st vs. arrival times + durations, chained
    # across at most `orders1` sequences) and quantity-scale bounds
    # (u <= M*v, q <= M*z, capped by pod/replenishment-station capacity),
    # so take the max of both tight bounds rather than one arbitrary
    # constant. 1,000,000 was orders of magnitude looser than necessary
    # for realistic instances, which weakens the LP relaxation and forces
    # CPLEX to branch far more than needed (very likely why every solve
    # was hitting the time limit without closing the MIP gap).
    M_time = int(5 * orders1 * (max(Arriving_times) + max(Duration))) + 10
    M_qty = int(5 * max(Pods_capacity, Replenishment_station_capacity)) + 10
    M = max(M_time, M_qty)
    print(f"[diag] tightened Big-M = {M} (was 1,000,000)")
    waves_capacity = np.full((1, waves), (num_replenishment_stations*Replenishment_station_capacity))
    replenishmnet_capacity = np.full((waves, num_replenishment_stations), Replenishment_station_capacity)
    sample = np.zeros((len(Demand1), len(Demand1[0]), waves))
    capacity_of_pods = np.full((waves, shelves), Pods_capacity)
    #Arriving_times = (10, 15, 20, 25) #(7, 12, 17, 22, 27)#(5, 10, 15, 20, 25, 30) #(7, 12, 17, 22, 27) #(10, 15, 20, 25, 30) #(7, 12, 17, 22, 27)#(7, 12, 17, 22, 27, 32) #(5, 10) #(6, 12, 18, 24, 30)  #(5, 10) # 
    #Duration = (5, 5, 5, 5) #(5, 5, 5, 5, 5, 5) #(5, 5) #(6, 6, 6, 6, 6) #(5, 5) # 
    

    #items = 11 #5 #7 #5  
    #orders1 = 8 #6 #50 #20 #15  
    #shelves = 3 #10 
    #num_picking_stations = 3 #2 #6 #4  
    #num_replenishment_stations = 3 #2 #6 #4  
    #waves = 4 #2 #6 #5   
    #Pods_capacity = 45 #15 #35 #25   
    #group_capacity = 2 #7 #3 #4 
    #Replenishment_station_capacity = 25 #22 #40   
    
    

#-----------------------------------------------------------------------------
# Build the model
#-----------------------------------------------------------------------------

# Create CPO model
    mdl = Model(name='Thesis')
    print(order_groups)


    qq = np.zeros((len(order_groups),shelves, items, waves))
    zz = np.zeros((len(order_groups),shelves, waves))
    vv = np.zeros((shelves, waves, num_replenishment_stations))
    uu = np.zeros((items, shelves, waves, num_replenishment_stations))

    xx = np.zeros((len(matrix1),len(order_groups)))
    x = {}
#for group_idx, group in enumerate(order_groups):
 #   for order in group:
    for g in Groups:
        for o in Orders:
            x[o, g] = mdl.binary_var(name=f'x_{o}_{g}')

#for g in Groups:
    #for o in Orders:
        #mdl.add_constraint(x[o, g] == 0)

# Set constraints based on order_groups
    for group_idx, group in enumerate(order_groups):
        for order in group:
            xx[order, group_idx] = 1
            mdl.add_constraint(x[order, group_idx] == 1)
    #print(xx)
    #print(yy)
    y= {}
    for g in Groups:
        for s in Picking_stations:
            for j in Sequences:
                y[g, s, j] =mdl.binary_var(name=f'y_{g}_{s}_{j}')

    for g in Groups:
        for s in Picking_stations:
            for j in Sequences:
                if yy[g, s, j] == 1:
                    mdl.add_constraint(y[g, s, j] == 1)
                
    z = {(i,j,k): mdl.binary_var(name="z%d%d%d" % (i,j,k)) for i in Groups for j in Shelves for k in Waves1}
    v = {(i,j,k): mdl.binary_var(name="v%d%d%d" % (i,j,k)) for i in Shelves for j in Waves1 for k in Replenishment_stations}
    q = {(i,j,k,ii): mdl.continuous_var(name="q%d%d%d%d" % (i,j,k,ii)) for i in Groups for j in Shelves for k in Items for ii in Waves1}
    u = {(i,j,k,ii): mdl.continuous_var(name="u%d%d%d%d" % (i,j,k,ii)) for i in Items for j in Shelves for k in Waves1 for ii in Replenishment_stations}
    ct = {(i,j,k): mdl.continuous_var(name="ct%d%d%d" % (i,j,k)) for i in Groups for j in Picking_stations for k in Sequences}
    st = {(i,j,k): mdl.continuous_var(name="st%d%d%d" % (i,j,k)) for i in Groups for j in Picking_stations for k in Sequences}
    ctt = {(i): mdl.continuous_var(name="stt%d" % (i)) for i in Orders}
    #ff = {(i,j,k): mdl.continuous_var(name="ff%d%d%d" % (i,j,k)) for i in Items for j in Waves for k in Replenishment_stations}
    l = {(i,j): mdl.binary_var(name="l%d%d" % (i,j)) for i in Groups for j in Waves1}

    ff= {}
    for i in Items:
        for j in Waves1:
            for k in Replenishment_stations:
                ff[i, j, k] =mdl.continuous_var(name=f'ff_{i}_{j}_{k}')
            
    waving_items = np.zeros((waves, len(matrix1[0])))
    replenishment_items = np.zeros((waves, len(matrix1[0]),num_replenishment_stations))
    pod_items = np.zeros((len(matrix1[0]),shelves ,waves, num_replenishment_stations))
    
    # ------------------------------------------------------------------------ waving heuristic -------------------------------------------------------
    # SINGLE-HORIZON DESIGN: run the wave-assignment heuristic across ALL
    # waves first (no solving in between), then build the MIP ONCE over
    # the full Waves set and solve ONCE. This replaces the previous
    # "solve after each wave, growing the horizon" structure. That
    # incremental design repeatedly produced infeasibilities (see the
    # long chain of conflict-report analyses this file's fixes are
    # based on): groups without demand yet still needed a provisional
    # l/z/st commitment to stay consistent with already-solved earlier
    # iterations, and those provisional commitments kept conflicting
    # with other constraints (demandeq, sequencing, st/ct bounds) in
    # different ways each time. Solving once, with every group's
    # demandeq defined over the complete wave horizon from the start,
    # removes that whole class of bug. It is also more faithful to the
    # textual description of Stage 3 in Section 4.3 ("...determine each
    # wave plan and then define the problem as an MIP model..."), and
    # avoids `waves` repeated MIP solves (each with its own solve time,
    # plus conflict-refiner/retry overhead on failure).
    Waves = Waves1  # full horizon, always

    # Discover processing order once, by station sequence (same priority
    # order the original incremental scan used), then run the random
    # item-to-wave heuristic wave-by-wave (still respecting each wave's
    # own capacity), accumulating into `sample`/`waving_items` for the
    # whole horizon before any constraint is built.
    Groups1_ordered = []
    for i in Sequences:
        for g in Groups:
            for s in Picking_stations:
                if yy[g, s, i] == 1:
                    if g not in Groups1_ordered:
                        Groups1_ordered.append(g)

    for waves_number in range(waves):
        exit_loop = False
        for g in Groups1_ordered:
            for j in range(len(order_groups[g])):
                while np.sum(matrix1[order_groups[g][j]], axis=0) > 0:
                    j1 = random.randint(0, len(matrix1[0]) - 1)
                    if matrix1[order_groups[g][j]][j1] > 0:
                        if waves_capacity[0][waves_number] >= matrix1[order_groups[g][j]][j1]:
                            sample[order_groups[g][j]][j1][waves_number] = matrix1[order_groups[g][j]][j1]
                            waves_capacity[0][waves_number] = waves_capacity[0][waves_number] - matrix1[order_groups[g][j]][j1]
                            matrix1[order_groups[g][j]][j1] = 0
                        else:
                            if waves_capacity[0][waves_number] > 0:
                                sample[order_groups[g][j]][j1][waves_number] = waves_capacity[0][waves_number]
                                matrix1[order_groups[g][j]][j1] = matrix1[order_groups[g][j]][j1] - waves_capacity[0][waves_number]
                                waves_capacity[0][waves_number] = 0
                                exit_loop = True
                                break
                            else:
                                exit_loop = True
                                break
            if exit_loop:
                break

        for j in range(len(matrix1[0])):
            for i in range(len(matrix1)):
                if sample[i][j][waves_number] > 0:
                    waving_items[waves_number][j] = waving_items[waves_number][j] + sample[i][j][waves_number]


    # ---- Neighborhood 3 (Algorithm 4 / Section 4.4, Reviewer R1-23): ----
    # "Exchange one item i in wave w with one item i' in wave w'."
    # waving_items[w][i] is the aggregate quantity of item i that the
    # wave-assignment heuristic above committed to wave w; this is exactly
    # the quantity the MIP takes as fixed (see the wavetotal_i*_w*
    # constraint just below). Swapping the two cells' quantities -- rather
    # than moving/adding quantity -- exactly preserves each wave's total
    # replenishment volume, so no additional capacity check is needed here.
    # wave_item_swaps is a list of (item1, wave1, item2, wave2) tuples;
    # None/[] (the default for every existing caller) leaves waving_items
    # untouched, so no previously reported result changes.
    if wave_item_swaps:
        for (item1, wave1, item2, wave2) in wave_item_swaps:
            waving_items[wave1][item1], waving_items[wave2][item2] = \
                waving_items[wave2][item2], waving_items[wave1][item1]

    groups_demand = np.zeros((1, len(order_groups)))
    for g in range(len(order_groups)):
        for j in range(len(order_groups[g])):
            groups_demand[0][g] = groups_demand[0][g] + np.sum(matrix1[order_groups[g][j]], axis=0)

    Groups1 = list(range(len(order_groups)))
    unmet = [g for g in Groups1 if groups_demand[0][g] > 0]
    if unmet:
        print(f"[warn] total wave capacity across the full horizon was insufficient "
              f"to fully supply {len(unmet)} group(s): {unmet}. Their demand is left "
              f"optional (deferred) rather than forced, to avoid an artificial conflict; "
              f"this points to an instance-sizing issue (waves x capacity too small) "
              f"rather than a modeling bug.")

# -------------------------------------------------------- constraints ------------------------------------------------------
#Constant variables based on heuristic:
    _build_t0 = time.time()
    print(f"[diag] starting constraint construction | groups={len(Groups1)} | "
          f"shelves={shelves} | waves={waves} | sequences={len(Sequences)}")

    # Total quantity of item i physically replenished in wave w is fixed
    # to what the wave-assignment heuristic decided; the split across
    # shelves (p) and replenishment stations (r) is left as a free MIP
    # decision, bounded only by the per-station/per-shelf capacity
    # constraints below.
    for i in Items:
        for w in Waves:
            mdl.add_constraint(
                mdl.sum(u[i, p, w, r] for p in Shelves for r in Replenishment_stations)
                == waving_items[w][i],
                ctname=f"wavetotal_i{i}_w{w}")

    for g in range(len(Groups1)):
        if groups_demand[0][Groups1[g]] == 0:
            mdl.add_constraint(mdl.sum(z[Groups1[g],p,w] for p in Shelves for w in Waves) >= 1,
                               ctname=f"zforce_g{Groups1[g]}_fullydemanded")

    for w in Waves:
        for r in Replenishment_stations:
            mdl.add_constraint(mdl.sum(u[i,p,w,r] for i in Items for p in Shelves) <= Replenishment_station_capacity)

    for w in Waves:
        for p in Shelves:
            mdl.add_constraint(mdl.sum(u[i,p,w,r] for i in Items for r in Replenishment_stations) <= Pods_capacity)

    for g in range(len(Groups1)):
        if groups_demand[0][Groups1[g]] == 0:
            for i in Items:
                mdl.add_constraint(mdl.sum(Demand1[o][i]*x[o,Groups1[g]] for o in Orders) == mdl.sum(q[Groups1[g],p,i,w] for p in Shelves for w in Waves),
                                   ctname=f"demandeq_g{Groups1[g]}_i{i}_fullydemanded")

    for w in Waves:
        mdl.add_constraint(mdl.sum(u[i,p,w,r] for r in Replenishment_stations for i in Items for p in Shelves) == mdl.sum(q[Groups1[g],p,i,w] for p in Shelves for i in Items for g in range(len(Groups1))),
                           ctname=f"wavebalance_w{w}")

    # Per-item / per-shelf / per-wave link between group picks (q) and
    # the actual physical items replenished there (u). Without this, only
    # the AGGREGATE total per wave was checked (wavebalance_w*), so q for
    # item i could be satisfied "on paper" using stock that was actually
    # a different item.
    for i in Items:
        for w in Waves:
            for p in Shelves:
                mdl.add_constraint(
                    mdl.sum(q[Groups1[g], p, i, w] for g in range(len(Groups1))) <=
                    mdl.sum(u[i, p, w, r] for r in Replenishment_stations),
                    ctname=f"qleu_i{i}_p{p}_w{w}")

    for g in range(len(Groups1)):
        for w in Waves:
            mdl.add_constraint(mdl.sum(z[Groups1[g],p,w] for p in Shelves) >= l[Groups1[g],w],
                               ctname=f"zgel_g{Groups1[g]}_w{w}")

    # Groups whose demand couldn't be fully placed even across the whole
    # horizon (see `unmet` warning above) are left optional; every other
    # group commits to exactly one first wave.
    for g in range(len(Groups1)):
        group_idx = Groups1[g]
        if groups_demand[0][group_idx] == 0:
            mdl.add_constraint(mdl.sum(l[group_idx,w] for w in Waves) == 1,
                               ctname=f"lsum1_g{group_idx}_fullydemanded")
        else:
            mdl.add_constraint(mdl.sum(l[group_idx,w] for w in Waves) <= 1,
                               ctname=f"lsum_le1_g{group_idx}_deferred")

    for g in range(len(Groups1)):
        for w1 in Waves:
            mdl.add_constraint(l[Groups1[g],w1] + (mdl.sum(z[Groups1[g],p,w2] for p in Shelves for w2 in Waves if w2 < w1)/(shelves*waves)) <= 1)

    for g in range(len(Groups1)):
        for s in Picking_stations:
            for j in Sequences:
                mdl.add_constraint(st[Groups1[g],s,j] >= mdl.sum(Arriving_times[w]*l[Groups1[g],w] for w in Waves) - M*(1- y[Groups1[g],s,j]))

    for g in range(len(Groups1)):
        for s in Picking_stations:
            for j in Sequences:
                mdl.add_constraint(st[Groups1[g],s,j] <= mdl.sum(Arriving_times[w]*l[Groups1[g],w] for w in Waves) + M*(1- y[Groups1[g],s,j]))

    for g in range(len(Groups1)):
        for p in Shelves:
            for w in Waves:
                for s in Picking_stations:
                    for j in Sequences:
                        mdl.add_constraint(ct[Groups1[g],s,j] >= z[Groups1[g],p,w]*Arriving_times[w] + Duration[w] - M*(1- y[Groups1[g],s,j]))

    print(f"[diag] constraints before sequencing block: {time.time()-_build_t0:.1f}s elapsed | "
          f"about to build O(groups^2 * sequences^2) = "
          f"{len(Groups1)**2 * len(Sequences)**2} sequencing constraint checks")
    _seq_t0 = time.time()
    _seq_done = 0
    _seq_total = len(Groups1)
    for g1 in range(len(Groups1)):
        for g2 in range(len(Groups1)):
            if g1 != g2:
                for j1 in Sequences:
                    for j2 in Sequences:
                        if j2 > j1:
                            for s in Picking_stations:
                                mdl.add_constraint(st[Groups1[g2],s,j2] >= ct[Groups1[g1],s,j1] - M*(2 - y[Groups1[g2],s,j2] - y[Groups1[g1],s,j1]))
        _seq_done += 1
        if time.time() - _seq_t0 > 5:  # heartbeat if this is taking a while
            print(f"[diag] sequencing block progress: {_seq_done}/{_seq_total} groups done, "
                  f"{time.time()-_seq_t0:.1f}s elapsed so far in this block")
    print(f"[diag] sequencing block finished: {time.time()-_seq_t0:.1f}s total")

    for p in Shelves:
        for w in Waves:
            for r in Replenishment_stations:
                mdl.add_constraint(mdl.sum(u[i,p,w,r] for i in Items) >= v[p,w,r])

    for p in Shelves:
        for w in Waves:
            for r in Replenishment_stations:
                mdl.add_constraint(mdl.sum(u[i,p,w,r] for i in Items) <= M*v[p,w,r])

    for g in range(len(Groups1)):
        for p in Shelves:
            for w in Waves:
                mdl.add_constraint(mdl.sum(q[Groups1[g],p,i,w] for i in Items) >= z[Groups1[g],p,w])

    for g in range(len(Groups1)):
        for p in Shelves:
            for w in Waves:
                mdl.add_constraint(mdl.sum(q[Groups1[g],p,i,w] for i in Items) <= M*z[Groups1[g],p,w])

    # Reviewer comment 11 flagged the ">= y" lower bound (forcing ct/st to
    # be >= 1 whenever y=1) as arbitrary and unjustified ("why not 0?").
    # Removed; the upper bound (<= M*y) below already correctly ties
    # ct/st to 0 when y=0, and the tight equalities above determine their
    # real value when y=1.
    for g in range(len(Groups1)):
        for s in Picking_stations:
            for j in Sequences:
                mdl.add_constraint(ct[Groups1[g],s,j] <= M*y[Groups1[g],s,j])       

    for g in range(len(Groups1)):
        for s in Picking_stations:
            for j in Sequences:
                mdl.add_constraint(st[Groups1[g],s,j] <= M*y[Groups1[g],s,j])                                 

#Objective function
    # Reviewer comment #9: the objective is described as average ORDER
    # fulfillment time but was computed as sum(ct)/|G| (average GROUP
    # completion time) -- not equivalent unless all groups have the same
    # number of orders. Fixed: weight each group's ct by its own order
    # count and divide by the total number of orders (|O| = orders1).
    mdl.total_objective = mdl.sum(
        ct[Groups1[g], s, j] * len(order_groups[Groups1[g]])
        for g in range(len(Groups1)) for s in Picking_stations for j in Sequences
    ) / orders1
    mdl.minimize(mdl.total_objective)

    unsupplied = [Groups1[g] for g in range(len(Groups1)) if groups_demand[0][Groups1[g]] > 0]
    print(f"[diag] single-horizon solve | total_groups={len(Groups1)} | "
          f"still_unsupplied(groups_demand>0)={len(unsupplied)} -> {unsupplied}")

    # Bound every individual solve so nothing can hang indefinitely inside
    # a VNS run that calls this function many times (up to ~3*max_iterations
    # times). SOLVE_TIME_LIMIT_SECONDS is a module-level constant (see top
    # of file) so it can be tuned in one place. log_output is off by
    # default now that this runs inside a metaheuristic loop -- turn it
    # back on manually here if you need to debug a specific solve.
    print(f"[diag] TOTAL model construction time: {time.time()-_build_t0:.1f}s "
          f"(this is NOT bounded by SOLVE_TIME_LIMIT_SECONDS -- if this number "
          f"is large, the bottleneck is Python building constraints, not CPLEX solving)")
    # Scale the effective time limit with instance size: larger instances
    # (more groups) sometimes just need more time to find their FIRST
    # feasible integer solution, not because they are actually infeasible
    # -- confirmed by conflict-refiner runs reporting "0 conflicts" on a
    # "no_solution" case where the exact same grouping succeeded on a
    # later attempt. Baseline SOLVE_TIME_LIMIT_SECONDS is tuned for small
    # instances (~7 groups); scale up roughly linearly with group count
    # beyond that, capped at 5x the baseline so a single call still can't
    # hang indefinitely inside VNS.
    _n_groups = len(Groups1)
    _scale = max(1.0, _n_groups / 7.0)
    _tl_eff = min(_tl * _scale, _tl * 5)
    mdl.parameters.timelimit = _tl_eff
    mdl.parameters.threads = SOLVE_THREADS
    mdl.parameters.mip.tolerances.mipgap = _gap
    print(f"[diag] solving with time_limit={_tl_eff:.1f}s (base {_tl}s x{_scale:.2f} "
          f"for {_n_groups} groups), mipgap={_gap}")
    _solve_t0 = time.time()
    solved = mdl.solve(log_output=False)
    _solve_dt = time.time() - _solve_t0
    print(f"[timing] solve took {_solve_dt:.1f}s (limit={_tl_eff:.1f}s)")

    if solved:
        total_cost = mdl.objective_value
        print("total_cost:", total_cost)
        for i in Groups:
            for j in Picking_stations:
                for k in Sequences:
                    if ct[i, j, k].solution_value > 0:
                        print("ct:",i,j,k,ct[i,j,k].solution_value)
    else:
        if run_conflict_refiner and ENABLE_CONFLICT_REFINER:
            print("no_solution -- running conflict refiner...")
            try:
                from docplex.mp.conflict_refiner import ConflictRefiner
                cr = ConflictRefiner()
                conflicts = cr.refine_conflict(mdl, display=True)
                conflict_file = "conflict_singlehorizon.txt"
                with open(conflict_file, "w") as f:
                    f.write(f"CONFLICT REPORT - single-horizon solve, "
                            f"num_groups={len(Groups1)}, "
                            f"groups_still_unsupplied={unsupplied}\n")
                    f.write("=" * 80 + "\n")
                    for c in conflicts:
                        f.write(f"{c}\n")
                print(f"  -> conflict report saved to {conflict_file}")
                print(f"  (if this said 'conflicts: 0', the model was likely just "
                      f"too slow to find a first feasible solution in the time "
                      f"limit, not actually infeasible -- consider raising the "
                      f"time limit for this instance size)")
            except Exception as e:
                print("  conflict refiner failed:", e)
        else:
            print("no_solution -- skipping conflict refiner on this attempt "
                  "(will run it on the final retry if still failing) -- "
                  "retrying with a fresh random draw and/or more time first, "
                  "since most failures at this stage are just timeouts, not "
                  "true infeasibility.")
        total_cost = 10000000

    return total_cost


def evaluate_order_groups(order_groups, yy, max_retries=3, time_limit=None, mipgap=None, wave_item_swaps=None):
    """
    Thin retry wrapper around _evaluate_order_groups_once.

    Most "no_solution" results turn out to be CPLEX simply not finding a
    first feasible integer solution within the time limit on larger
    instances (confirmed by conflict-refiner runs reporting 0 conflicts,
    and by the exact same grouping succeeding on a later attempt) rather
    than genuine infeasibility. Retrying with a fresh random draw is a
    cheap, low-risk mitigation either way: each retry is an independent
    shot at a feasible physical layout / an easier-to-solve instance of
    the same model. The conflict refiner (expensive on large models) is
    only run on the FINAL retry, so it doesn't waste time on attempts
    that are likely just unlucky timeouts.
    """
    last_cost = 10000000
    for attempt in range(1, max_retries + 1):
        cost = _evaluate_order_groups_once(
            order_groups, yy, time_limit=time_limit, mipgap=mipgap,
            run_conflict_refiner=(attempt == max_retries),
            wave_item_swaps=wave_item_swaps)
        if cost < 10000000:
            if attempt > 1:
                print(f"[retry] succeeded on attempt {attempt}/{max_retries}")
            return cost
        last_cost = cost
        print(f"[retry] attempt {attempt}/{max_retries} hit infeasibility, "
              f"retrying with a fresh random draw..." if attempt < max_retries
              else f"[retry] all {max_retries} attempts failed, giving up.")
    return last_cost


def _evaluate_order_groups_sequential_once(order_groups, yy, time_limit=None, mipgap=None, run_conflict_refiner=True):
    _tl = SOLVE_TIME_LIMIT_SECONDS if time_limit is None else time_limit
    _gap = SOLVE_MIPGAP if mipgap is None else mipgap

    ##########################################################Read the data###########################################################################

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'matrix1':
                matrix1_start_row = row + 1
                matrix1_start_col = col
                break

    matrix1_rows = 0
    while sheet.cell(row=matrix1_start_row + matrix1_rows, column=matrix1_start_col).value is not None:
        matrix1_rows += 1

    matrix1_cols = 0
    while sheet.cell(row=matrix1_start_row, column=matrix1_start_col + matrix1_cols).value is not None:
        matrix1_cols += 1

    # Read the matrix1 data
    matrix1 = []
    for row in range(matrix1_start_row, matrix1_start_row + matrix1_rows):
        row_data1 = []
        for col in range(matrix1_start_col, matrix1_start_col + matrix1_cols):
            row_data1.append(sheet.cell(row=row, column=col).value)
        matrix1.append(row_data1)
        
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'Demand1':
                demand1_start_row = row + 1
                demand1_start_col = col
                break

    demand1_rows = 0
    while sheet.cell(row=demand1_start_row + demand1_rows, column=demand1_start_col).value is not None:
        demand1_rows += 1

    demand1_cols = 0
    while sheet.cell(row=demand1_start_row, column=demand1_start_col + demand1_cols).value is not None:
        demand1_cols += 1

    # Read the Demand data
    Demand1 = []
    for row in range(demand1_start_row, demand1_start_row + demand1_rows):
        row_data2 = []
        for col in range(demand1_start_col, demand1_start_col + demand1_cols):
            row_data2.append(sheet.cell(row=row, column=col).value)
        Demand1.append(row_data2)

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'items':
                items_row = row
                items_col = col
                break

    items = sheet.cell(row=items_row + 1, column=items_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'orders1':
                orders1_row = row
                orders1_col = col
                break

    orders1 = sheet.cell(row=orders1_row + 1, column=orders1_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'shelves':
                shelves_row = row
                shelves_col = col
                break

    shelves = sheet.cell(row=shelves_row + 1, column=shelves_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'num_picking_stations':
                num_picking_stations_row = row
                num_picking_stations_col = col
                break

    num_picking_stations = sheet.cell(row=num_picking_stations_row + 1, column=num_picking_stations_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'num_replenishment_stations':
                num_replenishment_stations_row = row
                num_replenishment_stations_col = col
                break

    num_replenishment_stations = sheet.cell(row=num_replenishment_stations_row + 1, column=num_replenishment_stations_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'waves':
                waves_row = row
                waves_col = col
                break

    waves = sheet.cell(row=waves_row + 1, column=waves_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'Pods_capacity':
                Pods_capacity_row = row
                Pods_capacity_col = col
                break

    Pods_capacity = sheet.cell(row=Pods_capacity_row + 1, column=Pods_capacity_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'group_capacity':
                group_capacity_row = row
                group_capacity_col = col
                break

    group_capacity = sheet.cell(row=group_capacity_row + 1, column=group_capacity_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'Replenishment_station_capacity':
                Replenishment_station_capacity_row = row
                Replenishment_station_capacity_col = col
                break

    Replenishment_station_capacity = sheet.cell(row=Replenishment_station_capacity_row + 1, column=Replenishment_station_capacity_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'Arriving_times':
                arriving_start_row = row + 1
                arriving_start_col = col
                break

    arriving_rows = 0
    while sheet.cell(row=arriving_start_row + arriving_rows, column=arriving_start_col).value is not None:
        arriving_rows += 1

    arriving_cols = 0
    while sheet.cell(row=arriving_start_row, column=arriving_start_col + arriving_cols).value is not None:
        arriving_cols += 1

    # Read the data
    Arriving_times = []
    for row in range(arriving_start_row, arriving_start_row + arriving_rows):
        row_data = []
        for col in range(arriving_start_col, arriving_start_col + arriving_cols):
            row_data.append(sheet.cell(row=row, column=col).value)
        Arriving_times = row_data

#--------------------------------

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'Duration':
                Duration_start_row = row + 1
                Duration_start_col = col
                break

    Duration_rows = 0
    while sheet.cell(row=Duration_start_row + Duration_rows, column=Duration_start_col).value is not None:
        Duration_rows += 1

    Duration_cols = 0
    while sheet.cell(row=Duration_start_row, column=Duration_start_col + Duration_cols).value is not None:
        Duration_cols += 1

    # Read the data
    Duration = []
    for row in range(Duration_start_row, Duration_start_row + Duration_rows):
        row_data = []
        for col in range(Duration_start_col, Duration_start_col + Duration_cols):
            row_data.append(sheet.cell(row=row, column=col).value)
        Duration = row_data


    ##################################################################################################################################################


    Groups = list(range(len(order_groups)))
    Items = list(range(items))
    Orders = list(range(orders1))
    Sequences = list(range(orders1))
    Shelves = list(range(shelves))
    Picking_stations = list(range(num_picking_stations))
    Replenishment_stations = list(range(num_replenishment_stations))
    Waves1 = list(range(waves))
    # A tight Big-M instead of an arbitrary 1,000,000. Used both for
    # time-scale bounds (ct/st vs. arrival times + durations, chained
    # across at most `orders1` sequences) and quantity-scale bounds
    # (u <= M*v, q <= M*z, capped by pod/replenishment-station capacity),
    # so take the max of both tight bounds rather than one arbitrary
    # constant. 1,000,000 was orders of magnitude looser than necessary
    # for realistic instances, which weakens the LP relaxation and forces
    # CPLEX to branch far more than needed (very likely why every solve
    # was hitting the time limit without closing the MIP gap).
    M_time = int(5 * orders1 * (max(Arriving_times) + max(Duration))) + 10
    M_qty = int(5 * max(Pods_capacity, Replenishment_station_capacity)) + 10
    M = max(M_time, M_qty)
    print(f"[diag] tightened Big-M = {M} (was 1,000,000)")
    waves_capacity = np.full((1, waves), (num_replenishment_stations*Replenishment_station_capacity))
    replenishmnet_capacity = np.full((waves, num_replenishment_stations), Replenishment_station_capacity)
    sample = np.zeros((len(Demand1), len(Demand1[0]), waves))
    capacity_of_pods = np.full((waves, shelves), Pods_capacity)
    #Arriving_times = (10, 15, 20, 25) #(7, 12, 17, 22, 27)#(5, 10, 15, 20, 25, 30) #(7, 12, 17, 22, 27) #(10, 15, 20, 25, 30) #(7, 12, 17, 22, 27)#(7, 12, 17, 22, 27, 32) #(5, 10) #(6, 12, 18, 24, 30)  #(5, 10) # 
    #Duration = (5, 5, 5, 5) #(5, 5, 5, 5, 5, 5) #(5, 5) #(6, 6, 6, 6, 6) #(5, 5) # 
    

    #items = 11 #5 #7 #5  
    #orders1 = 8 #6 #50 #20 #15  
    #shelves = 3 #10 
    #num_picking_stations = 3 #2 #6 #4  
    #num_replenishment_stations = 3 #2 #6 #4  
    #waves = 4 #2 #6 #5   
    #Pods_capacity = 45 #15 #35 #25   
    #group_capacity = 2 #7 #3 #4 
    #Replenishment_station_capacity = 25 #22 #40   
    
    

#-----------------------------------------------------------------------------
# Build the model
#-----------------------------------------------------------------------------

# Create CPO model
    mdl = Model(name='Thesis')
    print(order_groups)


    qq = np.zeros((len(order_groups),shelves, items, waves))
    zz = np.zeros((len(order_groups),shelves, waves))
    vv = np.zeros((shelves, waves, num_replenishment_stations))
    uu = np.zeros((items, shelves, waves, num_replenishment_stations))

    xx = np.zeros((len(matrix1),len(order_groups)))
    x = {}
#for group_idx, group in enumerate(order_groups):
 #   for order in group:
    for g in Groups:
        for o in Orders:
            x[o, g] = mdl.binary_var(name=f'x_{o}_{g}')

#for g in Groups:
    #for o in Orders:
        #mdl.add_constraint(x[o, g] == 0)

# Set constraints based on order_groups
    for group_idx, group in enumerate(order_groups):
        for order in group:
            xx[order, group_idx] = 1
            mdl.add_constraint(x[order, group_idx] == 1)
    #print(xx)
    #print(yy)
    y= {}
    for g in Groups:
        for s in Picking_stations:
            for j in Sequences:
                y[g, s, j] =mdl.binary_var(name=f'y_{g}_{s}_{j}')

    for g in Groups:
        for s in Picking_stations:
            for j in Sequences:
                if yy[g, s, j] == 1:
                    mdl.add_constraint(y[g, s, j] == 1)
                
    z = {(i,j,k): mdl.binary_var(name="z%d%d%d" % (i,j,k)) for i in Groups for j in Shelves for k in Waves1}
    v = {(i,j,k): mdl.binary_var(name="v%d%d%d" % (i,j,k)) for i in Shelves for j in Waves1 for k in Replenishment_stations}
    q = {(i,j,k,ii): mdl.continuous_var(name="q%d%d%d%d" % (i,j,k,ii)) for i in Groups for j in Shelves for k in Items for ii in Waves1}
    u = {(i,j,k,ii): mdl.continuous_var(name="u%d%d%d%d" % (i,j,k,ii)) for i in Items for j in Shelves for k in Waves1 for ii in Replenishment_stations}
    ct = {(i,j,k): mdl.continuous_var(name="ct%d%d%d" % (i,j,k)) for i in Groups for j in Picking_stations for k in Sequences}
    st = {(i,j,k): mdl.continuous_var(name="st%d%d%d" % (i,j,k)) for i in Groups for j in Picking_stations for k in Sequences}
    ctt = {(i): mdl.continuous_var(name="stt%d" % (i)) for i in Orders}
    #ff = {(i,j,k): mdl.continuous_var(name="ff%d%d%d" % (i,j,k)) for i in Items for j in Waves for k in Replenishment_stations}
    l = {(i,j): mdl.binary_var(name="l%d%d" % (i,j)) for i in Groups for j in Waves1}

    ff= {}
    for i in Items:
        for j in Waves1:
            for k in Replenishment_stations:
                ff[i, j, k] =mdl.continuous_var(name=f'ff_{i}_{j}_{k}')
            
    waving_items = np.zeros((waves, len(matrix1[0])))
    replenishment_items = np.zeros((waves, len(matrix1[0]),num_replenishment_stations))
    pod_items = np.zeros((len(matrix1[0]),shelves ,waves, num_replenishment_stations))
    
    # ------------------------------------------------------------------------ waving heuristic -------------------------------------------------------
    # SINGLE-HORIZON DESIGN: run the wave-assignment heuristic across ALL
    # waves first (no solving in between), then build the MIP ONCE over
    # the full Waves set and solve ONCE. This replaces the previous
    # "solve after each wave, growing the horizon" structure. That
    # incremental design repeatedly produced infeasibilities (see the
    # long chain of conflict-report analyses this file's fixes are
    # based on): groups without demand yet still needed a provisional
    # l/z/st commitment to stay consistent with already-solved earlier
    # iterations, and those provisional commitments kept conflicting
    # with other constraints (demandeq, sequencing, st/ct bounds) in
    # different ways each time. Solving once, with every group's
    # demandeq defined over the complete wave horizon from the start,
    # removes that whole class of bug. It is also more faithful to the
    # textual description of Stage 3 in Section 4.3 ("...determine each
    # wave plan and then define the problem as an MIP model..."), and
    # avoids `waves` repeated MIP solves (each with its own solve time,
    # plus conflict-refiner/retry overhead on failure).
    Waves = Waves1  # full horizon, always

    # Discover processing order once, by station sequence (same priority
    # order the original incremental scan used), then run the random
    # item-to-wave heuristic wave-by-wave (still respecting each wave's
    # own capacity), accumulating into `sample`/`waving_items` for the
    # whole horizon before any constraint is built.
    Groups1_ordered = []
    for i in Sequences:
        for g in Groups:
            for s in Picking_stations:
                if yy[g, s, i] == 1:
                    if g not in Groups1_ordered:
                        Groups1_ordered.append(g)

    for waves_number in range(waves):
        exit_loop = False
        for g in Groups1_ordered:
            for j in range(len(order_groups[g])):
                while np.sum(matrix1[order_groups[g][j]], axis=0) > 0:
                    j1 = random.randint(0, len(matrix1[0]) - 1)
                    if matrix1[order_groups[g][j]][j1] > 0:
                        if waves_capacity[0][waves_number] >= matrix1[order_groups[g][j]][j1]:
                            sample[order_groups[g][j]][j1][waves_number] = matrix1[order_groups[g][j]][j1]
                            waves_capacity[0][waves_number] = waves_capacity[0][waves_number] - matrix1[order_groups[g][j]][j1]
                            matrix1[order_groups[g][j]][j1] = 0
                        else:
                            if waves_capacity[0][waves_number] > 0:
                                sample[order_groups[g][j]][j1][waves_number] = waves_capacity[0][waves_number]
                                matrix1[order_groups[g][j]][j1] = matrix1[order_groups[g][j]][j1] - waves_capacity[0][waves_number]
                                waves_capacity[0][waves_number] = 0
                                exit_loop = True
                                break
                            else:
                                exit_loop = True
                                break
            if exit_loop:
                break

        for j in range(len(matrix1[0])):
            for i in range(len(matrix1)):
                if sample[i][j][waves_number] > 0:
                    waving_items[waves_number][j] = waving_items[waves_number][j] + sample[i][j][waves_number]

    groups_demand = np.zeros((1, len(order_groups)))
    for g in range(len(order_groups)):
        for j in range(len(order_groups[g])):
            groups_demand[0][g] = groups_demand[0][g] + np.sum(matrix1[order_groups[g][j]], axis=0)

    Groups1 = list(range(len(order_groups)))
    unmet = [g for g in Groups1 if groups_demand[0][g] > 0]
    if unmet:
        print(f"[warn] total wave capacity across the full horizon was insufficient "
              f"to fully supply {len(unmet)} group(s): {unmet}. Their demand is left "
              f"optional (deferred) rather than forced, to avoid an artificial conflict; "
              f"this points to an instance-sizing issue (waves x capacity too small) "
              f"rather than a modeling bug.")

# -------------------------------------------------------- constraints ------------------------------------------------------
#Constant variables based on heuristic:
    _build_t0 = time.time()
    print(f"[diag] starting constraint construction | groups={len(Groups1)} | "
          f"shelves={shelves} | waves={waves} | sequences={len(Sequences)}")

    # === SEQUENTIAL / RANDOM REPLENISHMENT BASELINE (Section 5.3) ===
    # Unlike the integrated model (where u -- the shelf/station split of
    # replenished items -- is left FREE for the solver to optimize), this
    # baseline fixes replenishment quantities via a naive, demand-blind
    # random/greedy split representing conventional, uncoordinated
    # warehouse practice: replenishment decided independently of what
    # picking actually needs.
    pod_items_fixed = np.zeros((len(matrix1[0]), shelves, waves, num_replenishment_stations))
    for w in range(waves):
        item_order = list(range(len(matrix1[0])))
        random.shuffle(item_order)
        p_cursor, r_cursor = 0, 0
        p_cap = [Pods_capacity] * shelves
        r_cap = [Replenishment_station_capacity] * num_replenishment_stations
        for it in item_order:
            remaining = waving_items[w][it]
            while remaining > 1e-9:
                while r_cursor < num_replenishment_stations - 1 and r_cap[r_cursor] <= 1e-9:
                    r_cursor += 1
                take_r = min(remaining, max(r_cap[r_cursor], 0))
                if take_r <= 1e-9:
                    break
                r_cap[r_cursor] -= take_r
                chunk = take_r
                while chunk > 1e-9:
                    while p_cursor < shelves - 1 and p_cap[p_cursor] <= 1e-9:
                        p_cursor += 1
                    take_p = min(chunk, max(p_cap[p_cursor], 0))
                    if take_p <= 1e-9:
                        break
                    pod_items_fixed[it][p_cursor][w][r_cursor] += take_p
                    p_cap[p_cursor] -= take_p
                    chunk -= take_p
                remaining -= take_r

    for i in Items:
        for w in Waves:
            for p in Shelves:
                for r in Replenishment_stations:
                    mdl.add_constraint(u[i, p, w, r] == pod_items_fixed[i][p][w][r],
                                       ctname=f"seq_ufixed_i{i}_p{p}_w{w}_r{r}")

    for g in range(len(Groups1)):
        if groups_demand[0][Groups1[g]] == 0:
            mdl.add_constraint(mdl.sum(z[Groups1[g],p,w] for p in Shelves for w in Waves) >= 1,
                               ctname=f"zforce_g{Groups1[g]}_fullydemanded")

    for w in Waves:
        for r in Replenishment_stations:
            mdl.add_constraint(mdl.sum(u[i,p,w,r] for i in Items for p in Shelves) <= Replenishment_station_capacity)

    for w in Waves:
        for p in Shelves:
            mdl.add_constraint(mdl.sum(u[i,p,w,r] for i in Items for r in Replenishment_stations) <= Pods_capacity)

    for g in range(len(Groups1)):
        if groups_demand[0][Groups1[g]] == 0:
            for i in Items:
                mdl.add_constraint(mdl.sum(Demand1[o][i]*x[o,Groups1[g]] for o in Orders) == mdl.sum(q[Groups1[g],p,i,w] for p in Shelves for w in Waves),
                                   ctname=f"demandeq_g{Groups1[g]}_i{i}_fullydemanded")

    for w in Waves:
        mdl.add_constraint(mdl.sum(u[i,p,w,r] for r in Replenishment_stations for i in Items for p in Shelves) == mdl.sum(q[Groups1[g],p,i,w] for p in Shelves for i in Items for g in range(len(Groups1))),
                           ctname=f"wavebalance_w{w}")

    # Per-item / per-shelf / per-wave link between group picks (q) and
    # the actual physical items replenished there (u). Without this, only
    # the AGGREGATE total per wave was checked (wavebalance_w*), so q for
    # item i could be satisfied "on paper" using stock that was actually
    # a different item.
    for i in Items:
        for w in Waves:
            for p in Shelves:
                mdl.add_constraint(
                    mdl.sum(q[Groups1[g], p, i, w] for g in range(len(Groups1))) <=
                    mdl.sum(u[i, p, w, r] for r in Replenishment_stations),
                    ctname=f"qleu_i{i}_p{p}_w{w}")

    for g in range(len(Groups1)):
        for w in Waves:
            mdl.add_constraint(mdl.sum(z[Groups1[g],p,w] for p in Shelves) >= l[Groups1[g],w],
                               ctname=f"zgel_g{Groups1[g]}_w{w}")

    # Groups whose demand couldn't be fully placed even across the whole
    # horizon (see `unmet` warning above) are left optional; every other
    # group commits to exactly one first wave.
    for g in range(len(Groups1)):
        group_idx = Groups1[g]
        if groups_demand[0][group_idx] == 0:
            mdl.add_constraint(mdl.sum(l[group_idx,w] for w in Waves) == 1,
                               ctname=f"lsum1_g{group_idx}_fullydemanded")
        else:
            mdl.add_constraint(mdl.sum(l[group_idx,w] for w in Waves) <= 1,
                               ctname=f"lsum_le1_g{group_idx}_deferred")

    for g in range(len(Groups1)):
        for w1 in Waves:
            mdl.add_constraint(l[Groups1[g],w1] + (mdl.sum(z[Groups1[g],p,w2] for p in Shelves for w2 in Waves if w2 < w1)/(shelves*waves)) <= 1)

    for g in range(len(Groups1)):
        for s in Picking_stations:
            for j in Sequences:
                mdl.add_constraint(st[Groups1[g],s,j] >= mdl.sum(Arriving_times[w]*l[Groups1[g],w] for w in Waves) - M*(1- y[Groups1[g],s,j]))

    for g in range(len(Groups1)):
        for s in Picking_stations:
            for j in Sequences:
                mdl.add_constraint(st[Groups1[g],s,j] <= mdl.sum(Arriving_times[w]*l[Groups1[g],w] for w in Waves) + M*(1- y[Groups1[g],s,j]))

    for g in range(len(Groups1)):
        for p in Shelves:
            for w in Waves:
                for s in Picking_stations:
                    for j in Sequences:
                        mdl.add_constraint(ct[Groups1[g],s,j] >= z[Groups1[g],p,w]*Arriving_times[w] + Duration[w] - M*(1- y[Groups1[g],s,j]))

    print(f"[diag] constraints before sequencing block: {time.time()-_build_t0:.1f}s elapsed | "
          f"about to build O(groups^2 * sequences^2) = "
          f"{len(Groups1)**2 * len(Sequences)**2} sequencing constraint checks")
    _seq_t0 = time.time()
    _seq_done = 0
    _seq_total = len(Groups1)
    for g1 in range(len(Groups1)):
        for g2 in range(len(Groups1)):
            if g1 != g2:
                for j1 in Sequences:
                    for j2 in Sequences:
                        if j2 > j1:
                            for s in Picking_stations:
                                mdl.add_constraint(st[Groups1[g2],s,j2] >= ct[Groups1[g1],s,j1] - M*(2 - y[Groups1[g2],s,j2] - y[Groups1[g1],s,j1]))
        _seq_done += 1
        if time.time() - _seq_t0 > 5:  # heartbeat if this is taking a while
            print(f"[diag] sequencing block progress: {_seq_done}/{_seq_total} groups done, "
                  f"{time.time()-_seq_t0:.1f}s elapsed so far in this block")
    print(f"[diag] sequencing block finished: {time.time()-_seq_t0:.1f}s total")

    for p in Shelves:
        for w in Waves:
            for r in Replenishment_stations:
                mdl.add_constraint(mdl.sum(u[i,p,w,r] for i in Items) >= v[p,w,r])

    for p in Shelves:
        for w in Waves:
            for r in Replenishment_stations:
                mdl.add_constraint(mdl.sum(u[i,p,w,r] for i in Items) <= M*v[p,w,r])

    for g in range(len(Groups1)):
        for p in Shelves:
            for w in Waves:
                mdl.add_constraint(mdl.sum(q[Groups1[g],p,i,w] for i in Items) >= z[Groups1[g],p,w])

    for g in range(len(Groups1)):
        for p in Shelves:
            for w in Waves:
                mdl.add_constraint(mdl.sum(q[Groups1[g],p,i,w] for i in Items) <= M*z[Groups1[g],p,w])

    # Reviewer comment 11 flagged the ">= y" lower bound (forcing ct/st to
    # be >= 1 whenever y=1) as arbitrary and unjustified ("why not 0?").
    # Removed; the upper bound (<= M*y) below already correctly ties
    # ct/st to 0 when y=0, and the tight equalities above determine their
    # real value when y=1.
    for g in range(len(Groups1)):
        for s in Picking_stations:
            for j in Sequences:
                mdl.add_constraint(ct[Groups1[g],s,j] <= M*y[Groups1[g],s,j])       

    for g in range(len(Groups1)):
        for s in Picking_stations:
            for j in Sequences:
                mdl.add_constraint(st[Groups1[g],s,j] <= M*y[Groups1[g],s,j])                                 

#Objective function
    # Reviewer comment #9: the objective is described as average ORDER
    # fulfillment time but was computed as sum(ct)/|G| (average GROUP
    # completion time) -- not equivalent unless all groups have the same
    # number of orders. Fixed: weight each group's ct by its own order
    # count and divide by the total number of orders (|O| = orders1).
    mdl.total_objective = mdl.sum(
        ct[Groups1[g], s, j] * len(order_groups[Groups1[g]])
        for g in range(len(Groups1)) for s in Picking_stations for j in Sequences
    ) / orders1
    mdl.minimize(mdl.total_objective)

    unsupplied = [Groups1[g] for g in range(len(Groups1)) if groups_demand[0][Groups1[g]] > 0]
    print(f"[diag] sequential-baseline solve | total_groups={len(Groups1)} | "
          f"still_unsupplied(groups_demand>0)={len(unsupplied)} -> {unsupplied}")

    # Bound every individual solve so nothing can hang indefinitely inside
    # a VNS run that calls this function many times (up to ~3*max_iterations
    # times). SOLVE_TIME_LIMIT_SECONDS is a module-level constant (see top
    # of file) so it can be tuned in one place. log_output is off by
    # default now that this runs inside a metaheuristic loop -- turn it
    # back on manually here if you need to debug a specific solve.
    print(f"[diag] TOTAL model construction time: {time.time()-_build_t0:.1f}s "
          f"(this is NOT bounded by SOLVE_TIME_LIMIT_SECONDS -- if this number "
          f"is large, the bottleneck is Python building constraints, not CPLEX solving)")
    # Scale the effective time limit with instance size: larger instances
    # (more groups) sometimes just need more time to find their FIRST
    # feasible integer solution, not because they are actually infeasible
    # -- confirmed by conflict-refiner runs reporting "0 conflicts" on a
    # "no_solution" case where the exact same grouping succeeded on a
    # later attempt. Baseline SOLVE_TIME_LIMIT_SECONDS is tuned for small
    # instances (~7 groups); scale up roughly linearly with group count
    # beyond that, capped at 5x the baseline so a single call still can't
    # hang indefinitely inside VNS.
    _n_groups = len(Groups1)
    _scale = max(1.0, _n_groups / 7.0)
    _tl_eff = min(_tl * _scale, _tl * 5)
    mdl.parameters.timelimit = _tl_eff
    mdl.parameters.threads = SOLVE_THREADS
    mdl.parameters.mip.tolerances.mipgap = _gap
    print(f"[diag] solving with time_limit={_tl_eff:.1f}s (base {_tl}s x{_scale:.2f} "
          f"for {_n_groups} groups), mipgap={_gap}")
    _solve_t0 = time.time()
    solved = mdl.solve(log_output=False)
    _solve_dt = time.time() - _solve_t0
    print(f"[timing] solve took {_solve_dt:.1f}s (limit={_tl_eff:.1f}s)")

    if solved:
        total_cost = mdl.objective_value
        print("total_cost:", total_cost)
        for i in Groups:
            for j in Picking_stations:
                for k in Sequences:
                    if ct[i, j, k].solution_value > 0:
                        print("ct:",i,j,k,ct[i,j,k].solution_value)
    else:
        if run_conflict_refiner and ENABLE_CONFLICT_REFINER:
            print("no_solution -- running conflict refiner...")
            try:
                from docplex.mp.conflict_refiner import ConflictRefiner
                cr = ConflictRefiner()
                conflicts = cr.refine_conflict(mdl, display=True)
                conflict_file = "conflict_sequential.txt"
                with open(conflict_file, "w") as f:
                    f.write(f"CONFLICT REPORT - sequential-baseline solve, "
                            f"num_groups={len(Groups1)}, "
                            f"groups_still_unsupplied={unsupplied}\n")
                    f.write("=" * 80 + "\n")
                    for c in conflicts:
                        f.write(f"{c}\n")
                print(f"  -> conflict report saved to {conflict_file}")
                print(f"  (if this said 'conflicts: 0', the model was likely just "
                      f"too slow to find a first feasible solution in the time "
                      f"limit, not actually infeasible -- consider raising the "
                      f"time limit for this instance size)")
            except Exception as e:
                print("  conflict refiner failed:", e)
        else:
            print("no_solution -- skipping conflict refiner on this attempt "
                  "(will run it on the final retry if still failing) -- "
                  "retrying with a fresh random draw and/or more time first, "
                  "since most failures at this stage are just timeouts, not "
                  "true infeasibility.")
        total_cost = 10000000

    return total_cost


def evaluate_order_groups_sequential(order_groups, yy, max_retries=5, time_limit=None, mipgap=None):
    """
    Sequential / random-replenishment baseline (Section 5.3), for the
    Fig. 4 integrated-vs-sequential comparison. Same retry philosophy as
    evaluate_order_groups(), but with more retries by default: fixing u
    via a demand-blind random split is EXPECTED to fail more often than
    the integrated model (that is the whole point -- it represents
    uncoordinated, less efficient real-world practice), so a higher
    retry budget avoids conflating "this happens to be an unlucky draw"
    with "the scheme is fundamentally worse", while still letting a
    persistently-failing case show through as a real reliability
    difference worth reporting.
    """
    last_cost = 10000000
    for attempt in range(1, max_retries + 1):
        cost = _evaluate_order_groups_sequential_once(
            order_groups, yy, time_limit=time_limit, mipgap=mipgap,
            run_conflict_refiner=(attempt == max_retries))
        if cost < 10000000:
            if attempt > 1:
                print(f"[retry] sequential baseline succeeded on attempt {attempt}/{max_retries}")
            return cost
        last_cost = cost
        print(f"[retry] sequential baseline attempt {attempt}/{max_retries} hit infeasibility, "
              f"retrying with a fresh random draw..." if attempt < max_retries
              else f"[retry] sequential baseline: all {max_retries} attempts failed, giving up.")
    return last_cost


def _evaluate_order_groups_ruleBased_once(order_groups, yy, time_limit=None, mipgap=None, run_conflict_refiner=True):
    _tl = SOLVE_TIME_LIMIT_SECONDS if time_limit is None else time_limit
    _gap = SOLVE_MIPGAP if mipgap is None else mipgap

    ##########################################################Read the data###########################################################################

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'matrix1':
                matrix1_start_row = row + 1
                matrix1_start_col = col
                break

    matrix1_rows = 0
    while sheet.cell(row=matrix1_start_row + matrix1_rows, column=matrix1_start_col).value is not None:
        matrix1_rows += 1

    matrix1_cols = 0
    while sheet.cell(row=matrix1_start_row, column=matrix1_start_col + matrix1_cols).value is not None:
        matrix1_cols += 1

    # Read the matrix1 data
    matrix1 = []
    for row in range(matrix1_start_row, matrix1_start_row + matrix1_rows):
        row_data1 = []
        for col in range(matrix1_start_col, matrix1_start_col + matrix1_cols):
            row_data1.append(sheet.cell(row=row, column=col).value)
        matrix1.append(row_data1)
        
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'Demand1':
                demand1_start_row = row + 1
                demand1_start_col = col
                break

    demand1_rows = 0
    while sheet.cell(row=demand1_start_row + demand1_rows, column=demand1_start_col).value is not None:
        demand1_rows += 1

    demand1_cols = 0
    while sheet.cell(row=demand1_start_row, column=demand1_start_col + demand1_cols).value is not None:
        demand1_cols += 1

    # Read the Demand data
    Demand1 = []
    for row in range(demand1_start_row, demand1_start_row + demand1_rows):
        row_data2 = []
        for col in range(demand1_start_col, demand1_start_col + demand1_cols):
            row_data2.append(sheet.cell(row=row, column=col).value)
        Demand1.append(row_data2)

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'items':
                items_row = row
                items_col = col
                break

    items = sheet.cell(row=items_row + 1, column=items_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'orders1':
                orders1_row = row
                orders1_col = col
                break

    orders1 = sheet.cell(row=orders1_row + 1, column=orders1_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'shelves':
                shelves_row = row
                shelves_col = col
                break

    shelves = sheet.cell(row=shelves_row + 1, column=shelves_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'num_picking_stations':
                num_picking_stations_row = row
                num_picking_stations_col = col
                break

    num_picking_stations = sheet.cell(row=num_picking_stations_row + 1, column=num_picking_stations_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'num_replenishment_stations':
                num_replenishment_stations_row = row
                num_replenishment_stations_col = col
                break

    num_replenishment_stations = sheet.cell(row=num_replenishment_stations_row + 1, column=num_replenishment_stations_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'waves':
                waves_row = row
                waves_col = col
                break

    waves = sheet.cell(row=waves_row + 1, column=waves_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'Pods_capacity':
                Pods_capacity_row = row
                Pods_capacity_col = col
                break

    Pods_capacity = sheet.cell(row=Pods_capacity_row + 1, column=Pods_capacity_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'group_capacity':
                group_capacity_row = row
                group_capacity_col = col
                break

    group_capacity = sheet.cell(row=group_capacity_row + 1, column=group_capacity_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'Replenishment_station_capacity':
                Replenishment_station_capacity_row = row
                Replenishment_station_capacity_col = col
                break

    Replenishment_station_capacity = sheet.cell(row=Replenishment_station_capacity_row + 1, column=Replenishment_station_capacity_col).value

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'Arriving_times':
                arriving_start_row = row + 1
                arriving_start_col = col
                break

    arriving_rows = 0
    while sheet.cell(row=arriving_start_row + arriving_rows, column=arriving_start_col).value is not None:
        arriving_rows += 1

    arriving_cols = 0
    while sheet.cell(row=arriving_start_row, column=arriving_start_col + arriving_cols).value is not None:
        arriving_cols += 1

    # Read the data
    Arriving_times = []
    for row in range(arriving_start_row, arriving_start_row + arriving_rows):
        row_data = []
        for col in range(arriving_start_col, arriving_start_col + arriving_cols):
            row_data.append(sheet.cell(row=row, column=col).value)
        Arriving_times = row_data

#--------------------------------

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == 'Duration':
                Duration_start_row = row + 1
                Duration_start_col = col
                break

    Duration_rows = 0
    while sheet.cell(row=Duration_start_row + Duration_rows, column=Duration_start_col).value is not None:
        Duration_rows += 1

    Duration_cols = 0
    while sheet.cell(row=Duration_start_row, column=Duration_start_col + Duration_cols).value is not None:
        Duration_cols += 1

    # Read the data
    Duration = []
    for row in range(Duration_start_row, Duration_start_row + Duration_rows):
        row_data = []
        for col in range(Duration_start_col, Duration_start_col + Duration_cols):
            row_data.append(sheet.cell(row=row, column=col).value)
        Duration = row_data


    ##################################################################################################################################################


    Groups = list(range(len(order_groups)))
    Items = list(range(items))
    Orders = list(range(orders1))
    Sequences = list(range(orders1))
    Shelves = list(range(shelves))
    Picking_stations = list(range(num_picking_stations))
    Replenishment_stations = list(range(num_replenishment_stations))
    Waves1 = list(range(waves))
    # A tight Big-M instead of an arbitrary 1,000,000. Used both for
    # time-scale bounds (ct/st vs. arrival times + durations, chained
    # across at most `orders1` sequences) and quantity-scale bounds
    # (u <= M*v, q <= M*z, capped by pod/replenishment-station capacity),
    # so take the max of both tight bounds rather than one arbitrary
    # constant. 1,000,000 was orders of magnitude looser than necessary
    # for realistic instances, which weakens the LP relaxation and forces
    # CPLEX to branch far more than needed (very likely why every solve
    # was hitting the time limit without closing the MIP gap).
    M_time = int(5 * orders1 * (max(Arriving_times) + max(Duration))) + 10
    M_qty = int(5 * max(Pods_capacity, Replenishment_station_capacity)) + 10
    M = max(M_time, M_qty)
    print(f"[diag] tightened Big-M = {M} (was 1,000,000)")
    waves_capacity = np.full((1, waves), (num_replenishment_stations*Replenishment_station_capacity))
    replenishmnet_capacity = np.full((waves, num_replenishment_stations), Replenishment_station_capacity)
    sample = np.zeros((len(Demand1), len(Demand1[0]), waves))
    capacity_of_pods = np.full((waves, shelves), Pods_capacity)
    #Arriving_times = (10, 15, 20, 25) #(7, 12, 17, 22, 27)#(5, 10, 15, 20, 25, 30) #(7, 12, 17, 22, 27) #(10, 15, 20, 25, 30) #(7, 12, 17, 22, 27)#(7, 12, 17, 22, 27, 32) #(5, 10) #(6, 12, 18, 24, 30)  #(5, 10) # 
    #Duration = (5, 5, 5, 5) #(5, 5, 5, 5, 5, 5) #(5, 5) #(6, 6, 6, 6, 6) #(5, 5) # 
    

    #items = 11 #5 #7 #5  
    #orders1 = 8 #6 #50 #20 #15  
    #shelves = 3 #10 
    #num_picking_stations = 3 #2 #6 #4  
    #num_replenishment_stations = 3 #2 #6 #4  
    #waves = 4 #2 #6 #5   
    #Pods_capacity = 45 #15 #35 #25   
    #group_capacity = 2 #7 #3 #4 
    #Replenishment_station_capacity = 25 #22 #40   
    
    

#-----------------------------------------------------------------------------
# Build the model
#-----------------------------------------------------------------------------

# Create CPO model
    mdl = Model(name='Thesis')
    print(order_groups)


    qq = np.zeros((len(order_groups),shelves, items, waves))
    zz = np.zeros((len(order_groups),shelves, waves))
    vv = np.zeros((shelves, waves, num_replenishment_stations))
    uu = np.zeros((items, shelves, waves, num_replenishment_stations))

    xx = np.zeros((len(matrix1),len(order_groups)))
    x = {}
#for group_idx, group in enumerate(order_groups):
 #   for order in group:
    for g in Groups:
        for o in Orders:
            x[o, g] = mdl.binary_var(name=f'x_{o}_{g}')

#for g in Groups:
    #for o in Orders:
        #mdl.add_constraint(x[o, g] == 0)

# Set constraints based on order_groups
    for group_idx, group in enumerate(order_groups):
        for order in group:
            xx[order, group_idx] = 1
            mdl.add_constraint(x[order, group_idx] == 1)
    #print(xx)
    #print(yy)
    y= {}
    for g in Groups:
        for s in Picking_stations:
            for j in Sequences:
                y[g, s, j] =mdl.binary_var(name=f'y_{g}_{s}_{j}')

    for g in Groups:
        for s in Picking_stations:
            for j in Sequences:
                if yy[g, s, j] == 1:
                    mdl.add_constraint(y[g, s, j] == 1)
                
    z = {(i,j,k): mdl.binary_var(name="z%d%d%d" % (i,j,k)) for i in Groups for j in Shelves for k in Waves1}
    v = {(i,j,k): mdl.binary_var(name="v%d%d%d" % (i,j,k)) for i in Shelves for j in Waves1 for k in Replenishment_stations}
    q = {(i,j,k,ii): mdl.continuous_var(name="q%d%d%d%d" % (i,j,k,ii)) for i in Groups for j in Shelves for k in Items for ii in Waves1}
    u = {(i,j,k,ii): mdl.continuous_var(name="u%d%d%d%d" % (i,j,k,ii)) for i in Items for j in Shelves for k in Waves1 for ii in Replenishment_stations}
    ct = {(i,j,k): mdl.continuous_var(name="ct%d%d%d" % (i,j,k)) for i in Groups for j in Picking_stations for k in Sequences}
    st = {(i,j,k): mdl.continuous_var(name="st%d%d%d" % (i,j,k)) for i in Groups for j in Picking_stations for k in Sequences}
    ctt = {(i): mdl.continuous_var(name="stt%d" % (i)) for i in Orders}
    #ff = {(i,j,k): mdl.continuous_var(name="ff%d%d%d" % (i,j,k)) for i in Items for j in Waves for k in Replenishment_stations}
    l = {(i,j): mdl.binary_var(name="l%d%d" % (i,j)) for i in Groups for j in Waves1}

    ff= {}
    for i in Items:
        for j in Waves1:
            for k in Replenishment_stations:
                ff[i, j, k] =mdl.continuous_var(name=f'ff_{i}_{j}_{k}')
            
    waving_items = np.zeros((waves, len(matrix1[0])))
    replenishment_items = np.zeros((waves, len(matrix1[0]),num_replenishment_stations))
    pod_items = np.zeros((len(matrix1[0]),shelves ,waves, num_replenishment_stations))
    
    # ------------------------------------------------------------------------ waving heuristic -------------------------------------------------------
    # SINGLE-HORIZON DESIGN: run the wave-assignment heuristic across ALL
    # waves first (no solving in between), then build the MIP ONCE over
    # the full Waves set and solve ONCE. This replaces the previous
    # "solve after each wave, growing the horizon" structure. That
    # incremental design repeatedly produced infeasibilities (see the
    # long chain of conflict-report analyses this file's fixes are
    # based on): groups without demand yet still needed a provisional
    # l/z/st commitment to stay consistent with already-solved earlier
    # iterations, and those provisional commitments kept conflicting
    # with other constraints (demandeq, sequencing, st/ct bounds) in
    # different ways each time. Solving once, with every group's
    # demandeq defined over the complete wave horizon from the start,
    # removes that whole class of bug. It is also more faithful to the
    # textual description of Stage 3 in Section 4.3 ("...determine each
    # wave plan and then define the problem as an MIP model..."), and
    # avoids `waves` repeated MIP solves (each with its own solve time,
    # plus conflict-refiner/retry overhead on failure).
    Waves = Waves1  # full horizon, always

    # Discover processing order once, by station sequence (same priority
    # order the original incremental scan used), then run the random
    # item-to-wave heuristic wave-by-wave (still respecting each wave's
    # own capacity), accumulating into `sample`/`waving_items` for the
    # whole horizon before any constraint is built.
    Groups1_ordered = []
    for i in Sequences:
        for g in Groups:
            for s in Picking_stations:
                if yy[g, s, i] == 1:
                    if g not in Groups1_ordered:
                        Groups1_ordered.append(g)

    for waves_number in range(waves):
        exit_loop = False
        for g in Groups1_ordered:
            for j in range(len(order_groups[g])):
                while np.sum(matrix1[order_groups[g][j]], axis=0) > 0:
                    # RULE-BASED (Reviewer comment R2-14): instead of a random
                    # pick, deterministically choose the item with the largest
                    # remaining demand for this order (greedy by item demand),
                    # one of the alternatives the reviewer suggested.
                    j1 = int(np.argmax(matrix1[order_groups[g][j]]))
                    if matrix1[order_groups[g][j]][j1] > 0:
                        if waves_capacity[0][waves_number] >= matrix1[order_groups[g][j]][j1]:
                            sample[order_groups[g][j]][j1][waves_number] = matrix1[order_groups[g][j]][j1]
                            waves_capacity[0][waves_number] = waves_capacity[0][waves_number] - matrix1[order_groups[g][j]][j1]
                            matrix1[order_groups[g][j]][j1] = 0
                        else:
                            if waves_capacity[0][waves_number] > 0:
                                sample[order_groups[g][j]][j1][waves_number] = waves_capacity[0][waves_number]
                                matrix1[order_groups[g][j]][j1] = matrix1[order_groups[g][j]][j1] - waves_capacity[0][waves_number]
                                waves_capacity[0][waves_number] = 0
                                exit_loop = True
                                break
                            else:
                                exit_loop = True
                                break
            if exit_loop:
                break

        for j in range(len(matrix1[0])):
            for i in range(len(matrix1)):
                if sample[i][j][waves_number] > 0:
                    waving_items[waves_number][j] = waving_items[waves_number][j] + sample[i][j][waves_number]

    groups_demand = np.zeros((1, len(order_groups)))
    for g in range(len(order_groups)):
        for j in range(len(order_groups[g])):
            groups_demand[0][g] = groups_demand[0][g] + np.sum(matrix1[order_groups[g][j]], axis=0)

    Groups1 = list(range(len(order_groups)))
    unmet = [g for g in Groups1 if groups_demand[0][g] > 0]
    if unmet:
        print(f"[warn] total wave capacity across the full horizon was insufficient "
              f"to fully supply {len(unmet)} group(s): {unmet}. Their demand is left "
              f"optional (deferred) rather than forced, to avoid an artificial conflict; "
              f"this points to an instance-sizing issue (waves x capacity too small) "
              f"rather than a modeling bug.")

# -------------------------------------------------------- constraints ------------------------------------------------------
#Constant variables based on heuristic:
    _build_t0 = time.time()
    print(f"[diag] starting constraint construction | groups={len(Groups1)} | "
          f"shelves={shelves} | waves={waves} | sequences={len(Sequences)}")

    # Total quantity of item i physically replenished in wave w is fixed
    # to what the wave-assignment heuristic decided; the split across
    # shelves (p) and replenishment stations (r) is left as a free MIP
    # decision, bounded only by the per-station/per-shelf capacity
    # constraints below.
    for i in Items:
        for w in Waves:
            mdl.add_constraint(
                mdl.sum(u[i, p, w, r] for p in Shelves for r in Replenishment_stations)
                == waving_items[w][i],
                ctname=f"wavetotal_i{i}_w{w}")

    for g in range(len(Groups1)):
        if groups_demand[0][Groups1[g]] == 0:
            mdl.add_constraint(mdl.sum(z[Groups1[g],p,w] for p in Shelves for w in Waves) >= 1,
                               ctname=f"zforce_g{Groups1[g]}_fullydemanded")

    for w in Waves:
        for r in Replenishment_stations:
            mdl.add_constraint(mdl.sum(u[i,p,w,r] for i in Items for p in Shelves) <= Replenishment_station_capacity)

    for w in Waves:
        for p in Shelves:
            mdl.add_constraint(mdl.sum(u[i,p,w,r] for i in Items for r in Replenishment_stations) <= Pods_capacity)

    for g in range(len(Groups1)):
        if groups_demand[0][Groups1[g]] == 0:
            for i in Items:
                mdl.add_constraint(mdl.sum(Demand1[o][i]*x[o,Groups1[g]] for o in Orders) == mdl.sum(q[Groups1[g],p,i,w] for p in Shelves for w in Waves),
                                   ctname=f"demandeq_g{Groups1[g]}_i{i}_fullydemanded")

    for w in Waves:
        mdl.add_constraint(mdl.sum(u[i,p,w,r] for r in Replenishment_stations for i in Items for p in Shelves) == mdl.sum(q[Groups1[g],p,i,w] for p in Shelves for i in Items for g in range(len(Groups1))),
                           ctname=f"wavebalance_w{w}")

    # Per-item / per-shelf / per-wave link between group picks (q) and
    # the actual physical items replenished there (u). Without this, only
    # the AGGREGATE total per wave was checked (wavebalance_w*), so q for
    # item i could be satisfied "on paper" using stock that was actually
    # a different item.
    for i in Items:
        for w in Waves:
            for p in Shelves:
                mdl.add_constraint(
                    mdl.sum(q[Groups1[g], p, i, w] for g in range(len(Groups1))) <=
                    mdl.sum(u[i, p, w, r] for r in Replenishment_stations),
                    ctname=f"qleu_i{i}_p{p}_w{w}")

    for g in range(len(Groups1)):
        for w in Waves:
            mdl.add_constraint(mdl.sum(z[Groups1[g],p,w] for p in Shelves) >= l[Groups1[g],w],
                               ctname=f"zgel_g{Groups1[g]}_w{w}")

    # Groups whose demand couldn't be fully placed even across the whole
    # horizon (see `unmet` warning above) are left optional; every other
    # group commits to exactly one first wave.
    for g in range(len(Groups1)):
        group_idx = Groups1[g]
        if groups_demand[0][group_idx] == 0:
            mdl.add_constraint(mdl.sum(l[group_idx,w] for w in Waves) == 1,
                               ctname=f"lsum1_g{group_idx}_fullydemanded")
        else:
            mdl.add_constraint(mdl.sum(l[group_idx,w] for w in Waves) <= 1,
                               ctname=f"lsum_le1_g{group_idx}_deferred")

    for g in range(len(Groups1)):
        for w1 in Waves:
            mdl.add_constraint(l[Groups1[g],w1] + (mdl.sum(z[Groups1[g],p,w2] for p in Shelves for w2 in Waves if w2 < w1)/(shelves*waves)) <= 1)

    for g in range(len(Groups1)):
        for s in Picking_stations:
            for j in Sequences:
                mdl.add_constraint(st[Groups1[g],s,j] >= mdl.sum(Arriving_times[w]*l[Groups1[g],w] for w in Waves) - M*(1- y[Groups1[g],s,j]))

    for g in range(len(Groups1)):
        for s in Picking_stations:
            for j in Sequences:
                mdl.add_constraint(st[Groups1[g],s,j] <= mdl.sum(Arriving_times[w]*l[Groups1[g],w] for w in Waves) + M*(1- y[Groups1[g],s,j]))

    for g in range(len(Groups1)):
        for p in Shelves:
            for w in Waves:
                for s in Picking_stations:
                    for j in Sequences:
                        mdl.add_constraint(ct[Groups1[g],s,j] >= z[Groups1[g],p,w]*Arriving_times[w] + Duration[w] - M*(1- y[Groups1[g],s,j]))

    print(f"[diag] constraints before sequencing block: {time.time()-_build_t0:.1f}s elapsed | "
          f"about to build O(groups^2 * sequences^2) = "
          f"{len(Groups1)**2 * len(Sequences)**2} sequencing constraint checks")
    _seq_t0 = time.time()
    _seq_done = 0
    _seq_total = len(Groups1)
    for g1 in range(len(Groups1)):
        for g2 in range(len(Groups1)):
            if g1 != g2:
                for j1 in Sequences:
                    for j2 in Sequences:
                        if j2 > j1:
                            for s in Picking_stations:
                                mdl.add_constraint(st[Groups1[g2],s,j2] >= ct[Groups1[g1],s,j1] - M*(2 - y[Groups1[g2],s,j2] - y[Groups1[g1],s,j1]))
        _seq_done += 1
        if time.time() - _seq_t0 > 5:  # heartbeat if this is taking a while
            print(f"[diag] sequencing block progress: {_seq_done}/{_seq_total} groups done, "
                  f"{time.time()-_seq_t0:.1f}s elapsed so far in this block")
    print(f"[diag] sequencing block finished: {time.time()-_seq_t0:.1f}s total")

    for p in Shelves:
        for w in Waves:
            for r in Replenishment_stations:
                mdl.add_constraint(mdl.sum(u[i,p,w,r] for i in Items) >= v[p,w,r])

    for p in Shelves:
        for w in Waves:
            for r in Replenishment_stations:
                mdl.add_constraint(mdl.sum(u[i,p,w,r] for i in Items) <= M*v[p,w,r])

    for g in range(len(Groups1)):
        for p in Shelves:
            for w in Waves:
                mdl.add_constraint(mdl.sum(q[Groups1[g],p,i,w] for i in Items) >= z[Groups1[g],p,w])

    for g in range(len(Groups1)):
        for p in Shelves:
            for w in Waves:
                mdl.add_constraint(mdl.sum(q[Groups1[g],p,i,w] for i in Items) <= M*z[Groups1[g],p,w])

    # Reviewer comment 11 flagged the ">= y" lower bound (forcing ct/st to
    # be >= 1 whenever y=1) as arbitrary and unjustified ("why not 0?").
    # Removed; the upper bound (<= M*y) below already correctly ties
    # ct/st to 0 when y=0, and the tight equalities above determine their
    # real value when y=1.
    for g in range(len(Groups1)):
        for s in Picking_stations:
            for j in Sequences:
                mdl.add_constraint(ct[Groups1[g],s,j] <= M*y[Groups1[g],s,j])       

    for g in range(len(Groups1)):
        for s in Picking_stations:
            for j in Sequences:
                mdl.add_constraint(st[Groups1[g],s,j] <= M*y[Groups1[g],s,j])                                 

#Objective function
    # Reviewer comment #9: the objective is described as average ORDER
    # fulfillment time but was computed as sum(ct)/|G| (average GROUP
    # completion time) -- not equivalent unless all groups have the same
    # number of orders. Fixed: weight each group's ct by its own order
    # count and divide by the total number of orders (|O| = orders1).
    mdl.total_objective = mdl.sum(
        ct[Groups1[g], s, j] * len(order_groups[Groups1[g]])
        for g in range(len(Groups1)) for s in Picking_stations for j in Sequences
    ) / orders1
    mdl.minimize(mdl.total_objective)

    unsupplied = [Groups1[g] for g in range(len(Groups1)) if groups_demand[0][Groups1[g]] > 0]
    print(f"[diag] rule-based solve | total_groups={len(Groups1)} | "
          f"still_unsupplied(groups_demand>0)={len(unsupplied)} -> {unsupplied}")

    # Bound every individual solve so nothing can hang indefinitely inside
    # a VNS run that calls this function many times (up to ~3*max_iterations
    # times). SOLVE_TIME_LIMIT_SECONDS is a module-level constant (see top
    # of file) so it can be tuned in one place. log_output is off by
    # default now that this runs inside a metaheuristic loop -- turn it
    # back on manually here if you need to debug a specific solve.
    print(f"[diag] TOTAL model construction time: {time.time()-_build_t0:.1f}s "
          f"(this is NOT bounded by SOLVE_TIME_LIMIT_SECONDS -- if this number "
          f"is large, the bottleneck is Python building constraints, not CPLEX solving)")
    # Scale the effective time limit with instance size: larger instances
    # (more groups) sometimes just need more time to find their FIRST
    # feasible integer solution, not because they are actually infeasible
    # -- confirmed by conflict-refiner runs reporting "0 conflicts" on a
    # "no_solution" case where the exact same grouping succeeded on a
    # later attempt. Baseline SOLVE_TIME_LIMIT_SECONDS is tuned for small
    # instances (~7 groups); scale up roughly linearly with group count
    # beyond that, capped at 5x the baseline so a single call still can't
    # hang indefinitely inside VNS.
    _n_groups = len(Groups1)
    _scale = max(1.0, _n_groups / 7.0)
    _tl_eff = min(_tl * _scale, _tl * 5)
    mdl.parameters.timelimit = _tl_eff
    mdl.parameters.threads = SOLVE_THREADS
    mdl.parameters.mip.tolerances.mipgap = _gap
    print(f"[diag] solving with time_limit={_tl_eff:.1f}s (base {_tl}s x{_scale:.2f} "
          f"for {_n_groups} groups), mipgap={_gap}")
    _solve_t0 = time.time()
    solved = mdl.solve(log_output=False)
    _solve_dt = time.time() - _solve_t0
    print(f"[timing] solve took {_solve_dt:.1f}s (limit={_tl_eff:.1f}s)")

    if solved:
        total_cost = mdl.objective_value
        print("total_cost:", total_cost)
        for i in Groups:
            for j in Picking_stations:
                for k in Sequences:
                    if ct[i, j, k].solution_value > 0:
                        print("ct:",i,j,k,ct[i,j,k].solution_value)
    else:
        if run_conflict_refiner and ENABLE_CONFLICT_REFINER:
            print("no_solution -- running conflict refiner...")
            try:
                from docplex.mp.conflict_refiner import ConflictRefiner
                cr = ConflictRefiner()
                conflicts = cr.refine_conflict(mdl, display=True)
                conflict_file = "conflict_rulebased.txt"
                with open(conflict_file, "w") as f:
                    f.write(f"CONFLICT REPORT - rule-based solve, "
                            f"num_groups={len(Groups1)}, "
                            f"groups_still_unsupplied={unsupplied}\n")
                    f.write("=" * 80 + "\n")
                    for c in conflicts:
                        f.write(f"{c}\n")
                print(f"  -> conflict report saved to {conflict_file}")
                print(f"  (if this said 'conflicts: 0', the model was likely just "
                      f"too slow to find a first feasible solution in the time "
                      f"limit, not actually infeasible -- consider raising the "
                      f"time limit for this instance size)")
            except Exception as e:
                print("  conflict refiner failed:", e)
        else:
            print("no_solution -- skipping conflict refiner on this attempt "
                  "(will run it on the final retry if still failing) -- "
                  "retrying with a fresh random draw and/or more time first, "
                  "since most failures at this stage are just timeouts, not "
                  "true infeasibility.")
        total_cost = 10000000

    return total_cost


def evaluate_order_groups_rulebased(order_groups, yy, max_retries=3, time_limit=None, mipgap=None):
    """
    Rule-based (deterministic, greedy-by-item-demand) Stage 3, for the
    Reviewer R2-14 comparison against the random wave-assignment
    heuristic. Since the rule is deterministic, repeated calls with the
    same inputs give the same result (no variance) -- still wrapped with
    retries only as a safety net against the same rare timeout issue
    seen in the random version.
    """
    last_cost = 10000000
    for attempt in range(1, max_retries + 1):
        cost = _evaluate_order_groups_ruleBased_once(
            order_groups, yy, time_limit=time_limit, mipgap=mipgap,
            run_conflict_refiner=(attempt == max_retries))
        if cost < 10000000:
            if attempt > 1:
                print(f"[retry] rule-based Stage 3 succeeded on attempt {attempt}/{max_retries}")
            return cost
        last_cost = cost
        print(f"[retry] rule-based Stage 3 attempt {attempt}/{max_retries} hit infeasibility, "
              f"retrying..." if attempt < max_retries
              else f"[retry] rule-based Stage 3: all {max_retries} attempts failed, giving up.")
    return last_cost


def final_polish(order_groups, yy, time_limit=300, mipgap=0.001, max_retries=3):
    """
    Re-solve the BEST solution found by VNS with much tighter settings,
    for the number that actually gets reported in the paper/tables.

    During the VNS search itself, every candidate is solved fast-and-loose
    (SOLVE_TIME_LIMIT_SECONDS / SOLVE_MIPGAP, currently 10s / 5%) -- that
    is fine for RANKING candidates against each other (observed: the best
    solution found was stable across gap settings from 2% to 5%), but not
    precise enough to defend as the final reported objective value. Call
    this once, on the single best (order_groups, yy) pair VNS returns,
    with a long time limit and a near-zero gap, to get a trustworthy final
    number.

    Usage:
        best_groups, best_cost_search, best_yy = variable_neighborhood_search(...)
        final_cost = final_polish(best_groups, best_yy)
    """
    print(f"[final_polish] re-solving best solution with time_limit={time_limit}s, "
          f"mipgap={mipgap} (this is the number to report)")
    return evaluate_order_groups(order_groups, yy, max_retries=max_retries,
                                  time_limit=time_limit, mipgap=mipgap)

    
def perform_neighborhood_operation_1(order_groups):
    # Randomly select two different groups
    group_indices = np.random.choice(len(order_groups), size=2, replace=False)
    group1_index = group_indices[0]
    group2_index = group_indices[1]

    group1 = order_groups[group1_index]
    group2 = order_groups[group2_index]

    # Randomly select two different orders within the selected groups
    order_indices1 = np.random.choice(len(group1), size=1, replace=False)
    order_indices2 = np.random.choice(len(group2), size=1, replace=False)
    order1_index = order_indices1[0]
    order2_index = order_indices2[0]

    # Swap the selected orders between the groups
    group1[order1_index], group2[order2_index] = group2[order2_index], group1[order1_index]

    return order_groups

def perform_neighborhood_operation_3(order_groups):
    # Randomly select two different groups
    group_indices = np.random.choice(len(order_groups), size=2, replace=False)
    group1_index = group_indices[0]
    group2_index = group_indices[1]

    group1 = order_groups[group1_index]
    group2 = order_groups[group2_index]

    # Swap the selected groups
    order_groups[group1_index] = group2
    order_groups[group2_index] = group1

    return order_groups

def perform_neighborhood_operation_wave_item(current_swaps, items, waves):
    """
    Neighborhood 3 (Algorithm 4 / Section 4.4, Reviewer R1-23): "Exchange
    one item i in wave w with one item i' in wave w'."

    Unlike neighborhoods 1 and 2, this one does not perturb order_groups
    or yy -- it perturbs the aggregate item-to-wave replenishment
    quantities (waving_items) that the wave-assignment heuristic inside
    _evaluate_order_groups_once commits to before the MIP is solved (see
    that function's wave_item_swaps argument). current_swaps is the list
    of swaps already accepted for the current VNS solution; this proposes
    ONE additional swap on top, mirroring how neighborhoods 1/2 perturb a
    COPY of the current solution at each shaking step.
    """
    new_swaps = list(current_swaps)
    item1, item2 = np.random.choice(items, size=2, replace=False)
    wave1, wave2 = np.random.choice(waves, size=2, replace=False)
    new_swaps.append((int(item1), int(wave1), int(item2), int(wave2)))
    return new_swaps

#def variable_neighborhood_search_iteration(iteration, Sum_demand, orders, group_capacity, num_picking_stations, Demand, sequences):
def variable_neighborhood_search(Sum_demand, orders, group_capacity, max_iterations,
                                   num_picking_stations, Demand, sequences,
                                   active_neighborhoods=(1, 2, 3),
                                   items=None, waves=None):
    """
    Basic VNS (Mladenovic & Hansen, 1997): at each step, shake the current
    solution in neighborhood k, evaluate the resulting candidate, and
    either accept it (replacing the current/best solution and resetting
    k=1) or move on to the next, larger neighborhood (k=k+1). This
    replaces the previous implementation, which had three serious bugs:
      1. All three neighborhoods were applied unconditionally every
         iteration instead of one at a time (not real VNS).
      2. `current_order_groups = best_order_groups.copy()` was a SHALLOW
         copy, so mutating the "current" solution silently corrupted the
         "best" solution too (they shared the same inner group lists).
      3. `swap_neighborhood(...)`'s return value was discarded, so
         neighborhood 3 (station/sequence exchange) never actually did
         anything.
    k_max is reported explicitly (Reviewer comment R2-16 / R1-16).

    active_neighborhoods: subset of {1,2,3,4} to actually use. k=1,2,3 are
    UNCHANGED from every previously reported result (Tables 3, 6, 7,
    Figure 4, etc. all used the default active_neighborhoods=(1,2,3) and
    remain valid/reproducible as-is -- nothing about them changes here).

    k=4 is NEW: "exchange one item i in wave w with one item i' in wave
    w'" (Algorithm 4 / Section 4.4's third stated neighborhood, which was
    missing from the code -- see perform_neighborhood_operation_wave_item).
    It is opt-in only (never included unless explicitly requested in
    active_neighborhoods), so it cannot affect any previously reported
    number. items/waves (instance-level counts) must be supplied when 4
    is included.

    NOTE on k=1 vs k=3 (pointed out while building the R1-23 per-
    neighborhood ablation): perform_neighborhood_operation_3 (k=1, swaps
    two entire groups' contents while yy stays fixed) and
    swap_neighborhood (k=3, swaps two groups' (station, wave) slot in yy
    while group contents stay fixed) produce the SAME resulting
    group-to-(station,wave) assignment -- they are two different
    mechanisms for the same move. For the R1-23 per-neighborhood
    ablation, use k=2 (order exchange), k=3 (station/sequence exchange),
    and k=4 (item/wave exchange) as the three genuinely distinct
    neighborhoods; k=1 is redundant with k=3 and is kept only so the
    original active_neighborhoods=(1,2,3) default is 100% unchanged.
    """
    active_neighborhoods = list(active_neighborhoods)
    if 4 in active_neighborhoods:
        assert items is not None and waves is not None, \
            "items and waves must be supplied when neighborhood 4 (item/wave exchange) is active"

    best_order_groups = create_order_groups(Sum_demand, orders, group_capacity)
    best_yy, active, jj = assigning_order_groups(best_order_groups, num_picking_stations, Demand)
    best_cost = evaluate_order_groups(best_order_groups, best_yy)

    current_order_groups = copy.deepcopy(best_order_groups)
    current_yy = best_yy.copy()
    current_wave_item_swaps = []
    current_cost = best_cost

    k_max = max_iterations
    print(f"[VNS] k_max = {k_max}  |  initial cost = {best_cost}  |  "
          f"active_neighborhoods = {active_neighborhoods}")

    iteration = 0
    while iteration < k_max:
        k_idx = 0
        while k_idx < len(active_neighborhoods):
            k = active_neighborhoods[k_idx]
            # --- Shaking: perturb a COPY of the current solution in neighborhood k ---
            candidate_groups = copy.deepcopy(current_order_groups)
            candidate_yy = current_yy.copy()
            candidate_wave_item_swaps = list(current_wave_item_swaps)

            if k == 1:
                candidate_groups = perform_neighborhood_operation_3(candidate_groups)
            elif k == 2:
                candidate_groups = perform_neighborhood_operation_1(candidate_groups)
            elif k == 3:
                candidate_yy = swap_neighborhood(candidate_yy, sequences, num_picking_stations)
            elif k == 4:
                candidate_wave_item_swaps = perform_neighborhood_operation_wave_item(
                    current_wave_item_swaps, items, waves)

            candidate_cost = evaluate_order_groups(
                candidate_groups, candidate_yy,
                wave_item_swaps=candidate_wave_item_swaps if candidate_wave_item_swaps else None)

            if candidate_cost < current_cost:
                # Improvement: accept as the new current solution and
                # restart the neighborhood search from the first active one.
                current_order_groups = candidate_groups
                current_yy = candidate_yy
                current_wave_item_swaps = candidate_wave_item_swaps
                current_cost = candidate_cost
                if candidate_cost < best_cost:
                    best_order_groups = copy.deepcopy(candidate_groups)
                    best_yy = candidate_yy.copy()
                    best_cost = candidate_cost
                k_idx = 0
            else:
                # No improvement: move to the next neighborhood without
                # mutating the current solution.
                k_idx += 1

        iteration += 1

    return best_order_groups, best_cost, best_yy
#--------------------------------------------------------------------------------------------------------------------------

#Demand
#Demand = [[10, 3, 0, 4, 5, 2, 3, 0, 0, 0, 0],
 #         [4, 4, 4, 3, 0, 2, 2, 1, 3, 0, 1],
  #        [3, 7, 6, 0, 1, 5, 0, 2, 1, 2, 1],
   #       [0, 5, 2, 2, 3, 0, 1, 0, 2, 1, 7],
    #      [1, 2, 1, 1, 1, 0, 2, 3, 5, 6, 0],
     #     [5, 12, 3, 0, 4, 1, 2, 1, 0, 3, 2],
      #    [6, 3, 1, 4, 0, 1, 4, 0, 5, 0, 6],
       #   [2, 2, 3, 6, 7, 0, 0, 1, 2, 3, 4]]





#Demand = [[10, 5, 1, 2, 6],
 #              [4, 5, 4, 3, 2],
  #             [5, 6, 3, 1, 2,],
   #            [2, 5, 1, 2, 4],
    #           [2, 2, 3, 1, 1],
     #          [4, 5, 3, 1, 4],
      #         [5, 3, 2, 4, 5],
       #        [2, 3, 4, 6, 3],
        #       [4, 2, 4, 2, 5],
         #      [4, 4, 6, 3, 1],
          #     [2, 5, 3, 2, 3],
           #    [3, 2, 2, 4, 0],
            #   [8, 7, 2, 2, 3],
             #  [8, 0, 4, 5, 1],
              # [2, 3, 4, 5, 8],
               #[2, 0, 4, 3, 4],
               #[5, 1, 2, 7, 7],
               #[7, 4, 1, 2, 4],
               #[1, 2, 0, 1, 3],
               #[3, 5, 5, 6, 2],
               #[4, 6, 3, 2, 1],
               #[7, 5, 2, 6, 2],
               #[4, 2, 4, 3, 4]]

#Demand = [[10, 4, 4, 4, 0],
 #        [7, 3, 0, 2, 4],
  #       [4, 0, 6, 5, 2],
   #      [2, 3, 4, 0, 1],
    #     [0, 4, 3, 2, 5],
     #    [1, 2, 1, 2, 2]]

Sum_demand = np.sum(Demand, axis=1)
print("sss:",sum(Sum_demand))



for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=row, column=col).value == 'orders1':
            orders1_row = row
            orders1_col = col
            break

orders1 = sheet.cell(row=orders1_row + 1, column=orders1_col).value
orders = np.arange(orders1)

#orders = [0, 1, 2, 3, 4]
#orders = [0, 1, 2, 3, 4, 5]
#orders = [0, 1, 2, 3, 4, 5, 6, 7]
#orders = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
#orders = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]
#orders = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19]
#orders = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24]
#orders = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29]
#orders = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34]
#orders = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39]
#orders = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49]
#orders = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59]
#orders = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 62, 63, 64, 65, 66, 67, 68, 69]
#orders = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 77, 78, 79]
#orders = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89]
#orders = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96, 97, 98, 99]

          

for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=row, column=col).value == 'num_picking_stations':
            num_picking_stations_row = row
            num_picking_stations_col = col
            break

num_picking_stations = sheet.cell(row=num_picking_stations_row + 1, column=num_picking_stations_col).value

for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=row, column=col).value == 'group_capacity':
            group_capacity_row = row
            group_capacity_col = col
            break

group_capacity = sheet.cell(row=group_capacity_row + 1, column=group_capacity_col).value

for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=row, column=col).value == 'max_iterations':
            max_iterations_row = row
            max_iterations_col = col
            break

max_iterations = sheet.cell(row=max_iterations_row + 1, column=max_iterations_col).value

for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=row, column=col).value == 'sequences':
            sequences_row = row
            sequences_col = col
            break

sequences = sheet.cell(row=sequences_row + 1, column=sequences_col).value



#sequences =  5 #3 #8 


if __name__ == "__main__":
    num_threads = 7
    iterations_per_thread = max_iterations // num_threads


    #------------------------------------------------------------ Varibale Neighborhood Search -------------------------------------------------------------
    start = time.time()
    #print("start:",start)

    #args_list = [(iterations_per_thread, Sum_demand, orders, group_capacity, num_picking_stations, Demand, sequences)] * num_threads

    #with concurrent.futures.ThreadPoolExecutor(max_workers=num_threads) as executor:
    #    iteration_results = list(executor.map(variable_neighborhood_search_iteration, *zip(*args_list)))

    #best_result = min(iteration_results, key=lambda x: x[1])
    #best_order_groups, best_cost = best_result

    best_order_groups, best_cost, best_yy = variable_neighborhood_search(Sum_demand, orders,group_capacity, max_iterations, num_picking_stations, Demand, sequences)

    print(f"Best result after iterations (fast search estimate): {best_cost}")

    #print("Best order groups:", best_order_groups)
    print("Best cost (fast search estimate):", best_cost)

    # Final polish: re-solve just the single best solution with a tight gap
    # and a long time budget, for the precise number to report in the paper.
    final_cost = final_polish(best_order_groups, best_yy)
    print("Final polished cost (report this one):", final_cost)

    end = time.time()
    print(end - start)


