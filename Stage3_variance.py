# Table 5: Stage 3 replication variance.
"""
Table 5 -- Stage 3 Replication Variance
=========================================
Place this file in the same folder as RMFS_main.py and the instance file.

What it does:
  - Runs Stage 1 (FFD order grouping) and Stage 2 (station assignment) once
    and then holds them fixed.
  - Repeats only Stage 3 (the randomized wave-planning heuristic plus the
    reduced MIP solve, i.e. evaluate_order_groups) N times with different
    random seeds.
  - This produces Table 5: with grouping and station assignment fixed, it
    isolates how much the randomness in wave planning alone affects the
    final objective.

Usage:
  set RMFS_DATA_FILE=Data_XX.xlsx
  set TABLE5_N_REPLICATIONS=5
  set TABLE5_OUTPUT_FILE=table5_XX.xlsx
  python Stage3_variance.py

Output: table5_<label>.xlsx (or the name given via TABLE5_OUTPUT_FILE)
"""

import sys
_CPLEX_API = r'C:\Program Files\IBM\ILOG\CPLEX_Studio221\cplex\python\3.7\x64_win64'
# On another machine, set the CPLEX_PYTHON_API environment variable to the
# folder holding the CPLEX Python API, e.g.
#   Linux : /opt/ibm/ILOG/CPLEX_Studio221/cplex/python/3.7/x86-64_linux
#   macOS : /Applications/CPLEX_Studio221/cplex/python/3.7/x86-64_osx
import os
_CPLEX_API = os.environ.get('CPLEX_PYTHON_API', _CPLEX_API)
if os.path.isdir(_CPLEX_API):
    sys.path = [_CPLEX_API] + sys.path
import random
import numpy as np
import time
import os
import openpyxl

SEED = 42
random.seed(SEED)
np.random.seed(SEED)

N_REPLICATIONS = int(os.environ.get('TABLE5_N_REPLICATIONS', '5'))
OUTPUT_FILE = os.environ.get('TABLE5_OUTPUT_FILE', 'table5_results.xlsx')
INSTANCE_LABEL = os.environ.get('TABLE5_INSTANCE_LABEL', os.environ.get('RMFS_DATA_FILE', 'Data11.xlsx'))
FAIL_SENTINEL = 10000000

DATA_FILE = os.environ.get('RMFS_DATA_FILE', 'Data11.xlsx')
workbook = openpyxl.load_workbook(DATA_FILE)
sheet = workbook.active

def read_matrix(sheet, name):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == name:
                sr, sc = row + 1, col
                rows, cols = 0, 0
                while sheet.cell(row=sr + rows, column=sc).value is not None:
                    rows += 1
                while sheet.cell(row=sr, column=sc + cols).value is not None:
                    cols += 1
                data = []
                for r in range(sr, sr + rows):
                    row_data = [sheet.cell(row=r, column=c).value for c in range(sc, sc + cols)]
                    data.append(row_data)
                return data
    return None

def read_scalar(sheet, name):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == name:
                return sheet.cell(row=row + 1, column=col).value
    return None

Demand               = read_matrix(sheet, 'Demand')
Sum_demand           = np.sum(Demand, axis=1)
orders1              = read_scalar(sheet, 'orders1')
orders               = np.arange(orders1)
num_picking_stations = read_scalar(sheet, 'num_picking_stations')
group_capacity       = read_scalar(sheet, 'group_capacity')

from RMFS_main import create_order_groups, assigning_order_groups, evaluate_order_groups


# Guard against a silent instance mismatch. evaluate_order_groups and the other
# RMFS_main functions read the demand matrix, capacities and wave data from
# RMFS_main's own module-level worksheet, NOT from the arguments passed in. If
# this script and RMFS_main were pointed at different workbooks, every result
# would be computed from a mixture of two instances and no error would be
# raised. Both now read RMFS_DATA_FILE, so this check should never fire.
import RMFS_main as _rmfs_main
if os.path.abspath(_rmfs_main.DATA_FILE) != os.path.abspath(DATA_FILE):
    raise RuntimeError(
        "Instance mismatch: this script loaded %r but RMFS_main loaded %r. "
        "Set the RMFS_DATA_FILE environment variable so that both use the "
        "same workbook." % (DATA_FILE, _rmfs_main.DATA_FILE))

print(f"Instance label: {INSTANCE_LABEL}")
print(f"Instance: {orders1} orders | {num_picking_stations} picking stations | "
      f"group_capacity={group_capacity}")
print("=" * 65)

# -- Stage 1 + Stage 2 are run once and then held fixed --
fixed_groups = create_order_groups(Sum_demand, orders, group_capacity)
fixed_yy, _, _ = assigning_order_groups(fixed_groups, num_picking_stations, Demand)
print(f"Fixed grouping (Stage 1+2, will NOT change across replications): {fixed_groups}")
print("=" * 65)

costs, times = [], []
for rep in range(N_REPLICATIONS):
    seed = SEED + rep
    random.seed(seed)
    np.random.seed(seed)
    print(f"\nReplication {rep+1}/{N_REPLICATIONS}  (seed={seed}) -- Stage 3 only")
    t0 = time.time()
    cost = evaluate_order_groups(fixed_groups, fixed_yy)
    dt = time.time() - t0
    costs.append(cost)
    times.append(dt)
    status = "" if cost < FAIL_SENTINEL else "  [FAILED]"
    print(f"  -> objective={cost:.4f}  time={dt:.1f}s{status}")

ok_costs = [c for c in costs if c < FAIL_SENTINEL]
n_ok = len(ok_costs)
mean_c = np.mean(ok_costs) if ok_costs else None
std_c = np.std(ok_costs) if ok_costs else None
cv = (std_c / mean_c * 100) if (mean_c and mean_c > 0) else None

print(f"\n\n{'='*65}")
print(f"TABLE 5 RESULT -- {INSTANCE_LABEL}")
print(f"{'='*65}")
print(f"Successful replications: {n_ok}/{N_REPLICATIONS}")
if mean_c is not None:
    print(f"Mean objective: {mean_c:.4f}  |  Std. Dev.: {std_c:.4f}"
          + (f"  |  Coefficient of variation: {cv:.2f}%" if cv is not None else ""))
else:
    print("⚠️  ALL replications failed.")
print("=" * 65)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Table 5"
ws.append(["Instance", INSTANCE_LABEL])
ws.append(["Fixed grouping (Stage 1+2)", str(fixed_groups)])
ws.append([])
header = ["Instance"] + [f"Rep {i+1}" for i in range(N_REPLICATIONS)] + ["Mean", "Std. Dev."]
ws.append(header)
row = [INSTANCE_LABEL] + [round(c, 4) if c < FAIL_SENTINEL else "FAILED" for c in costs]
row += [round(mean_c, 4) if mean_c is not None else "N/A",
        round(std_c, 4) if std_c is not None else "N/A"]
ws.append(row)

ws.append([])
ws.append(["Replication", "Seed", "Objective", "Time (s)", "Failed?"])
for rep, (c, t) in enumerate(zip(costs, times)):
    ws.append([rep + 1, SEED + rep, round(c, 4) if c < FAIL_SENTINEL else "FAILED",
               round(t, 2), c >= FAIL_SENTINEL])

wb.save(OUTPUT_FILE)
print(f"\n✅ Results saved to {OUTPUT_FILE}")
