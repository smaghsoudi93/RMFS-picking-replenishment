# Table 3 re-run: full pipeline (Construction + VNS + final_polish) on ONE
# instance, repeated N times, to get the mean/std objective & time for the
# "Solution of our proposal" columns -- with the fixed model, tightened
# Big-M, corrected objective, and single-horizon solve.
"""
Table 3 Re-run (single instance)
==================================
Place this file in the same folder as:
  - RMFS_main.py
  - the instance file of interest (e.g. Data1.xlsx)

Like run_sa.py / ablation_study.py, this script works on one instance at a
time. Run it separately for each row of Table 3 (6, 8, 15, ... 100 orders):

  set RMFS_DATA_FILE=Data_6orders.xlsx
  set TABLE3_OUTPUT_FILE=table3_row_6orders.xlsx
  set TABLE3_N_REPLICATIONS=5
  python Standard_deviation.py

Timing note: for large instances (60-100 orders) a full replication
(construction + complete VNS run + final_polish) can take a long time; use
fewer replications for those (e.g. 2-3 instead of 5).

Output:
  - console log (each replication plus the final mean / std)
  - table3_row_<label>.xlsx (or the name given via TABLE3_OUTPUT_FILE)
"""

import sys
_CPLEX_API = r'C:\Program Files\IBM\ILOG\CPLEX_Studio221\cplex\python\3.7\x64_win64'
# On another machine, set the CPLEX_PYTHON_API environment variable to the
# folder holding the CPLEX Python API, e.g.
#   Linux : /opt/ibm/ILOG/CPLEX_Studio221/cplex/python/3.7/x86-64_linux
#   macOS : /Applications/CPLEX_Studio221/cplex/python/3.7/x86-64_osx
import os
_CPLEX_API = os.environ.get('CPLEX_PYTHON_API', _CPLEX_API)
if os.path.isdir(_CPLEX_API):
    sys.path = [_CPLEX_API] + sys.path
import random
import numpy as np
import time
import os
import openpyxl

SEED = 42
random.seed(SEED)
np.random.seed(SEED)

N_REPLICATIONS = int(os.environ.get('TABLE3_N_REPLICATIONS', '5'))
OUTPUT_FILE = os.environ.get('TABLE3_OUTPUT_FILE', 'table3_row_results.xlsx')
INSTANCE_LABEL = os.environ.get('TABLE3_INSTANCE_LABEL', os.environ.get('RMFS_DATA_FILE', 'Data11.xlsx'))
FAIL_SENTINEL = 10000000

# -- Independent Excel loading (same pattern as ablation_study.py / run_sa.py) --
DATA_FILE = os.environ.get('RMFS_DATA_FILE', 'Data11.xlsx')
workbook = openpyxl.load_workbook(DATA_FILE)
sheet = workbook.active

def read_matrix(sheet, name):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == name:
                sr, sc = row + 1, col
                rows, cols = 0, 0
                while sheet.cell(row=sr + rows, column=sc).value is not None:
                    rows += 1
                while sheet.cell(row=sr, column=sc + cols).value is not None:
                    cols += 1
                data = []
                for r in range(sr, sr + rows):
                    row_data = [sheet.cell(row=r, column=c).value for c in range(sc, sc + cols)]
                    data.append(row_data)
                return data
    return None

def read_scalar(sheet, name):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == name:
                return sheet.cell(row=row + 1, column=col).value
    return None

Demand                    = read_matrix(sheet, 'Demand')
Sum_demand                = np.sum(Demand, axis=1)
orders1                   = read_scalar(sheet, 'orders1')
orders                    = np.arange(orders1)
num_picking_stations      = read_scalar(sheet, 'num_picking_stations')
num_replenishment_stations = read_scalar(sheet, 'num_replenishment_stations')
group_capacity            = read_scalar(sheet, 'group_capacity')
max_iterations            = read_scalar(sheet, 'max_iterations')
sequences                 = read_scalar(sheet, 'sequences')

# -- Import only the functions from RMFS_main.py, not its module-level state --
from RMFS_main import (
    create_order_groups, assigning_order_groups,
    variable_neighborhood_search, final_polish,
)


# Guard against a silent instance mismatch. evaluate_order_groups and the other
# RMFS_main functions read the demand matrix, capacities and wave data from
# RMFS_main's own module-level worksheet, NOT from the arguments passed in. If
# this script and RMFS_main were pointed at different workbooks, every result
# would be computed from a mixture of two instances and no error would be
# raised. Both now read RMFS_DATA_FILE, so this check should never fire.
import RMFS_main as _rmfs_main
if os.path.abspath(_rmfs_main.DATA_FILE) != os.path.abspath(DATA_FILE):
    raise RuntimeError(
        "Instance mismatch: this script loaded %r but RMFS_main loaded %r. "
        "Set the RMFS_DATA_FILE environment variable so that both use the "
        "same workbook." % (DATA_FILE, _rmfs_main.DATA_FILE))

print(f"Instance label: {INSTANCE_LABEL}")
print(f"Instance: {orders1} orders | {num_picking_stations} picking stations | "
      f"{num_replenishment_stations} replenishment stations | max_iterations={max_iterations}")
print("=" * 65)


def run_one_replication(seed):
    random.seed(seed)
    np.random.seed(seed)
    t0 = time.time()

    best_groups, best_cost_search, best_yy = variable_neighborhood_search(
        Sum_demand, orders, group_capacity, max_iterations,
        num_picking_stations, Demand, sequences)

    final_cost = final_polish(best_groups, best_yy)
    elapsed = time.time() - t0
    return final_cost, elapsed


def run_table3_row(n_replications=None):
    n_replications = n_replications or N_REPLICATIONS
    costs, times = [], []

    for rep in range(n_replications):
        seed = SEED + rep
        print(f"\nReplication {rep+1}/{n_replications}  (seed={seed})")
        print("-" * 65)
        cost, t = run_one_replication(seed)
        costs.append(cost)
        times.append(t)
        status = "" if cost < FAIL_SENTINEL else "  [FAILED]"
        print(f"  -> objective={cost:.4f}  time={t:.1f}s{status}")

    ok_costs = [c for c in costs if c < FAIL_SENTINEL]
    ok_times = [t for c, t in zip(costs, times) if c < FAIL_SENTINEL]
    n_ok = len(ok_costs)

    mean_c = np.mean(ok_costs) if ok_costs else None
    std_c = np.std(ok_costs) if ok_costs else None
    mean_t = np.mean(ok_times) if ok_times else None
    std_t = np.std(ok_times) if ok_times else None

    print(f"\n\n{'='*65}")
    print(f"TABLE 3 ROW RESULT -- {INSTANCE_LABEL}")
    print(f"{'='*65}")
    print(f"Successful replications: {n_ok}/{n_replications}")
    if mean_c is not None:
        print(f"Mean objective (our proposal): {mean_c:.4f}  (std: {std_c:.4f})")
        print(f"Mean time (s):                 {mean_t:.2f}  (std: {std_t:.4f})")
        print(f"\n-> Table 3 'Objective function' column: {mean_c:.4f}")
        print(f"-> Table 3 'Computing time (s)' column:  {mean_t:.2f}")
        print(f"-> Table 3b 'Std. Dev. (Objective)':      {std_c:.4f}")
        print(f"-> Table 3b 'Std. Dev. (Time)':           {std_t:.4f}")
    else:
        print("⚠️  ALL replications failed -- no usable result for this instance.")
    print("=" * 65)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Table 3 Row"
    ws.append(["Instance", INSTANCE_LABEL])
    ws.append(["Orders", orders1, "Picking stations", num_picking_stations,
               "Replenishment stations", num_replenishment_stations])
    ws.append([])
    ws.append(["Replication", "Seed", "Objective", "Time (s)", "Failed?"])
    for rep, (c, t) in enumerate(zip(costs, times)):
        ws.append([rep + 1, SEED + rep,
                   round(c, 4) if c < FAIL_SENTINEL else "FAILED",
                   round(t, 2), c >= FAIL_SENTINEL])

    ws.append([])
    ws.append(["Summary (Table 3 / Table 3b values)"])
    ws.append(["Mean Objective", "Std Dev Objective", "Mean Time (s)", "Std Dev Time",
               "Successes", "Total Reps"])
    ws.append([
        round(mean_c, 4) if mean_c is not None else "N/A",
        round(std_c, 4) if std_c is not None else "N/A",
        round(mean_t, 2) if mean_t is not None else "N/A",
        round(std_t, 4) if std_t is not None else "N/A",
        n_ok, n_replications,
    ])

    wb.save(OUTPUT_FILE)
    print(f"\n✅ Results saved to {OUTPUT_FILE}")
    return costs, times


if __name__ == "__main__":
    run_table3_row()
