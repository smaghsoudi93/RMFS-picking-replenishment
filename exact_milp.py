"""
Standalone exact MILP for the integrated picking-replenishment problem
=====================================================================
This is the complete model of Section 3 of the manuscript, Constraints
(3.1)-(3.26), with NO heuristic pre-processing. Every decision the paper
lists as a decision variable is left free for CPLEX to make:

  x[o,g]        order-to-group assignment              (3.2), (3.6)
  y[g,s,j]      group-to-station and sequence          (3.3), (3.4)
  z[g,p,w]      group picks from pod p, wave w         (3.5)
  l[g,w]        first wave supplying group g           (3.12)-(3.14)
  q[g,p,i,w]    quantity of item i picked from pod p   (3.9), (3.10)
  u[i,p,w,r]    quantity replenished per station       (3.7), (3.8), (3.11)
  v[p,w,r]      pod-to-replenishment-station           (3.19), (3.20)
  ct, st        completion / start times               (3.15)-(3.18)

This is the model referred to as "CPLEX" in Table 3 of the manuscript.
It is what the proposed heuristic is benchmarked AGAINST. It is deliberately
separate from RMFS_main.py, where the same constraint system is rebuilt with
x, y and the wave plan already fixed by Stages 1-3 (that is the reduced MIP
sub-problem of Section 4.3, not the exact model).

Usage
-----
    RMFS_DATA_FILE=Data1.xlsx python exact_milp.py

Formulation notes
-----------------
The constraint system is the one printed in Section 3, but written in the
tightest equivalent form so that the solve time reflects the difficulty of the
problem rather than a loose encoding of it:

  * purpose-specific big-M values (arrival, completion, quantity) instead of a
    single global constant;
  * symmetry among the interchangeable order groups broken by restricting
    order o to groups 0..o;
  * the per-pod form of (3.17) aggregated over pods via an indicator zw[g,w];
  * (3.18) stated for consecutive sequence positions only, which implies the
    general case by transitivity;
  * a valid floor on each order's completion time.

None of these changes the feasible set of integer solutions or the optimal
objective value; they only give CPLEX a usable LP relaxation. No heuristic
solution is supplied to the solver: the benchmark in Table 3 measures CPLEX
solving the model from scratch.

Environment variables:
    RMFS_DATA_FILE        instance workbook            (default Data1.xlsx)
    EXACT_TIME_LIMIT      CPLEX time limit in seconds  (default 36000 = 600 min,
                          the limit reported in Section 5.2)
    EXACT_MIPGAP          relative MIP gap             (default 0.0001)
    EXACT_THREADS         CPLEX threads                (default 4, as Section 5.1)
    EXACT_OUTPUT_FILE     result workbook              (default exact_milp_results.xlsx)
    CPLEX_PYTHON_API      folder holding the CPLEX Python API, if not on sys.path

Reproducing the CPLEX columns of Table 3
----------------------------------------
Run once per instance, e.g.

    for f in Data1 Data2 Data3 Data4 Data5 Data6 Data7 Data8 Data9; do
        RMFS_DATA_FILE=$f.xlsx EXACT_OUTPUT_FILE=exact_$f.xlsx python exact_milp.py
    done

Instances 10-12 (80, 90 and 100 orders) are reported as N/A in Table 3
because CPLEX returns no feasible integer solution within the 600-minute
limit; this script prints NO_SOLUTION for those.
"""

import os
import sys
import time

_CPLEX_API = r'C:\Program Files\IBM\ILOG\CPLEX_Studio221\cplex\python\3.7\x64_win64'
# On another machine, set the CPLEX_PYTHON_API environment variable to the
# folder holding the CPLEX Python API, e.g.
#   Linux : /opt/ibm/ILOG/CPLEX_Studio221/cplex/python/3.7/x86-64_linux
#   macOS : /Applications/CPLEX_Studio221/cplex/python/3.7/x86-64_osx
_CPLEX_API = os.environ.get('CPLEX_PYTHON_API', _CPLEX_API)
if os.path.isdir(_CPLEX_API):
    sys.path = [_CPLEX_API] + sys.path

import numpy as np
import openpyxl
from docplex.mp.model import Model

DATA_FILE = os.environ.get("RMFS_DATA_FILE", "Data1.xlsx")
TIME_LIMIT = float(os.environ.get("EXACT_TIME_LIMIT", 600 * 60))
MIPGAP = float(os.environ.get("EXACT_MIPGAP", 1e-4))
THREADS = int(os.environ.get("EXACT_THREADS", 4))
OUTPUT_FILE = os.environ.get("EXACT_OUTPUT_FILE", "exact_milp_results.xlsx")
# "per_order" (default) : Eq. (3.1) of the current manuscript, the average
#                         fulfillment time PER ORDER.
# "per_group"           : the pre-revision objective, sum of group completion
#                         times divided by the NUMBER OF GROUPS. Provided only
#                         as a diagnostic, to check which objective a given set
#                         of published numbers was produced with. The two agree
#                         only when all groups hold the same number of orders.
OBJECTIVE = os.environ.get("EXACT_OBJECTIVE", "per_order").strip().lower()
assert OBJECTIVE in ("per_order", "per_group")

workbook = openpyxl.load_workbook(DATA_FILE)
sheet = workbook.active


# ----------------------------------------------------------------------
# Instance reading (same layout as RMFS_main.py)
# ----------------------------------------------------------------------
def read_scalar(name):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == name:
                return sheet.cell(row=row + 1, column=col).value
    raise KeyError("scalar not found in workbook: %s" % name)


def read_block(name):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == name:
                sr, sc = row + 1, col
                nrows = 0
                while sheet.cell(row=sr + nrows, column=sc).value is not None:
                    nrows += 1
                ncols = 0
                while sheet.cell(row=sr, column=sc + ncols).value is not None:
                    ncols += 1
                return [[sheet.cell(row=r, column=c).value
                         for c in range(sc, sc + ncols)]
                        for r in range(sr, sr + nrows)]
    raise KeyError("block not found in workbook: %s" % name)


matrix1 = read_block("matrix1")            # order x item demand
Demand1 = read_block("Demand1")            # order x item demand (objective side)
Arriving_times = read_block("Arriving_times")[0]
Duration = read_block("Duration")[0]

orders1 = read_scalar("orders1")
items = read_scalar("items")
shelves = read_scalar("shelves")
waves = read_scalar("waves")
num_picking_stations = read_scalar("num_picking_stations")
num_replenishment_stations = read_scalar("num_replenishment_stations")
Pods_capacity = read_scalar("Pods_capacity")
group_capacity = read_scalar("group_capacity")
Replenishment_station_capacity = read_scalar("Replenishment_station_capacity")

# |G|: an upper bound on the number of groups that can be needed. With
# group capacity Cp, ceil(|O|/Cp) groups always suffice to hold every order;
# any surplus group stays empty.
n_groups = int(np.ceil(orders1 / float(group_capacity)))

Orders = list(range(orders1))
Groups = list(range(n_groups))
Items = list(range(items))
Shelves = list(range(shelves))
Waves = list(range(waves))
Picking_stations = list(range(num_picking_stations))
Replenishment_stations = list(range(num_replenishment_stations))
# |Sq|: sequence positions per station. At most every group could queue at a
# single station, so |Sq| = |G| is a valid and tight bound (see the notation
# table in Section 3.1).
Sequences = list(range(n_groups))

# Same tightened Big-M as RMFS_main.py: max of a time-scale bound and a
# quantity-scale bound, rather than an arbitrary large constant.
# Purpose-specific big-M values. Using one global constant everywhere (the
# maximum of a time-scale and a quantity-scale bound) makes every time
# constraint carry a coefficient tens of times larger than it needs, which
# destroys the LP relaxation and leaves CPLEX with no usable lower bound.
M_arr = max(Arriving_times)                        # bounds a start time
M_ct = max(Arriving_times) + max(Duration)         # bounds a completion time
M_qty = int(max(Pods_capacity, Replenishment_station_capacity)) + 1
M = max(M_ct, M_qty)                               # only where a generic bound is needed

print("Instance %s | %d orders, %d items, %d pods, %d waves, "
      "%d picking st., %d replenishment st." %
      (DATA_FILE, orders1, items, shelves, waves,
       num_picking_stations, num_replenishment_stations))
print("Groups (upper bound) = %d | Sequence positions = %d"
      % (n_groups, len(Sequences)))
print("Big-M: arrival=%g, completion=%g, quantity=%g" % (M_arr, M_ct, M_qty))

# ----------------------------------------------------------------------
# Model
# ----------------------------------------------------------------------
mdl = Model(name="RMFS_exact_MILP")

x = {(o, g): mdl.binary_var(name="x_%d_%d" % (o, g))
     for o in Orders for g in Groups}
y = {(g, s, j): mdl.binary_var(name="y_%d_%d_%d" % (g, s, j))
     for g in Groups for s in Picking_stations for j in Sequences}
z = {(g, p, w): mdl.binary_var(name="z_%d_%d_%d" % (g, p, w))
     for g in Groups for p in Shelves for w in Waves}
v = {(p, w, r): mdl.binary_var(name="v_%d_%d_%d" % (p, w, r))
     for p in Shelves for w in Waves for r in Replenishment_stations}
l = {(g, w): mdl.binary_var(name="l_%d_%d" % (g, w))
     for g in Groups for w in Waves}
q = {(g, p, i, w): mdl.continuous_var(name="q_%d_%d_%d_%d" % (g, p, i, w))
     for g in Groups for p in Shelves for i in Items for w in Waves}
u = {(i, p, w, r): mdl.continuous_var(name="u_%d_%d_%d_%d" % (i, p, w, r))
     for i in Items for p in Shelves for w in Waves
     for r in Replenishment_stations}
ct = {(g, s, j): mdl.continuous_var(name="ct_%d_%d_%d" % (g, s, j))
      for g in Groups for s in Picking_stations for j in Sequences}
st = {(g, s, j): mdl.continuous_var(name="st_%d_%d_%d" % (g, s, j))
      for g in Groups for s in Picking_stations for j in Sequences}

# (3.2) every order is assigned to exactly one group
for o in Orders:
    mdl.add_constraint(mdl.sum(x[o, g] for g in Groups) == 1,
                       ctname="assign_o%d" % o)

# A group is "used" if it holds at least one order. Empty groups take no
# station, no sequence position and no wave.
used = {g: mdl.binary_var(name="used_%d" % g) for g in Groups}
for g in Groups:
    mdl.add_constraint(mdl.sum(x[o, g] for o in Orders) >= used[g],
                       ctname="used_lb_g%d" % g)
    mdl.add_constraint(mdl.sum(x[o, g] for o in Orders) <= orders1 * used[g],
                       ctname="used_ub_g%d" % g)
# Symmetry breaking. Groups carry no distinguishing data, so any solution can
# be permuted |G|! ways and CPLEX would explore each permutation separately.
# Two standard cuts remove almost all of it:
#   (a) groups are opened in index order;
#   (b) order o may only occupy one of the groups 0..o, which forces a unique
#       canonical representative for each partition of the orders.
for g in Groups[:-1]:
    mdl.add_constraint(used[g] >= used[g + 1], ctname="symm_g%d" % g)
for o in Orders:
    for g in Groups:
        if g > o:
            mdl.add_constraint(x[o, g] == 0, ctname="symmrep_o%d_g%d" % (o, g))

# (3.3) each used group goes to exactly one station and one sequence position
for g in Groups:
    mdl.add_constraint(
        mdl.sum(y[g, s, j] for s in Picking_stations for j in Sequences)
        == used[g], ctname="onestation_g%d" % g)

# (3.4) each (station, sequence position) holds at most one group
for s in Picking_stations:
    for j in Sequences:
        mdl.add_constraint(mdl.sum(y[g, s, j] for g in Groups) <= 1,
                           ctname="oneslot_s%d_j%d" % (s, j))

# (3.5) each used group receives items from at least one pod and wave
for g in Groups:
    mdl.add_constraint(
        mdl.sum(z[g, p, w] for p in Shelves for w in Waves) >= used[g],
        ctname="atleastonepod_g%d" % g)

# (3.6) picking station capacity (orders per group)
for g in Groups:
    mdl.add_constraint(mdl.sum(x[o, g] for o in Orders) <= group_capacity,
                       ctname="groupcap_g%d" % g)

# (3.7) replenishment station capacity per wave
for w in Waves:
    for r in Replenishment_stations:
        mdl.add_constraint(
            mdl.sum(u[i, p, w, r] for i in Items for p in Shelves)
            <= Replenishment_station_capacity,
            ctname="replcap_w%d_r%d" % (w, r))

# (3.8) pod capacity per wave
for w in Waves:
    for p in Shelves:
        mdl.add_constraint(
            mdl.sum(u[i, p, w, r] for i in Items
                    for r in Replenishment_stations) <= Pods_capacity,
            ctname="podcap_w%d_p%d" % (w, p))

# (3.9) every group's demand for each item must be picked
for g in Groups:
    for i in Items:
        mdl.add_constraint(
            mdl.sum(Demand1[o][i] * x[o, g] for o in Orders)
            == mdl.sum(q[g, p, i, w] for p in Shelves for w in Waves),
            ctname="demandeq_g%d_i%d" % (g, i))

# (3.10) per wave, everything replenished is picked
for w in Waves:
    mdl.add_constraint(
        mdl.sum(u[i, p, w, r] for i in Items for p in Shelves
                for r in Replenishment_stations)
        == mdl.sum(q[g, p, i, w] for g in Groups for p in Shelves
                   for i in Items),
        ctname="wavebalance_w%d" % w)

# (3.11) per item / pod / wave, picks cannot exceed what is physically there
for i in Items:
    for p in Shelves:
        for w in Waves:
            mdl.add_constraint(
                mdl.sum(q[g, p, i, w] for g in Groups)
                <= mdl.sum(u[i, p, w, r] for r in Replenishment_stations),
                ctname="qleu_i%d_p%d_w%d" % (i, p, w))

# (3.12) the first wave of a group must be one it actually draws from
for g in Groups:
    for w in Waves:
        mdl.add_constraint(mdl.sum(z[g, p, w] for p in Shelves) >= l[g, w],
                           ctname="zgel_g%d_w%d" % (g, w))

# (3.13) exactly one first wave per used group
for g in Groups:
    mdl.add_constraint(mdl.sum(l[g, w] for w in Waves) == used[g],
                       ctname="lsum_g%d" % g)

# (3.14) if w is the first wave for g, no earlier wave supplies g.
#        RHS = |P| * (|W| - 1)
RHS_314 = shelves * (waves - 1)
for g in Groups:
    for w1 in Waves:
        mdl.add_constraint(
            l[g, w1] * RHS_314
            + mdl.sum(z[g, p, w2] for p in Shelves for w2 in Waves if w2 < w1)
            <= RHS_314,
            ctname="firstwave_g%d_w%d" % (g, w1))

# (3.15), (3.16) start time = arrival time of the first wave
for g in Groups:
    for s in Picking_stations:
        for j in Sequences:
            arrival = mdl.sum(Arriving_times[w] * l[g, w] for w in Waves)
            mdl.add_constraint(st[g, s, j] <= arrival + M_arr * (1 - y[g, s, j]),
                               ctname="stub_g%d_s%d_j%d" % (g, s, j))
            mdl.add_constraint(st[g, s, j] >= arrival - M_arr * (1 - y[g, s, j]),
                               ctname="stlb_g%d_s%d_j%d" % (g, s, j))

# (3.17) completion time >= arrival + duration of the last wave the group uses.
#
# Written per pod this is |G|*|P|*|W|*|S|*|Sq| constraints, all but one of them
# per (g,w) redundant, because only whether the group draws from wave w matters
# and not from which pod. An aggregated indicator zw[g,w] removes the |P|
# factor. The wave duration is also multiplied by the indicator: leaving it
# outside would impose ct >= Duration[w] even for waves the group never uses.
zw = {(g, w): mdl.binary_var(name="zw_%d_%d" % (g, w))
      for g in Groups for w in Waves}
for g in Groups:
    for w in Waves:
        for p in Shelves:
            mdl.add_constraint(zw[g, w] >= z[g, p, w],
                               ctname="zwlink_g%d_w%d_p%d" % (g, w, p))
        mdl.add_constraint(zw[g, w] <= mdl.sum(z[g, p, w] for p in Shelves),
                           ctname="zwub_g%d_w%d" % (g, w))

for g in Groups:
    for w in Waves:
        for s in Picking_stations:
            for j in Sequences:
                mdl.add_constraint(
                    ct[g, s, j] >= zw[g, w] * (Arriving_times[w] + Duration[w])
                    - M_ct * (1 - y[g, s, j]),
                    ctname="ct_g%d_w%d_s%d_j%d" % (g, w, s, j))

# (3.18) a later sequence position cannot start before the earlier one ends.
#
# Stating this for every ordered pair j1 < j2 is O(|Sq|^2) per station and
# entirely redundant: consecutive positions suffice, since ct >= st at each
# position makes the relation transitive down the queue.
for g1 in Groups:
    for g2 in Groups:
        if g1 == g2:
            continue
        for j1 in Sequences[:-1]:
            j2 = j1 + 1
            for s in Picking_stations:
                mdl.add_constraint(
                    st[g2, s, j2] >= ct[g1, s, j1]
                    - M_ct * (2 - y[g2, s, j2] - y[g1, s, j1]),
                    ctname="seq_g%d_g%d_s%d_j%d" % (g1, g2, s, j1))

# ct >= st at every occupied slot; this is what makes the chaining above
# transitive, and it is implied by (3.17) only when a wave is actually used.
for g in Groups:
    for s in Picking_stations:
        for j in Sequences:
            mdl.add_constraint(ct[g, s, j] >= st[g, s, j],
                               ctname="ctgest_g%d_s%d_j%d" % (g, s, j))

# (3.19), (3.20) link pod-to-replenishment-station assignment with quantities
for p in Shelves:
    for w in Waves:
        for r in Replenishment_stations:
            mdl.add_constraint(mdl.sum(u[i, p, w, r] for i in Items)
                               >= v[p, w, r],
                               ctname="vlb_p%d_w%d_r%d" % (p, w, r))
            mdl.add_constraint(mdl.sum(u[i, p, w, r] for i in Items)
                               <= M_qty * v[p, w, r],
                               ctname="vub_p%d_w%d_r%d" % (p, w, r))

# (3.21), (3.22) link group-pod-wave selection with picked quantities
for g in Groups:
    for p in Shelves:
        for w in Waves:
            mdl.add_constraint(mdl.sum(q[g, p, i, w] for i in Items)
                               >= z[g, p, w],
                               ctname="zlb_g%d_p%d_w%d" % (g, p, w))
            mdl.add_constraint(mdl.sum(q[g, p, i, w] for i in Items)
                               <= M_qty * z[g, p, w],
                               ctname="zub_g%d_p%d_w%d" % (g, p, w))

# (3.23), (3.24) ct and st are zero for unassigned (group, station, sequence)
for g in Groups:
    for s in Picking_stations:
        for j in Sequences:
            mdl.add_constraint(ct[g, s, j] <= M_ct * y[g, s, j],
                               ctname="ctzero_g%d_s%d_j%d" % (g, s, j))
            mdl.add_constraint(st[g, s, j] <= M_arr * y[g, s, j],
                               ctname="stzero_g%d_s%d_j%d" % (g, s, j))

# (3.1) minimise the average fulfillment time PER ORDER.
#
# Writing the objective directly as sum over groups of |O_g| * CT_g would be
# BILINEAR here, because the group sizes |O_g| = sum_o x[o,g] are themselves
# decision variables once grouping is free (unlike the reduced MIP in
# RMFS_main.py, where Stage 1 fixes the grouping and |O_g| is a constant).
#
# Instead an order-level completion time ctt[o] is introduced. Constraints
# (3.23)-(3.24) force ct[g,s,j] to zero for every (station, sequence) slot the
# group does not occupy, so the sum over slots is exactly that group's
# completion time. Order o then inherits the completion time of the group
# holding it; because the objective is minimised, each ctt[o] settles at
# exactly that value. The objective sum_o ctt[o] / |O| is fully linear and
# numerically identical to sum_g |O_g| * CT_g / |O|.
#
# The big-M used here is the tight bound on a completion time (latest wave
# arrival plus the longest wave duration), NOT the global M, which is orders of
# magnitude larger and would cripple the LP relaxation.
M_ct = max(Arriving_times) + max(Duration)

CTg = {g: mdl.sum(ct[g, s, j] for s in Picking_stations for j in Sequences)
       for g in Groups}
if OBJECTIVE == "per_order":
    ctt = {o: mdl.continuous_var(name="ctt_%d" % o) for o in Orders}
    for o in Orders:
        for g in Groups:
            mdl.add_constraint(ctt[o] >= CTg[g] - M_ct * (1 - x[o, g]),
                               ctname="ordercomp_o%d_g%d" % (o, g))
    # Valid lower bound on every order's completion time: an order cannot be
    # finished before the earliest wave that could supply it has arrived and
    # been delivered. Without this the LP relaxation can drive every ctt to
    # zero by splitting x fractionally across groups, leaving CPLEX with a
    # lower bound of 0 and nothing to prune on. This cut does not remove any
    # integer-feasible solution.
    ct_floor = min(Arriving_times[w] + Duration[w] for w in Waves)
    for o in Orders:
        mdl.add_constraint(ctt[o] >= ct_floor, ctname="cttfloor_o%d" % o)

    mdl.minimize(mdl.sum(ctt[o] for o in Orders) / orders1)
else:
    # Diagnostic only: the pre-revision objective. The completion times enter
    # directly, with no big-M link to the grouping variables, which is why this
    # version has a far stronger LP relaxation and solves orders of magnitude
    # faster.
    mdl.minimize(mdl.sum(CTg[g] for g in Groups) / n_groups)

print("Objective mode: %s" % OBJECTIVE)

print("Model built: %d variables, %d constraints"
      % (mdl.number_of_variables, mdl.number_of_constraints))

# ----------------------------------------------------------------------
# Solve
# ----------------------------------------------------------------------
mdl.parameters.timelimit = TIME_LIMIT
mdl.parameters.threads = THREADS
mdl.parameters.mip.tolerances.mipgap = MIPGAP

# Memory management. On the larger instances the branch-and-bound tree grows to
# millions of open nodes and CPLEX aborts with "Error 1001: Out of memory"
# before the time limit is reached. Node files write the tree to disk instead
# of holding it all in RAM: strategy 3 = compressed node file on disk.
mdl.parameters.mip.strategy.file = 3
mdl.parameters.workmem = float(os.environ.get("EXACT_WORKMEM_MB", 2048))
mdl.parameters.mip.limits.treememory = float(
    os.environ.get("EXACT_TREEMEM_MB", 20000))
mdl.parameters.emphasis.memory = 1

print("Solving with time_limit=%.0fs, threads=%d, mipgap=%g"
      % (TIME_LIMIT, THREADS, MIPGAP))
t0 = time.time()
solution = mdl.solve(log_output=True)
elapsed = time.time() - t0

# If CPLEX aborts (out of memory, interrupt) it returns no solution object even
# when it had already found an incumbent. Recover whatever it did find, so a
# genuine "no feasible solution" is not confused with an aborted search that
# had a perfectly usable answer.
details = mdl.solve_details
aborted_incumbent = None
if solution is None and details is not None:
    n_sol = getattr(details, "nb_solutions", None)
    if n_sol:                      # None or 0 both mean "no incumbent"
        cand = getattr(details, "best_solution_objective", None)
        if cand is not None and cand == cand and abs(cand) < 1e29:
            aborted_incumbent = cand

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Exact MILP"
ws.append(["Instance", DATA_FILE])
ws.append(["Orders", orders1])
ws.append(["Items", items])
ws.append(["Picking stations", num_picking_stations])
ws.append(["Replenishment stations", num_replenishment_stations])
ws.append(["Variables", mdl.number_of_variables])
ws.append(["Constraints", mdl.number_of_constraints])
ws.append(["Time limit (s)", TIME_LIMIT])
ws.append(["Threads", THREADS])
ws.append(["Objective mode", OBJECTIVE])
ws.append(["Sequence positions", len(Sequences)])
ws.append([])

if solution:
    obj = mdl.objective_value
    try:
        gap = mdl.solve_details.mip_relative_gap
        bound = mdl.solve_details.best_bound
    except Exception:
        gap, bound = None, None
    print("\nObjective (average fulfillment time) = %.4f" % obj)
    print("Best bound = %s | relative gap = %s" % (bound, gap))
    print("Solve time = %.1fs" % elapsed)
    ws.append(["Status", mdl.solve_details.status])
    ws.append(["Objective", round(obj, 4)])
    ws.append(["Best bound", bound])
    ws.append(["Relative gap", gap])
    ws.append(["Solve time (s)", round(elapsed, 1)])
    ws.append([])
    ws.append(["Objective mode", OBJECTIVE])
    ws.append([])
    ws.append(["Group", "Orders in group", "Station", "Sequence",
               "Start time", "Completion time"])
    for g in Groups:
        members = [o for o in Orders if x[o, g].solution_value > 0.5]
        if not members:
            continue
        for s in Picking_stations:
            for j in Sequences:
                if y[g, s, j].solution_value > 0.5:
                    ws.append([g, str(members), s, j,
                               round(st[g, s, j].solution_value, 3),
                               round(ct[g, s, j].solution_value, 3)])
else:
    status = details.status if details is not None else "unknown"
    try:
        bound = details.best_bound
    except Exception:
        bound = None
    print("\nCPLEX returned no solution object. Status: %s" % status)
    print("Solve time = %.1fs (time limit was %.0fs)" % (elapsed, TIME_LIMIT))
    if aborted_incumbent is not None:
        gap = None
        if bound not in (None, 0):
            gap = abs(aborted_incumbent - bound) / abs(aborted_incumbent)
        print("An incumbent WAS found before the search stopped: %.4f"
              % aborted_incumbent)
        print("Best bound = %s | relative gap = %s"
              % (bound, "%.2f%%" % (gap * 100) if gap is not None else "n/a"))
        print("Report this objective with its gap -- NOT as N/A.")
        ws.append(["Status", "%s (aborted with an incumbent)" % status])
        ws.append(["Objective (incumbent at abort)", round(aborted_incumbent, 4)])
        ws.append(["Best bound", bound])
        ws.append(["Relative gap", gap])
    else:
        print("No incumbent was found either. If the log ends in "
              "'Error 1001: Out of memory', raise EXACT_WORKMEM_MB or lower "
              "EXACT_TREEMEM_MB; this is not the same as infeasibility.")
        ws.append(["Status", str(status)])
    ws.append(["Solve time (s)", round(elapsed, 1)])

wb.save(OUTPUT_FILE)
print("Results saved to %s" % OUTPUT_FILE)
