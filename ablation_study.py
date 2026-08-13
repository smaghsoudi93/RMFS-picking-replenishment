"""
Ablation Study
==============
Place this file in the same folder as:
  - RMFS_main.py
  - Data11.xlsx

Three configurations are compared:
  Config 1: Construction only  (Stages 1-3, no VNS)
  Config 2: Full algorithm     (Stages 1-3 + VNS) -- the proposed method
  Config 3: Random init + VNS  (random grouping + VNS)

Output:
  - console log
  - ablation_results.xlsx
"""

import sys
_CPLEX_API = r'C:\Program Files\IBM\ILOG\CPLEX_Studio221\cplex\python\3.7\x64_win64'
# On another machine, set the CPLEX_PYTHON_API environment variable to the
# folder holding the CPLEX Python API, e.g.
#   Linux : /opt/ibm/ILOG/CPLEX_Studio221/cplex/python/3.7/x86-64_linux
#   macOS : /Applications/CPLEX_Studio221/cplex/python/3.7/x86-64_osx
import os
_CPLEX_API = os.environ.get('CPLEX_PYTHON_API', _CPLEX_API)
if os.path.isdir(_CPLEX_API):
    sys.path = [_CPLEX_API] + sys.path

from docplex.mp.model import Model
import random
import numpy as np
import copy
import time
import openpyxl

# -- Fixed seed --
SEED = 42
random.seed(SEED)
np.random.seed(SEED)

# ══════════════════════════════════════════════════════
# Read instance data from Excel
# ══════════════════════════════════════════════════════
# The instance is selected with the RMFS_DATA_FILE environment variable, the
# same convention used by every other script in this repository.
DATA_FILE = os.environ.get('RMFS_DATA_FILE', 'Data1.xlsx')
workbook = openpyxl.load_workbook(DATA_FILE)
sheet = workbook.active

def read_matrix(sheet, name):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == name:
                sr, sc = row + 1, col
                rows, cols = 0, 0
                while sheet.cell(row=sr + rows, column=sc).value is not None:
                    rows += 1
                while sheet.cell(row=sr, column=sc + cols).value is not None:
                    cols += 1
                data = []
                for r in range(sr, sr + rows):
                    row_data = []
                    for c in range(sc, sc + cols):
                        row_data.append(sheet.cell(row=r, column=c).value)
                    data.append(row_data)
                return data
    return None

def read_scalar(sheet, name):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == name:
                return sheet.cell(row=row + 1, column=col).value
    return None

Demand               = read_matrix(sheet, 'Demand')
Sum_demand           = np.sum(Demand, axis=1)
orders1              = read_scalar(sheet, 'orders1')
orders               = np.arange(orders1)
num_picking_stations = read_scalar(sheet, 'num_picking_stations')
group_capacity       = read_scalar(sheet, 'group_capacity')
max_iterations       = read_scalar(sheet, 'max_iterations')
sequences            = read_scalar(sheet, 'sequences')

print(f"Instance: {orders1} orders | {num_picking_stations} stations | "
      f"group_cap={group_capacity} | max_iter={max_iterations}")
print("=" * 65)

# ══════════════════════════════════════════════════════
# Functions
# ══════════════════════════════════════════════════════

def create_order_groups(Sum_demand, orders, group_capacity):
    """Stage 1: Modified FFD"""
    order_groups = []
    remaining_orders = list(Sum_demand.copy())
    remaining_numbers = list(orders.copy())
    group_count = 0
    order_groups.append([])

    while len(remaining_orders) > 0:
        sorted_pairs = sorted(zip(remaining_orders, remaining_numbers), reverse=True)
        remaining_orders = [d for d, n in sorted_pairs]
        remaining_numbers = [n for d, n in sorted_pairs]
        order_groups[group_count].append(remaining_numbers[0])
        remaining_orders.pop(0)
        remaining_numbers.pop(0)

        if len(remaining_orders) > 0:
            if len(order_groups[group_count]) < group_capacity:
                sorted_pairs = sorted(zip(remaining_orders, remaining_numbers))
                remaining_orders = [d for d, n in sorted_pairs]
                remaining_numbers = [n for d, n in sorted_pairs]
                order_groups[group_count].append(remaining_numbers[0])
                remaining_orders.pop(0)
                remaining_numbers.pop(0)
            else:
                group_count += 1
                order_groups.append([])
                sorted_pairs = sorted(zip(remaining_orders, remaining_numbers))
                remaining_orders = [d for d, n in sorted_pairs]
                remaining_numbers = [n for d, n in sorted_pairs]
                order_groups[group_count].append(remaining_numbers[0])
                remaining_orders.pop(0)
                remaining_numbers.pop(0)

        if len(order_groups[group_count]) >= group_capacity and len(remaining_orders) > 0:
            group_count += 1
            order_groups.append([])

    return order_groups


def create_random_order_groups(orders, group_capacity):
    """Alternative Stage 1: random grouping, used by Config 3."""
    shuffled = list(orders)
    random.shuffle(shuffled)
    groups = []
    for i in range(0, len(shuffled), group_capacity):
        chunk = shuffled[i:i + group_capacity]
        if chunk:
            groups.append(chunk)
    return groups


def assigning_order_groups(order_groups, num_picking_stations, Demand):
    """Stage 2: Station assignment"""
    yy = np.zeros((len(order_groups), num_picking_stations, len(Demand)))
    active = np.zeros((len(order_groups), num_picking_stations, len(Demand)))
    ss, jj = 0, 0
    for g in range(len(order_groups)):
        yy[g, ss, jj] = 1
        active[g, ss, jj] = 1
        ss += 1
        if ss >= num_picking_stations:
            jj += 1
            ss = 0
    return yy, active, jj


def find_ss_jj(y, group, sequences, num_picking_stations):
    for ss in range(num_picking_stations):
        for jj in range(sequences):
            if y[group][ss][jj] == 1:
                return ss, jj
    return 0, 0


def swap_neighborhood(yy, sequences, num_picking_stations):
    new_y = yy.copy()
    group1, group2 = random.sample(range(len(new_y)), 2)
    ss1, jj1 = find_ss_jj(new_y, group1, sequences, num_picking_stations)
    ss2, jj2 = find_ss_jj(new_y, group2, sequences, num_picking_stations)
    new_y[group1][ss1][jj1] = 0
    new_y[group2][ss2][jj2] = 0
    new_y[group2][ss1][jj1] = 1
    new_y[group1][ss2][jj2] = 1
    return new_y


def perform_neighborhood_operation_1(order_groups):
    if len(order_groups) < 2:
        return order_groups
    g1, g2 = random.sample(range(len(order_groups)), 2)
    if order_groups[g1] and order_groups[g2]:
        i1 = random.randint(0, len(order_groups[g1]) - 1)
        i2 = random.randint(0, len(order_groups[g2]) - 1)
        order_groups[g1][i1], order_groups[g2][i2] = \
            order_groups[g2][i2], order_groups[g1][i1]
    return order_groups


def perform_neighborhood_operation_3(order_groups):
    if len(order_groups) < 2:
        return order_groups
    g1, g2 = random.sample(range(len(order_groups)), 2)
    order_groups[g1], order_groups[g2] = order_groups[g2], order_groups[g1]
    return order_groups


# Imported from the main implementation.
from RMFS_main import (
    evaluate_order_groups, variable_neighborhood_search,
    ACTIVE_NEIGHBORHOODS,
)


# Guard against a silent instance mismatch. evaluate_order_groups and the other
# RMFS_main functions read the demand matrix, capacities and wave data from
# RMFS_main's own module-level worksheet, NOT from the arguments passed in. If
# this script and RMFS_main were pointed at different workbooks, every result
# would be computed from a mixture of two instances and no error would be
# raised. Both now read RMFS_DATA_FILE, so this check should never fire.
import RMFS_main as _rmfs_main
if os.path.abspath(_rmfs_main.DATA_FILE) != os.path.abspath(DATA_FILE):
    raise RuntimeError(
        "Instance mismatch: this script loaded %r but RMFS_main loaded %r. "
        "Set the RMFS_DATA_FILE environment variable so that both use the "
        "same workbook." % (DATA_FILE, _rmfs_main.DATA_FILE))


# NOTE: this function used to be a local re-implementation of the VNS loop.
# It applied every move in each iteration (which is not the VNS scheme) and
# discarded the return value of swap_neighborhood, so the station/sequence
# neighborhood never took effect. Table 6 is now produced by calling the same
# variable_neighborhood_search used everywhere else, so "Construction + VNS"
# here means exactly what it means in Tables 3 and 7.
def run_vns(order_groups, yy, max_iterations, sequences, num_picking_stations):
    """Run the VNS from RMFS_main, starting from the given solution."""
    best_groups, best_cost, _ = variable_neighborhood_search(
        Sum_demand, orders, group_capacity, max_iterations,
        num_picking_stations, Demand, sequences,
        active_neighborhoods=ACTIVE_NEIGHBORHOODS,
        initial_order_groups=order_groups, initial_yy=yy)
    return best_groups, best_cost


# ══════════════════════════════════════════════════════
# The three configurations
# ══════════════════════════════════════════════════════

def config1_construction_only():
    t0 = time.time()
    groups = create_order_groups(Sum_demand, orders, group_capacity)
    yy, _, _ = assigning_order_groups(groups, num_picking_stations, Demand)
    cost = evaluate_order_groups(groups, yy)
    return cost, round(time.time() - t0, 2)


def config2_full_algorithm():
    t0 = time.time()
    groups = create_order_groups(Sum_demand, orders, group_capacity)
    yy, _, _ = assigning_order_groups(groups, num_picking_stations, Demand)
    best_groups, best_cost = run_vns(
        groups, yy, max_iterations, sequences, num_picking_stations)
    return best_cost, round(time.time() - t0, 2)


def config3_random_vns():
    t0 = time.time()
    groups = create_random_order_groups(orders, group_capacity)
    yy, _, _ = assigning_order_groups(groups, num_picking_stations, Demand)
    best_groups, best_cost = run_vns(
        groups, yy, max_iterations, sequences, num_picking_stations)
    return best_cost, round(time.time() - t0, 2)


# ══════════════════════════════════════════════════════
# Run the ablation study
# ══════════════════════════════════════════════════════

def run_ablation(n_replications=5):
    configs = [
        ('Construction only (Stages 1-3)', config1_construction_only),
        ('Full algorithm (Stages 1-3 + VNS)', config2_full_algorithm),
        ('Random init + VNS', config3_random_vns),
    ]
    results = {name: {'costs': [], 'times': []} for name, _ in configs}

    for rep in range(n_replications):
        seed = SEED + rep
        random.seed(seed)
        np.random.seed(seed)
        print(f"\nReplication {rep+1}/{n_replications}  (seed={seed})")
        print("-" * 65)

        for name, func in configs:
            print(f"  {name}...", end=' ', flush=True)
            cost, t = func()
            results[name]['costs'].append(cost)
            results[name]['times'].append(t)
            print(f"cost={cost:.4f}  time={t}s")

    # -- Results --
    print(f"\n\n{'='*75}")
    print("ABLATION STUDY RESULTS")
    print(f"{'='*75}")
    print(f"{'Configuration':<38} {'Mean':>8} {'Std':>8} {'Time':>8} {'Improv%':>9}")
    print("-" * 75)

    baseline = np.mean(results['Construction only (Stages 1-3)']['costs'])

    for name, _ in configs:
        data = results[name]
        mean_c = np.mean(data['costs'])
        std_c  = np.std(data['costs'])
        mean_t = np.mean(data['times'])
        improv = (baseline - mean_c) / baseline * 100 if baseline > 0 else 0
        print(f"{name:<38} {mean_c:>8.4f} {std_c:>8.4f} {mean_t:>8.2f} {improv:>8.1f}%")

    print("=" * 75)
    print(f"Note: Improv% = improvement over Construction-only baseline")
    print(f"Seeds: {SEED} to {SEED + n_replications - 1}")

    # -- Save to Excel --
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ablation Results"
    ws.append(["Configuration", "Replication", "Seed", "Cost", "Time (s)"])
    for name, _ in configs:
        for rep, (c, t) in enumerate(zip(results[name]['costs'], results[name]['times'])):
            ws.append([name, rep + 1, SEED + rep, round(c, 4), round(t, 2)])

    ws.append([])
    ws.append(["Summary"])
    ws.append(["Configuration", "Mean Cost", "Std Cost", "Mean Time (s)", "Improvement (%)"])
    for name, _ in configs:
        data = results[name]
        mean_c = np.mean(data['costs'])
        std_c  = np.std(data['costs'])
        mean_t = np.mean(data['times'])
        improv = (baseline - mean_c) / baseline * 100 if baseline > 0 else 0
        ws.append([name, round(mean_c,4), round(std_c,4), round(mean_t,2), round(improv,2)])

    wb.save(os.environ.get('ABLATION_OUTPUT_FILE', 'ablation_results.xlsx'))
    print("\n✅ Results saved to ablation_results.xlsx")
    return results


if __name__ == "__main__":
    # Use n_replications=3 if runtime is a concern
    run_ablation(n_replications=5)
