# SA runner v2: updated for the single-horizon RMFS_main module.
"""
Simulated Annealing (SA) for comparison with VNS
=================================================
Place this file in the same folder as:
  - RMFS_main.py
  - Data11.xlsx

SA uses the same neighborhood structures as the VNS, so the comparison
isolates the effect of the search framework rather than the moves.

SA parameters:
  - T0 (initial temperature): estimated automatically (see below)
  - T_min (final temperature): 0.01
  - alpha (cooling rate): 0.95
  - max_iterations: the same kmax used by the main algorithm
"""

import sys
_CPLEX_API = r'C:\Program Files\IBM\ILOG\CPLEX_Studio221\cplex\python\3.7\x64_win64'
# On another machine, set the CPLEX_PYTHON_API environment variable to the
# folder holding the CPLEX Python API, e.g.
#   Linux : /opt/ibm/ILOG/CPLEX_Studio221/cplex/python/3.7/x86-64_linux
#   macOS : /Applications/CPLEX_Studio221/cplex/python/3.7/x86-64_osx
import os
_CPLEX_API = os.environ.get('CPLEX_PYTHON_API', _CPLEX_API)
if os.path.isdir(_CPLEX_API):
    sys.path = [_CPLEX_API] + sys.path
import random
import numpy as np
import copy
import time
import os
import math
import openpyxl

# -- Fixed seed --
SEED = 42
random.seed(SEED)
np.random.seed(SEED)

SA_N_REPLICATIONS = int(os.environ.get('SA_N_REPLICATIONS', '3'))
SA_OUTPUT_FILE = os.environ.get('SA_OUTPUT_FILE', 'vns_vs_sa_results.xlsx')
SA_INSTANCE_LABEL = os.environ.get('SA_INSTANCE_LABEL', os.environ.get('RMFS_DATA_FILE', 'Data11.xlsx'))

# ══════════════════════════════════════════════════════
# Read instance data from Excel
# ══════════════════════════════════════════════════════
DATA_FILE = os.environ.get('RMFS_DATA_FILE', 'Data11.xlsx')
workbook = openpyxl.load_workbook(DATA_FILE)
sheet = workbook.active

def read_matrix(sheet, name):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == name:
                sr, sc = row + 1, col
                rows, cols = 0, 0
                while sheet.cell(row=sr + rows, column=sc).value is not None:
                    rows += 1
                while sheet.cell(row=sr, column=sc + cols).value is not None:
                    cols += 1
                data = []
                for r in range(sr, sr + rows):
                    row_data = [sheet.cell(row=r, column=c).value
                                for c in range(sc, sc + cols)]
                    data.append(row_data)
                return data
    return None

def read_scalar(sheet, name):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == name:
                return sheet.cell(row=row + 1, column=col).value
    return None

Demand               = read_matrix(sheet, 'Demand')
Sum_demand           = np.sum(Demand, axis=1)
orders1              = read_scalar(sheet, 'orders1')
orders               = np.arange(orders1)
num_picking_stations = read_scalar(sheet, 'num_picking_stations')
group_capacity       = read_scalar(sheet, 'group_capacity')
max_iterations       = read_scalar(sheet, 'max_iterations')
sequences            = read_scalar(sheet, 'sequences')

print(f"Instance: {orders1} orders | {num_picking_stations} stations | "
      f"group_cap={group_capacity} | max_iter={max_iterations}")

# ══════════════════════════════════════════════════════
# Helper functions (the same moves as the VNS)
# ══════════════════════════════════════════════════════

def create_order_groups(Sum_demand, orders, group_capacity):
    order_groups = []
    remaining_orders = list(Sum_demand.copy())
    remaining_numbers = list(orders.copy())
    group_count = 0
    order_groups.append([])

    while len(remaining_orders) > 0:
        sorted_pairs = sorted(zip(remaining_orders, remaining_numbers), reverse=True)
        remaining_orders = [d for d, n in sorted_pairs]
        remaining_numbers = [n for d, n in sorted_pairs]
        order_groups[group_count].append(remaining_numbers[0])
        remaining_orders.pop(0)
        remaining_numbers.pop(0)

        if len(remaining_orders) > 0:
            if len(order_groups[group_count]) < group_capacity:
                sorted_pairs = sorted(zip(remaining_orders, remaining_numbers))
                remaining_orders = [d for d, n in sorted_pairs]
                remaining_numbers = [n for d, n in sorted_pairs]
                order_groups[group_count].append(remaining_numbers[0])
                remaining_orders.pop(0)
                remaining_numbers.pop(0)
            else:
                group_count += 1
                order_groups.append([])
                sorted_pairs = sorted(zip(remaining_orders, remaining_numbers))
                remaining_orders = [d for d, n in sorted_pairs]
                remaining_numbers = [n for d, n in sorted_pairs]
                order_groups[group_count].append(remaining_numbers[0])
                remaining_orders.pop(0)
                remaining_numbers.pop(0)

        if len(order_groups[group_count]) >= group_capacity and len(remaining_orders) > 0:
            group_count += 1
            order_groups.append([])

    return order_groups


def assigning_order_groups(order_groups, num_picking_stations, Demand):
    yy = np.zeros((len(order_groups), num_picking_stations, len(Demand)))
    active = np.zeros((len(order_groups), num_picking_stations, len(Demand)))
    ss, jj = 0, 0
    for g in range(len(order_groups)):
        yy[g, ss, jj] = 1
        active[g, ss, jj] = 1
        ss += 1
        if ss >= num_picking_stations:
            jj += 1
            ss = 0
    return yy, active, jj


def find_ss_jj(y, group, sequences, num_picking_stations):
    for ss in range(num_picking_stations):
        for jj in range(sequences):
            if y[group][ss][jj] == 1:
                return ss, jj
    return 0, 0


def get_neighbor(order_groups, yy, sequences, num_picking_stations):
    """
    Pick one of the neighborhoods at random and apply it to a copy of the
    current solution.
    """
    new_groups = copy.deepcopy(order_groups)
    new_yy = yy.copy()

    neighborhood = random.choice([1, 2, 3])

    if neighborhood == 1:
        # Swap entire groups
        if len(new_groups) >= 2:
            g1, g2 = random.sample(range(len(new_groups)), 2)
            new_groups[g1], new_groups[g2] = new_groups[g2], new_groups[g1]

    elif neighborhood == 2:
        # Swap orders between groups
        if len(new_groups) >= 2:
            g1, g2 = random.sample(range(len(new_groups)), 2)
            if new_groups[g1] and new_groups[g2]:
                i1 = random.randint(0, len(new_groups[g1]) - 1)
                i2 = random.randint(0, len(new_groups[g2]) - 1)
                new_groups[g1][i1], new_groups[g2][i2] = \
                    new_groups[g2][i2], new_groups[g1][i1]

    elif neighborhood == 3:
        # Swap station assignments
        if len(new_yy) >= 2:
            gr1, gr2 = random.sample(range(len(new_yy)), 2)
            ss1, jj1 = find_ss_jj(new_yy, gr1, sequences, num_picking_stations)
            ss2, jj2 = find_ss_jj(new_yy, gr2, sequences, num_picking_stations)
            new_yy[gr1][ss1][jj1] = 0
            new_yy[gr2][ss2][jj2] = 0
            new_yy[gr2][ss1][jj1] = 1
            new_yy[gr1][ss2][jj2] = 1

    return new_groups, new_yy


# evaluate_order_groups from the main implementation (single-horizon variant)
from RMFS_main import evaluate_order_groups


# Guard against a silent instance mismatch. evaluate_order_groups and the other
# RMFS_main functions read the demand matrix, capacities and wave data from
# RMFS_main's own module-level worksheet, NOT from the arguments passed in. If
# this script and RMFS_main were pointed at different workbooks, every result
# would be computed from a mixture of two instances and no error would be
# raised. Both now read RMFS_DATA_FILE, so this check should never fire.
import RMFS_main as _rmfs_main
if os.path.abspath(_rmfs_main.DATA_FILE) != os.path.abspath(DATA_FILE):
    raise RuntimeError(
        "Instance mismatch: this script loaded %r but RMFS_main loaded %r. "
        "Set the RMFS_DATA_FILE environment variable so that both use the "
        "same workbook." % (DATA_FILE, _rmfs_main.DATA_FILE))


def estimate_initial_temperature(order_groups, yy, n_samples=10, acceptance_rate=0.8):
    """
    Estimate a suitable initial temperature: target an initial acceptance rate
    of about 80% for worsening solutions, which helps escape local optima.
    """
    costs = []
    current_cost = evaluate_order_groups(order_groups, yy)

    for _ in range(n_samples):
        new_groups, new_yy = get_neighbor(
            order_groups, yy, sequences, num_picking_stations)
        new_cost = evaluate_order_groups(new_groups, new_yy)
        costs.append(abs(new_cost - current_cost))

    avg_delta = np.mean(costs) if costs else 1.0
    # T0 = -delta / ln(acceptance_rate)
    T0 = -avg_delta / math.log(acceptance_rate) if avg_delta > 0 else 1.0
    return max(T0, 0.1)


def simulated_annealing(Sum_demand, orders, group_capacity, max_iterations,
                        num_picking_stations, Demand, sequences,
                        T_min=0.01, alpha=0.95):
    """
    Simulated Annealing using the same neighborhoods as the VNS.

    Parameters:
        alpha: cooling rate (0.9 to 0.99). Smaller = cools faster = less
               exploration; larger = cools slower = more exploration.
        T_min: stopping temperature
    """
    # -- Construction phase (Stages 1-3) --
    order_groups = create_order_groups(Sum_demand, orders, group_capacity)
    yy, _, _ = assigning_order_groups(order_groups, num_picking_stations, Demand)
    current_cost = evaluate_order_groups(order_groups, yy)

    best_groups = copy.deepcopy(order_groups)
    best_yy = yy.copy()
    best_cost = current_cost

    # -- Estimate the initial temperature --
    print("  Estimating initial temperature...", end=' ', flush=True)
    T0 = estimate_initial_temperature(order_groups, yy)
    T = T0
    print(f"T0 = {T0:.4f}")

    # ── SA loop ──
    iteration = 0
    accepted = 0
    rejected = 0

    while iteration < max_iterations and T > T_min:
        # Generate a neighbor
        new_groups, new_yy = get_neighbor(
            order_groups, yy, sequences, num_picking_stations)
        new_cost = evaluate_order_groups(new_groups, new_yy)

        delta = new_cost - current_cost

        # SA acceptance criterion
        if delta < 0:
            # Better solution: always accepted
            order_groups = new_groups
            yy = new_yy
            current_cost = new_cost
            accepted += 1

            if new_cost < best_cost:
                best_groups = copy.deepcopy(new_groups)
                best_yy = new_yy.copy()
                best_cost = new_cost

        else:
            # Worse: accepted with probability exp(-delta/T)
            prob = math.exp(-delta / T) if T > 0 else 0
            if random.random() < prob:
                order_groups = new_groups
                yy = new_yy
                current_cost = new_cost
                accepted += 1
            else:
                rejected += 1

        # Cool the temperature
        T *= alpha
        iteration += 1

    total = accepted + rejected
    acc_rate = accepted / total * 100 if total > 0 else 0
    print(f"  Final T={T:.6f} | Accepted={accepted} ({acc_rate:.1f}%) | "
          f"Best={best_cost:.4f}")

    return best_groups, best_cost


# ══════════════════════════════════════════════════════
# VNS vs SA comparison
# ══════════════════════════════════════════════════════

def run_comparison(n_replications=None):
    """
    Compare VNS and SA on the same instances.
    """
    n_replications = n_replications or SA_N_REPLICATIONS
    print(f"Instance label: {SA_INSTANCE_LABEL}")
    configs = {
        'VNS': {'costs': [], 'times': []},
        'SA':  {'costs': [], 'times': []},
    }

    for rep in range(n_replications):
        seed = SEED + rep
        random.seed(seed)
        np.random.seed(seed)

        print(f"\n{'='*65}")
        print(f"Replication {rep+1}/{n_replications}  (seed={seed})")
        print('='*65)

        # ── VNS ──
        print("Running VNS...")
        t0 = time.time()
        # The main implementation's VNS (returns groups, cost, yy)
        from RMFS_main import variable_neighborhood_search
        _, vns_cost, _ = variable_neighborhood_search(
            Sum_demand, orders, group_capacity, max_iterations,
            num_picking_stations, Demand, sequences)
        vns_time = round(time.time() - t0, 2)
        configs['VNS']['costs'].append(vns_cost)
        configs['VNS']['times'].append(vns_time)
        print(f"  VNS: cost={vns_cost:.4f}  time={vns_time}s")

        # ── SA ──
        print("Running SA...")
        t0 = time.time()
        _, sa_cost = simulated_annealing(
            Sum_demand, orders, group_capacity, max_iterations,
            num_picking_stations, Demand, sequences,
            T_min=0.01, alpha=0.95)
        sa_time = round(time.time() - t0, 2)
        configs['SA']['costs'].append(sa_cost)
        configs['SA']['times'].append(sa_time)
        print(f"  SA:  cost={sa_cost:.4f}  time={sa_time}s")

    # -- Results --
    print(f"\n\n{'='*65}")
    print("VNS vs SA COMPARISON RESULTS")
    print(f"{'='*65}")
    print(f"{'Method':<10} {'Mean(ok)':>10} {'Std(ok)':>10} {'Time(s)':>10} {'Success':>9} {'vs VNS':>10}")
    print("-" * 65)

    FAIL_SENTINEL = 10000000
    stats = {}
    for name, data in configs.items():
        ok_c = [c for c in data['costs'] if c < FAIL_SENTINEL]
        ok_t = [t for c, t in zip(data['costs'], data['times']) if c < FAIL_SENTINEL]
        stats[name] = {
            'mean_c': np.mean(ok_c) if ok_c else None,
            'std_c': np.std(ok_c) if ok_c else None,
            'mean_t': np.mean(ok_t) if ok_t else None,
            'n_ok': len(ok_c),
        }

    vns_mean = stats['VNS']['mean_c']

    for name, data in configs.items():
        s = stats[name]
        mean_str = f"{s['mean_c']:.4f}" if s['mean_c'] is not None else "N/A"
        std_str = f"{s['std_c']:.4f}" if s['std_c'] is not None else "N/A"
        mean_t_str = f"{s['mean_t']:.2f}" if s['mean_t'] is not None else "N/A"
        succ_str = f"{s['n_ok']}/{n_replications}"
        if vns_mean and s['mean_c'] is not None:
            diff = (s['mean_c'] - vns_mean) / vns_mean * 100
            sign = "+" if diff > 0 else ""
            diff_str = f"{sign}{diff:.1f}%"
        else:
            diff_str = "N/A"
        print(f"{name:<10} {mean_str:>10} {std_str:>10} {mean_t_str:>10} {succ_str:>9} {diff_str:>10}")

    print("=" * 65)
    print("'vs VNS' = % difference from VNS mean cost (+ = worse, - = better); means exclude failed runs")

    # -- Save to Excel --
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VNS vs SA"
    ws.append(["Instance", SA_INSTANCE_LABEL])

    ws.append(["Method", "Replication", "Seed", "Cost", "Time (s)", "Failed?"])
    for name, data in configs.items():
        for rep, (c, t) in enumerate(zip(data['costs'], data['times'])):
            ws.append([name, rep + 1, SEED + rep,
                       round(c, 4) if c < FAIL_SENTINEL else "FAILED",
                       round(t, 2), c >= FAIL_SENTINEL])

    ws.append([])
    ws.append(["Summary (means over successful replications only)"])
    ws.append(["Method", "Mean Cost (ok)", "Std Cost (ok)", "Mean Time (s)", "Successes", "Total Reps", "vs VNS (%)"])

    for name, data in configs.items():
        s = stats[name]
        if vns_mean and s['mean_c'] is not None:
            diff = (s['mean_c'] - vns_mean) / vns_mean * 100
        else:
            diff = None
        ws.append([name,
                   round(s['mean_c'], 4) if s['mean_c'] is not None else "N/A",
                   round(s['std_c'], 4) if s['std_c'] is not None else "N/A",
                   round(s['mean_t'], 2) if s['mean_t'] is not None else "N/A",
                   s['n_ok'], n_replications,
                   round(diff, 2) if diff is not None else "N/A"])

    wb.save(SA_OUTPUT_FILE)
    print(f"\n✅ Results saved to {SA_OUTPUT_FILE}")

    return configs


if __name__ == "__main__":
    run_comparison()
