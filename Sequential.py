# Figure 4: Integrated vs. Sequential (random replenishment) comparison.
"""
Figure 4 -- Integrated vs. Sequential Replenishment
======================================================
این فایل رو در همون پوشه‌ای بذار که RMFS_main.py و instance موردنظر هستن.

کاری که می‌کنه:
  - همون گروه‌بندی (Stage 1+2) رو یک‌بار می‌سازه.
  - برای هر replication، هم evaluate_order_groups (روش integrated، که
    replenishment رو خودش بهینه می‌کنه) و هم evaluate_order_groups_sequential
    (روش sequential/random، طبق توضیح Section 5.3 مقاله) رو روی همون
    گروه‌بندی صدا می‌زنه.
  - در پایان: میانگین/std هرکدوم، درصد بهبود integrated نسبت به sequential،
    و یک paired t-test.

⚠️ نکته: چون روش sequential عمداً replenishment رو تصادفی (نه بهینه)
می‌سازه، طبیعیه که گاهی بیشتر از روش integrated fail بشه -- این خودش با
ادعای مقاله («integrated بهتر/پایدارتره») سازگاره؛ نرخ شکست هر دو روش
جدا گزارش می‌شه.

استفاده:
  set RMFS_DATA_FILE=Data_XX.xlsx
  set FIG4_N_REPLICATIONS=5
  set FIG4_OUTPUT_FILE=fig4_XX.xlsx
  python run_figure4.py
"""

import sys
sys.path = [r'C:\Program Files\IBM\ILOG\CPLEX_Studio221\cplex\python\3.7\x64_win64'] + sys.path
import random
import numpy as np
import time
import os
import openpyxl

SEED = 42
random.seed(SEED)
np.random.seed(SEED)

N_REPLICATIONS = int(os.environ.get('FIG4_N_REPLICATIONS', '5'))
OUTPUT_FILE = os.environ.get('FIG4_OUTPUT_FILE', 'fig4_results.xlsx')
INSTANCE_LABEL = os.environ.get('FIG4_INSTANCE_LABEL', os.environ.get('RMFS_DATA_FILE', 'Data11.xlsx'))
FAIL_SENTINEL = 10000000

DATA_FILE = os.environ.get('RMFS_DATA_FILE', 'Data11.xlsx')
workbook = openpyxl.load_workbook(DATA_FILE)
sheet = workbook.active

def read_matrix(sheet, name):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == name:
                sr, sc = row + 1, col
                rows, cols = 0, 0
                while sheet.cell(row=sr + rows, column=sc).value is not None:
                    rows += 1
                while sheet.cell(row=sr, column=sc + cols).value is not None:
                    cols += 1
                data = []
                for r in range(sr, sr + rows):
                    row_data = [sheet.cell(row=r, column=c).value for c in range(sc, sc + cols)]
                    data.append(row_data)
                return data
    return None

def read_scalar(sheet, name):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == name:
                return sheet.cell(row=row + 1, column=col).value
    return None

Demand               = read_matrix(sheet, 'Demand')
Sum_demand           = np.sum(Demand, axis=1)
orders1              = read_scalar(sheet, 'orders1')
orders               = np.arange(orders1)
num_picking_stations = read_scalar(sheet, 'num_picking_stations')
group_capacity       = read_scalar(sheet, 'group_capacity')

from RMFS_main import (
    create_order_groups, assigning_order_groups,
    evaluate_order_groups, evaluate_order_groups_sequential,
)

print(f"Instance label: {INSTANCE_LABEL}")
print(f"Instance: {orders1} orders | {num_picking_stations} picking stations")
print("=" * 65)

fixed_groups = create_order_groups(Sum_demand, orders, group_capacity)
fixed_yy, _, _ = assigning_order_groups(fixed_groups, num_picking_stations, Demand)
print(f"Fixed grouping (Stage 1+2, shared by both schemes): {fixed_groups}")
print("=" * 65)

integrated_costs, sequential_costs = [], []
for rep in range(N_REPLICATIONS):
    seed = SEED + rep
    print(f"\nReplication {rep+1}/{N_REPLICATIONS}  (seed={seed})")
    print("-" * 65)

    random.seed(seed)
    np.random.seed(seed)
    print("  Running integrated (optimized replenishment)...")
    t0 = time.time()
    c_int = evaluate_order_groups(fixed_groups, fixed_yy)
    print(f"  -> integrated objective={c_int:.4f}  time={time.time()-t0:.1f}s"
          + ("  [FAILED]" if c_int >= FAIL_SENTINEL else ""))
    integrated_costs.append(c_int)

    random.seed(seed)
    np.random.seed(seed)
    print("  Running sequential (random replenishment)...")
    t0 = time.time()
    c_seq = evaluate_order_groups_sequential(fixed_groups, fixed_yy)
    print(f"  -> sequential objective={c_seq:.4f}  time={time.time()-t0:.1f}s"
          + ("  [FAILED]" if c_seq >= FAIL_SENTINEL else ""))
    sequential_costs.append(c_seq)

# ── تحلیل: فقط replicationهایی که هر دو موفق بودن، برای paired t-test ──
paired_int, paired_seq = [], []
for ci, cs in zip(integrated_costs, sequential_costs):
    if ci < FAIL_SENTINEL and cs < FAIL_SENTINEL:
        paired_int.append(ci)
        paired_seq.append(cs)

n_int_ok = sum(1 for c in integrated_costs if c < FAIL_SENTINEL)
n_seq_ok = sum(1 for c in sequential_costs if c < FAIL_SENTINEL)

print(f"\n\n{'='*65}")
print(f"FIGURE 4 RESULT -- {INSTANCE_LABEL}")
print(f"{'='*65}")
print(f"Integrated successes: {n_int_ok}/{N_REPLICATIONS}")
print(f"Sequential successes: {n_seq_ok}/{N_REPLICATIONS}")

if paired_int:
    mean_int = np.mean(paired_int)
    mean_seq = np.mean(paired_seq)
    improv = (mean_seq - mean_int) / mean_seq * 100 if mean_seq else float('nan')
    print(f"Mean integrated objective (paired): {mean_int:.4f}")
    print(f"Mean sequential objective (paired): {mean_seq:.4f}")
    print(f"Improvement of integrated over sequential: {improv:.2f}%")

    diffs = [s - i for s, i in zip(paired_seq, paired_int)]
    n = len(diffs)
    mean_d = np.mean(diffs)
    std_d = np.std(diffs, ddof=1) if n > 1 else 0.0
    t_stat = (float('inf') if mean_d != 0 else 0.0) if std_d == 0 else mean_d / (std_d / np.sqrt(n))
    dof = n - 1
    try:
        from scipy import stats as _stats
        p_value = 2 * (1 - _stats.t.cdf(abs(t_stat), dof)) if dof > 0 else float('nan')
        print(f"Paired t-test (n={n}, dof={dof}): t={t_stat:.3f}, p={p_value:.4f} [scipy exact]")
    except ImportError:
        print(f"Paired t-test (n={n}, dof={dof}): t={t_stat:.3f} "
              f"(install scipy for exact p-value)")
else:
    print("⚠️  No replication had both schemes succeed -- no usable comparison.")

print("=" * 65)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Figure 4 Data"
ws.append(["Instance", INSTANCE_LABEL])
ws.append(["Replication", "Seed", "Integrated Obj.", "Integrated Failed?",
           "Sequential Obj.", "Sequential Failed?"])
for rep, (ci, cs) in enumerate(zip(integrated_costs, sequential_costs)):
    ws.append([rep + 1, SEED + rep,
               round(ci, 4) if ci < FAIL_SENTINEL else "FAILED", ci >= FAIL_SENTINEL,
               round(cs, 4) if cs < FAIL_SENTINEL else "FAILED", cs >= FAIL_SENTINEL])

ws.append([])
ws.append(["Summary"])
ws.append(["Scheme", "Mean (paired ok)", "Successes", "Total Reps"])
ws.append(["Integrated", round(np.mean(paired_int), 4) if paired_int else "N/A", n_int_ok, N_REPLICATIONS])
ws.append(["Sequential", round(np.mean(paired_seq), 4) if paired_seq else "N/A", n_seq_ok, N_REPLICATIONS])

if paired_int:
    ws.append([])
    ws.append(["Improvement of integrated vs sequential (%)", round(improv, 2)])
    ws.append(["Paired t-test: t-statistic", "dof", "p-value"])
    try:
        ws.append([round(t_stat, 4) if t_stat != float('inf') else "inf", dof, round(p_value, 4)])
    except NameError:
        ws.append([round(t_stat, 4) if t_stat != float('inf') else "inf", dof, "N/A (install scipy)"])

wb.save(OUTPUT_FILE)
print(f"\n✅ Results saved to {OUTPUT_FILE}")
