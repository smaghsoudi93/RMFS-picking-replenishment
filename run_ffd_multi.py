# Multi-instance FFD comparison orchestrator.
"""
FFD Comparison Across Multiple Instances
==========================================
این فایل رو در همون پوشه‌ای بذار که:
  - Main_Thesis_Final_version_DEBUG.py
  - ffd_comparison.py
  - همه‌ی فایل‌های instance (Data*.xlsx) که می‌خوای تست کنی
هستن.

⚠️ حتماً اول لیست INSTANCE_FILES پایین رو با اسم واقعی فایل‌های
instance خودت پر کن (همون‌هایی که برای Table 3 استفاده کردی).

کاری که می‌کنه:
  - برای هر instance، ffd_comparison.py رو به‌صورت یک subprocess کاملاً
    مستقل اجرا می‌کنه (RMFS_DATA_FILE رو روی همون فایل ست می‌کنه). این
    کار عمداً subprocess-based هست، نه import مستقیم توی همون پردازش،
    چون Main_Thesis_Final_version_DEBUG فقط یک‌بار در طول عمر پردازش
    Excel رو می‌خونه -- برای instance بعدی باید یک پردازش پایتون کاملاً
    تازه باشه.
  - نتیجه‌ی هر instance (فایل ffd_comparison_<label>.xlsx) رو می‌خونه.
  - یک جدول خلاصه‌ی across-instance می‌سازه: میانگین هر روش روی هر
    instance، به‌علاوه‌ی یک paired t-test نهایی که «میانگین هر instance»
    (نه تک‌تک replicationها) رو بین Modified FFD و Standard FFD مقایسه
    می‌کنه -- این همون سطح مقایسه‌ایه که واقعاً به سؤال «آیا FFD
    اصلاح‌شده به‌طور کلی بهتره؟» جواب می‌ده.

خروجی:
  - ffd_comparison_<label>.xlsx برای هر instance (توسط ffd_comparison.py)
  - ffd_multi_instance_summary.xlsx (خلاصه‌ی نهایی across-instance)
"""

import os
import sys
import subprocess
import numpy as np
import openpyxl

# ══════════════════════════════════════════════════════
# ⚠️ این لیست رو با اسم واقعی فایل‌های instance خودت پر کن
# (همون‌هایی که برای Table 3 استفاده کردی: 6, 8, 15, 20, 30, ... orders)
# ══════════════════════════════════════════════════════
INSTANCE_FILES = [
    "Data3.xlsx",
    "Data4.xlsx",
    "Data5.xlsx",
    "Data6.xlsx",
]

N_REPLICATIONS_PER_INSTANCE = 3  # کاهش داده شد برای صرفه‌جویی در وقت

FFD_COMPARISON_SCRIPT = "run_ffd_single.py"


def run_one_instance(instance_file):
    label = os.path.splitext(os.path.basename(instance_file))[0]
    output_file = f"ffd_comparison_{label}.xlsx"

    if not os.path.exists(instance_file):
        print(f"⚠️  SKIPPING {instance_file} -- file not found in this folder.")
        return None

    env = os.environ.copy()
    env["RMFS_DATA_FILE"] = instance_file
    env["FFD_N_REPLICATIONS"] = str(N_REPLICATIONS_PER_INSTANCE)
    env["FFD_OUTPUT_FILE"] = output_file
    env["FFD_INSTANCE_LABEL"] = label

    print(f"\n{'#'*70}")
    print(f"# Running FFD comparison for instance: {instance_file}")
    print(f"{'#'*70}")

    result = subprocess.run(
        [sys.executable, FFD_COMPARISON_SCRIPT],
        env=env,
        capture_output=False,  # let it print live to the console
    )
    if result.returncode != 0:
        print(f"⚠️  {instance_file} failed (exit code {result.returncode}); skipping.")
        return None

    if not os.path.exists(output_file):
        print(f"⚠️  Expected output {output_file} not found; skipping.")
        return None

    return output_file


def read_instance_result(output_file):
    """Extract per-method mean costs (successes only) + success counts."""
    wb = openpyxl.load_workbook(output_file, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # find the "Summary (...)" marker row, then read the method rows after it
    summary_start = None
    for i, row in enumerate(rows):
        if row and row[0] and str(row[0]).startswith("Summary"):
            summary_start = i
            break
    if summary_start is None:
        return None

    methods = {}
    for row in rows[summary_start + 2:]:
        if not row or row[0] is None:
            break
        # New layout: Method, Mean Cost(ok), Std Cost(ok), Mean Time,
        # Successes, Total Reps, Improvement vs Random (%)
        name, mean_c, std_c, mean_t, n_ok, n_total, improv = row[:7]
        mean_c = float(mean_c) if isinstance(mean_c, (int, float)) else None
        std_c = float(std_c) if isinstance(std_c, (int, float)) else None
        methods[name] = {"mean_cost": mean_c, "std_cost": std_c,
                          "mean_time": mean_t, "improv_vs_random": improv,
                          "n_ok": n_ok, "n_total": n_total}
    return methods


def paired_t_test(sample_a, sample_b):
    diffs = [a - b for a, b in zip(sample_a, sample_b)]
    n = len(diffs)
    mean_d = np.mean(diffs)
    std_d = np.std(diffs, ddof=1) if n > 1 else 0.0
    if std_d == 0:
        t_stat = float('inf') if mean_d != 0 else 0.0
    else:
        t_stat = mean_d / (std_d / np.sqrt(n))
    dof = n - 1
    try:
        from scipy import stats as _stats
        p_value = 2 * (1 - _stats.t.cdf(abs(t_stat), dof)) if dof > 0 else float('nan')
        return t_stat, dof, p_value, "scipy (exact)"
    except ImportError:
        crit_05 = {1: 12.706, 2: 4.303, 3: 3.182, 4: 2.776, 5: 2.571,
                   6: 2.447, 7: 2.365, 8: 2.306, 9: 2.262, 10: 2.228}
        closest_dof = min(crit_05.keys(), key=lambda k: abs(k - dof)) if dof > 0 else 1
        crit = crit_05[closest_dof]
        verdict = "p < 0.05 (approx.)" if abs(t_stat) > crit else "p >= 0.05 (approx.)"
        return t_stat, dof, verdict, "manual approx. (install scipy for exact p-value)"


def main():
    per_instance_results = {}  # label -> methods dict

    for instance_file in INSTANCE_FILES:
        output_file = run_one_instance(instance_file)
        if output_file is None:
            continue
        label = os.path.splitext(os.path.basename(instance_file))[0]
        methods = read_instance_result(output_file)
        if methods:
            per_instance_results[label] = methods

    if not per_instance_results:
        print("\nNo instance results collected -- check INSTANCE_FILES and that "
              "the files exist in this folder.")
        return

    # ── Cross-instance summary ──
    print(f"\n\n{'='*90}")
    print("CROSS-INSTANCE SUMMARY (mean cost [successes/total] per method, per instance)")
    print(f"{'='*90}")
    method_names = ['Modified FFD (proposed)', 'Standard FFD', 'Random grouping']
    header = f"{'Instance':<12}" + "".join(f"{m:<26}" for m in method_names)
    print(header)
    print("-" * 90)
    for label, methods in per_instance_results.items():
        row = f"{label:<12}"
        for m in method_names:
            info = methods.get(m, {})
            mc = info.get('mean_cost')
            n_ok = info.get('n_ok', '?')
            n_tot = info.get('n_total', '?')
            cell = f"{mc:.4f} [{n_ok}/{n_tot}]" if mc is not None else f"N/A [{n_ok}/{n_tot}]"
            row += f"{cell:<26}"
        print(row)

    # Only use instances where BOTH methods have a real (non-None) mean
    # for the instance-level comparison -- an instance where one method
    # failed entirely contributes no information to "is Modified better
    # than Standard", so it must be excluded rather than compared as 0/None.
    usable_labels = [
        l for l in per_instance_results
        if per_instance_results[l].get('Modified FFD (proposed)', {}).get('mean_cost') is not None
        and per_instance_results[l].get('Standard FFD', {}).get('mean_cost') is not None
    ]
    modified_means = [per_instance_results[l]['Modified FFD (proposed)']['mean_cost'] for l in usable_labels]
    standard_means = [per_instance_results[l]['Standard FFD']['mean_cost'] for l in usable_labels]

    print("-" * 90)
    print(f"Instances usable for Modified-vs-Standard comparison "
          f"(both had >=1 successful replication): {usable_labels}")
    if len(modified_means) >= 2:
        t_stat, dof, p_or_verdict, method_used = paired_t_test(modified_means, standard_means)
        overall_improv = (np.mean(standard_means) - np.mean(modified_means)) / np.mean(standard_means) * 100
        print(f"Instance-level paired t-test (Modified FFD vs Standard FFD), "
              f"n={len(modified_means)} instances, dof={dof}:")
        print(f"  t={t_stat:.3f}, p={p_or_verdict}  [{method_used}]")
        print(f"  Overall mean improvement of Modified FFD vs Standard FFD: {overall_improv:.2f}%")
    else:
        t_stat, dof, p_or_verdict, method_used, overall_improv = (
            float('nan'), 0, f"N/A -- only {len(modified_means)} usable instance(s)", "n/a", float('nan'))
        print(f"⚠️  Instance-level t-test NOT computed: only {len(modified_means)} instance(s) "
              f"had both methods succeed at least once. Need >= 2 usable instances.")

    # ── Save summary workbook ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cross-Instance Summary"
    ws.append(["Instance"] + sum([[f"{m} - Mean", f"{m} - Successes/Total"] for m in method_names], []))
    for label, methods in per_instance_results.items():
        row = [label]
        for m in method_names:
            info = methods.get(m, {})
            row.append(info.get('mean_cost'))
            row.append(f"{info.get('n_ok', '?')}/{info.get('n_total', '?')}")
        ws.append(row)

    ws.append([])
    ws.append([f"Usable instances for t-test (both methods succeeded >=1x): {usable_labels}"])
    if len(modified_means) >= 2:
        ws.append(["Instance-level paired t-test: Modified FFD vs Standard FFD"])
        ws.append(["t-statistic", "dof", "p-value / verdict", "method", "Overall improvement (%)", "n instances"])
        ws.append([round(t_stat, 4) if t_stat == t_stat and t_stat != float('inf') else str(t_stat),
                   dof, str(p_or_verdict), method_used, round(overall_improv, 2), len(modified_means)])

    wb.save("ffd_multi_instance_summary.xlsx")
    print("\n✅ Cross-instance summary saved to ffd_multi_instance_summary.xlsx")
    print("   (per-instance detail files: ffd_comparison_<label>.xlsx)")


if __name__ == "__main__":
    main()
