# FFD comparison runner: modified FFD vs. standard FFD vs. random grouping,
# evaluated BEFORE any VNS refinement (Stage 1-3 only). This isolates the
# effect of the grouping heuristic itself, independent of VNS's ability to
# later erase a poor starting point (as seen in the ablation study).
"""
FFD Grouping Comparison (pre-VNS)
==================================
Place this file in the same folder as:
  - RMFS_main.py
  - Data11.xlsx

Three grouping strategies are compared, each WITHOUT VNS (the raw Stage 1-3
output):
  1. Modified FFD (alternating decreasing/increasing) -- the proposed method
  2. Standard FFD (decreasing order only)
  3. Random grouping

Output:
  - console log
  - ffd_comparison_results.xlsx

Note: the numbers are reported exactly as produced, whichever direction they
point. This is what Table 4 of the manuscript reports.
"""

import sys
_CPLEX_API = r'C:\Program Files\IBM\ILOG\CPLEX_Studio221\cplex\python\3.7\x64_win64'
# On another machine, set the CPLEX_PYTHON_API environment variable to the
# folder holding the CPLEX Python API, e.g.
#   Linux : /opt/ibm/ILOG/CPLEX_Studio221/cplex/python/3.7/x86-64_linux
#   macOS : /Applications/CPLEX_Studio221/cplex/python/3.7/x86-64_osx
import os
_CPLEX_API = os.environ.get('CPLEX_PYTHON_API', _CPLEX_API)
if os.path.isdir(_CPLEX_API):
    sys.path = [_CPLEX_API] + sys.path
import random
import numpy as np
import time
import os
import openpyxl

SEED = 42
random.seed(SEED)
np.random.seed(SEED)

# Configurable via env vars so a multi-instance orchestrator (see
# run_ffd_multi.py) can control each subprocess run without editing
# this file each time.
N_REPLICATIONS = int(os.environ.get('FFD_N_REPLICATIONS', '10'))
OUTPUT_FILE = os.environ.get('FFD_OUTPUT_FILE', 'ffd_comparison_results.xlsx')
INSTANCE_LABEL = os.environ.get('FFD_INSTANCE_LABEL', os.environ.get('RMFS_DATA_FILE', 'Data11.xlsx'))

DATA_FILE = os.environ.get('RMFS_DATA_FILE', 'Data11.xlsx')
workbook = openpyxl.load_workbook(DATA_FILE)
sheet = workbook.active

def read_matrix(sheet, name):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == name:
                sr, sc = row + 1, col
                rows, cols = 0, 0
                while sheet.cell(row=sr + rows, column=sc).value is not None:
                    rows += 1
                while sheet.cell(row=sr, column=sc + cols).value is not None:
                    cols += 1
                data = []
                for r in range(sr, sr + rows):
                    row_data = [sheet.cell(row=r, column=c).value
                                for c in range(sc, sc + cols)]
                    data.append(row_data)
                return data
    return None

def read_scalar(sheet, name):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == name:
                return sheet.cell(row=row + 1, column=col).value
    return None

Demand               = read_matrix(sheet, 'Demand')
Sum_demand           = np.sum(Demand, axis=1)
orders1              = read_scalar(sheet, 'orders1')
orders               = np.arange(orders1)
num_picking_stations = read_scalar(sheet, 'num_picking_stations')
group_capacity       = read_scalar(sheet, 'group_capacity')
sequences            = read_scalar(sheet, 'sequences')

print(f"Instance: {orders1} orders | {num_picking_stations} stations | "
      f"group_cap={group_capacity}")
print("=" * 65)

# ══════════════════════════════════════════════════════
# The three grouping strategies
# ══════════════════════════════════════════════════════

def create_order_groups_modified_ffd(Sum_demand, orders, group_capacity):
    """Modified FFD: alternating decreasing/increasing (Algorithm 1)"""
    order_groups = []
    remaining_orders = list(Sum_demand.copy())
    remaining_numbers = list(orders.copy())
    group_count = 0
    order_groups.append([])

    while len(remaining_orders) > 0:
        sorted_pairs = sorted(zip(remaining_orders, remaining_numbers), reverse=True)
        remaining_orders = [d for d, n in sorted_pairs]
        remaining_numbers = [n for d, n in sorted_pairs]
        order_groups[group_count].append(remaining_numbers[0])
        remaining_orders.pop(0)
        remaining_numbers.pop(0)

        if len(remaining_orders) > 0:
            if len(order_groups[group_count]) < group_capacity:
                sorted_pairs = sorted(zip(remaining_orders, remaining_numbers))
                remaining_orders = [d for d, n in sorted_pairs]
                remaining_numbers = [n for d, n in sorted_pairs]
                order_groups[group_count].append(remaining_numbers[0])
                remaining_orders.pop(0)
                remaining_numbers.pop(0)
            else:
                group_count += 1
                order_groups.append([])
                sorted_pairs = sorted(zip(remaining_orders, remaining_numbers))
                remaining_orders = [d for d, n in sorted_pairs]
                remaining_numbers = [n for d, n in sorted_pairs]
                order_groups[group_count].append(remaining_numbers[0])
                remaining_orders.pop(0)
                remaining_numbers.pop(0)

        if len(order_groups[group_count]) >= group_capacity and len(remaining_orders) > 0:
            group_count += 1
            order_groups.append([])

    return order_groups


def create_order_groups_standard_ffd(Sum_demand, orders, group_capacity):
    """Standard FFD: always decreasing, first-fit into groups with room."""
    pairs = sorted(zip(Sum_demand, orders), reverse=True)
    order_groups = []
    for demand, order in pairs:
        placed = False
        for g in order_groups:
            if len(g) < group_capacity:
                g.append(order)
                placed = True
                break
        if not placed:
            order_groups.append([order])
    return order_groups


def create_random_order_groups(orders, group_capacity):
    """Random grouping (no demand-based logic at all)."""
    shuffled = list(orders)
    random.shuffle(shuffled)
    groups = []
    for i in range(0, len(shuffled), group_capacity):
        chunk = shuffled[i:i + group_capacity]
        if chunk:
            groups.append(chunk)
    return groups


def assigning_order_groups(order_groups, num_picking_stations, Demand):
    """Stage 2: Station assignment (unchanged, identical for all 3 methods)."""
    yy = np.zeros((len(order_groups), num_picking_stations, len(Demand)))
    ss, jj = 0, 0
    for g in range(len(order_groups)):
        yy[g, ss, jj] = 1
        ss += 1
        if ss >= num_picking_stations:
            jj += 1
            ss = 0
    return yy


# evaluate_order_groups from the main implementation. Only Stages 1-3 are
# exercised: this function is called directly and variable_neighborhood_search
# is deliberately never invoked.
from RMFS_main import evaluate_order_groups


# Guard against a silent instance mismatch. evaluate_order_groups and the other
# RMFS_main functions read the demand matrix, capacities and wave data from
# RMFS_main's own module-level worksheet, NOT from the arguments passed in. If
# this script and RMFS_main were pointed at different workbooks, every result
# would be computed from a mixture of two instances and no error would be
# raised. Both now read RMFS_DATA_FILE, so this check should never fire.
import RMFS_main as _rmfs_main
if os.path.abspath(_rmfs_main.DATA_FILE) != os.path.abspath(DATA_FILE):
    raise RuntimeError(
        "Instance mismatch: this script loaded %r but RMFS_main loaded %r. "
        "Set the RMFS_DATA_FILE environment variable so that both use the "
        "same workbook." % (DATA_FILE, _rmfs_main.DATA_FILE))


def run_one(method_name, group_fn):
    t0 = time.time()
    if method_name == 'Random grouping':
        groups = group_fn(orders, group_capacity)
    else:
        groups = group_fn(Sum_demand, orders, group_capacity)
    yy = assigning_order_groups(groups, num_picking_stations, Demand)
    cost = evaluate_order_groups(groups, yy)
    return cost, round(time.time() - t0, 2)


def paired_t_test(sample_a, sample_b):
    """
    Paired t-test without requiring scipy (falls back to scipy.stats if
    available, for an exact p-value; otherwise reports the t-statistic
    and an approximate p-value from a small built-in t-table).
    """
    diffs = [a - b for a, b in zip(sample_a, sample_b)]
    n = len(diffs)
    mean_d = np.mean(diffs)
    std_d = np.std(diffs, ddof=1) if n > 1 else 0.0
    if std_d == 0:
        t_stat = float('inf') if mean_d != 0 else 0.0
    else:
        t_stat = mean_d / (std_d / np.sqrt(n))
    dof = n - 1

    try:
        from scipy import stats as _stats
        p_value = 2 * (1 - _stats.t.cdf(abs(t_stat), dof)) if dof > 0 else float('nan')
        return t_stat, dof, p_value, "scipy (exact)"
    except ImportError:
        # Rough two-sided critical values for common dof at alpha=0.05,
        # just to give a directional read without scipy installed.
        crit_05 = {1: 12.706, 2: 4.303, 3: 3.182, 4: 2.776, 5: 2.571,
                   6: 2.447, 7: 2.365, 8: 2.306, 9: 2.262, 10: 2.228,
                   15: 2.131, 20: 2.086, 30: 2.042}
        closest_dof = min(crit_05.keys(), key=lambda k: abs(k - dof)) if dof > 0 else 1
        crit = crit_05[closest_dof]
        verdict = "p < 0.05 (approx.)" if abs(t_stat) > crit else "p >= 0.05 (approx.)"
        return t_stat, dof, verdict, "manual approx. (install scipy for exact p-value)"


FAIL_SENTINEL = 10000000

def split_success_fail(costs, times):
    succ_costs, succ_times, n_fail = [], [], 0
    for c, t in zip(costs, times):
        if c >= FAIL_SENTINEL:
            n_fail += 1
        else:
            succ_costs.append(c)
            succ_times.append(t)
    return succ_costs, succ_times, n_fail


def run_ffd_comparison(n_replications=None):
    n_replications = n_replications or N_REPLICATIONS
    methods = [
        ('Modified FFD (proposed)', create_order_groups_modified_ffd),
        ('Standard FFD', create_order_groups_standard_ffd),
        ('Random grouping', create_random_order_groups),
    ]
    results = {name: {'costs': [], 'times': []} for name, _ in methods}

    print(f"Instance label: {INSTANCE_LABEL}")
    for rep in range(n_replications):
        seed = SEED + rep
        random.seed(seed)
        np.random.seed(seed)
        print(f"\nReplication {rep+1}/{n_replications}  (seed={seed})")
        print("-" * 65)
        for name, fn in methods:
            print(f"  {name}...", end=' ', flush=True)
            cost, t = run_one(name, fn)
            results[name]['costs'].append(cost)
            results[name]['times'].append(t)
            print(f"cost={cost:.4f}  time={t}s"
                  + ("  [FAILED]" if cost >= FAIL_SENTINEL else ""))

    print(f"\n\n{'='*70}")
    print(f"FFD GROUPING COMPARISON RESULTS (pre-VNS objective) -- {INSTANCE_LABEL}")
    print(f"{'='*70}")
    print(f"{'Method':<28} {'Mean(ok)':>9} {'Std(ok)':>9} {'Success':>9} {'vs Random %':>12}")
    print("-" * 70)

    stats_by_method = {}
    for name, _ in methods:
        data = results[name]
        succ_costs, succ_times, n_fail = split_success_fail(data['costs'], data['times'])
        n_ok = len(succ_costs)
        mean_c = np.mean(succ_costs) if n_ok > 0 else float('nan')
        std_c = np.std(succ_costs) if n_ok > 0 else float('nan')
        mean_t = np.mean(succ_times) if n_ok > 0 else float('nan')
        stats_by_method[name] = {
            'mean_c': mean_c, 'std_c': std_c, 'mean_t': mean_t,
            'n_ok': n_ok, 'n_fail': n_fail,
        }

    random_mean = stats_by_method['Random grouping']['mean_c']
    for name, _ in methods:
        s = stats_by_method[name]
        improv = ((random_mean - s['mean_c']) / random_mean * 100
                   if (random_mean and random_mean == random_mean and s['mean_c'] == s['mean_c']) else float('nan'))
        succ_str = f"{s['n_ok']}/{n_replications}"
        mean_str = f"{s['mean_c']:.4f}" if s['mean_c'] == s['mean_c'] else "N/A"
        std_str = f"{s['std_c']:.4f}" if s['std_c'] == s['std_c'] else "N/A"
        improv_str = f"{improv:.1f}%" if improv == improv else "N/A"
        print(f"{name:<28} {mean_str:>9} {std_str:>9} {succ_str:>9} {improv_str:>12}")

    print("-" * 70)
    if stats_by_method['Modified FFD (proposed)']['n_ok'] < n_replications or \
       stats_by_method['Standard FFD']['n_ok'] < n_replications:
        print(f"⚠️  Some replications failed (infeasible/timeout) and were EXCLUDED "
              f"from the means above -- see 'Success' column. Report both the "
              f"mean-of-successes AND the success rate honestly; do not treat "
              f"the mean as if it came from a full, clean sample.")

    # Paired t-test only over replications where BOTH methods succeeded.
    paired_mod, paired_std = [], []
    for c_mod, c_std in zip(results['Modified FFD (proposed)']['costs'],
                             results['Standard FFD']['costs']):
        if c_mod < FAIL_SENTINEL and c_std < FAIL_SENTINEL:
            paired_mod.append(c_mod)
            paired_std.append(c_std)

    if len(paired_mod) >= 2:
        t_stat, dof, p_or_verdict, method_used = paired_t_test(paired_mod, paired_std)
        print(f"Paired t-test (Modified FFD vs Standard FFD), using {len(paired_mod)} "
              f"replications where BOTH succeeded (out of {n_replications} total), "
              f"dof={dof}: t={t_stat:.3f}, p={p_or_verdict}  [{method_used}]")
    else:
        t_stat, dof, p_or_verdict, method_used = (float('nan'), 0,
            f"N/A -- only {len(paired_mod)} replication(s) had both methods succeed", "n/a")
        print(f"⚠️  Paired t-test NOT computed: only {len(paired_mod)} replication(s) had "
              f"both Modified FFD and Standard FFD succeed. Need >= 2 for a t-test.")
    print("=" * 70)
    print("Report these numbers as-is, whichever direction they point.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FFD Comparison"
    ws.append(["Instance", INSTANCE_LABEL])
    ws.append(["Method", "Replication", "Seed", "Cost", "Time (s)", "Failed?"])
    for name, _ in methods:
        for rep, (c, t) in enumerate(zip(results[name]['costs'], results[name]['times'])):
            ws.append([name, rep + 1, SEED + rep, round(c, 4) if c < FAIL_SENTINEL else "FAILED",
                       round(t, 2), c >= FAIL_SENTINEL])

    ws.append([])
    ws.append(["Summary (means computed over SUCCESSFUL replications only)"])
    ws.append(["Method", "Mean Cost (ok only)", "Std Cost (ok only)", "Mean Time (s)",
               "Successes", "Total Reps", "Improvement vs Random (%)"])
    for name, _ in methods:
        s = stats_by_method[name]
        improv = ((random_mean - s['mean_c']) / random_mean * 100
                   if (random_mean == random_mean and s['mean_c'] == s['mean_c']) else None)
        ws.append([name,
                   round(s['mean_c'], 4) if s['mean_c'] == s['mean_c'] else "N/A",
                   round(s['std_c'], 4) if s['std_c'] == s['std_c'] else "N/A",
                   round(s['mean_t'], 2) if s['mean_t'] == s['mean_t'] else "N/A",
                   s['n_ok'], n_replications,
                   round(improv, 2) if improv is not None else "N/A"])

    ws.append([])
    ws.append(["Paired t-test: Modified FFD vs Standard FFD (successful-pairs only)"])
    ws.append(["t-statistic", "dof", "p-value / verdict", "method", "n paired reps used"])
    ws.append([round(t_stat, 4) if t_stat == t_stat and t_stat != float('inf') else str(t_stat),
               dof, str(p_or_verdict), method_used, len(paired_mod)])

    wb.save(OUTPUT_FILE)
    print(f"\n✅ Results saved to {OUTPUT_FILE}")
    return results


if __name__ == "__main__":
    run_ffd_comparison()
